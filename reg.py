import streamlit as st
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime, time, timedelta
from zoneinfo import ZoneInfo

BAHRAIN_TZ = ZoneInfo('Asia/Bahrain')
def now_bh(): return datetime.now(BAHRAIN_TZ)
from streamlit_geolocation import streamlit_geolocation
import math
import random
import string
import time as time_module
import urllib.parse
from io import BytesIO
import pandas as pd

try:
    from streamlit_local_storage import LocalStorage
    localS = LocalStorage()
    LOCAL_STORAGE_OK = True
except Exception:
    localS = None
    LOCAL_STORAGE_OK = False

st.set_page_config(page_title="نظام الحضور والانصراف", page_icon="🕘", layout="centered")
st.markdown("""
<style>
.block-container {
    padding-top: 0.9rem !important;
}

section.main > div {
    padding-top: 0.6rem !important;
}

html, body {
    margin: 0 !important;
    padding: 0 !important;
}

[data-testid="stImage"] {
    margin-top: 12px !important;
}
</style>
""", unsafe_allow_html=True)
# ─── إعدادات ───────────────────────────────────────────────────
SHEET_ID       = "1svkfgRq4-osKr86_2WJQFZShuoy8Ek5DOiUaaHKL-6Y"
SCHOOL_LAT     = 26.216371784473964
SCHOOL_LON     = 50.54035843289093
ALLOWED_RADIUS = 1000
ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "Afaf1234")
DEVICE_COOLDOWN_MINUTES = 10  # لم يعد مستخدماً كمهلة أساسية؛ القفل الآن طوال اليوم إلا إذا عطله الأدمن
DEVICE_LOCK_STRICT_FULL_DAY = True
APP_URL = "https://attendance-jsgs.streamlit.app/"

# أعمدة sheet1
COL_DATE=1; COL_DAY=2; COL_SCHOOL=3; COL_TASK=4; COL_SUPPORT=5
COL_NAME=6; COL_ID=7; COL_ATTEND=8; COL_LATE_REASON=9
COL_DEPART=10; COL_DEPART_REASON=11; COL_EXIT=12; COL_RETURN=13; COL_ATTEMPT=14
COL_WORK_START=15; COL_EXPECTED_END=16; COL_WORK_HOURS=17; COL_EXTRA_HOURS=18
COL_WORK_STATUS=19; COL_DAILY_TYPE=20; COL_AUTO_CLOSE=21; COL_CARE_CONF=22; COL_REG_TYPE=23

schools = [
    "مدرسة المنامة الثانوية للبنات",
    "مدرسة النور الثانوية للبنات",
    "مدرسة المعرفة الثانوية للبنات",
    "مدرسة الرفاع الغربي الثانوية للبنات",
    "مدرسة جدحفص الثانوية للبنات"
]

TASKS_MAIN = [
    "مصححة — اللغة العربية","مصححة — اللغة الإنجليزية","مصححة — الرياضيات",
    "مصححة — الفيزياء","مصححة — الكيمياء","مصححة — الأحياء",
    "مصححة — العلوم التجارية","مصححة — المواد الاجتماعية","مصححة — التربية الإسلامية",
    "مصححة — التربية الأسرية","مصححة — التربية الفنية","مصححة — الحاسب الآلي",
    "مصححة — التربية البدنية","كنترول خارجي — دعم فني",
    "كنترول خارجي — رصد الدرجات","كنترول خارجي — ضبط مركزي",
]
TASKS_SUPPORT = [
    "دعم — اللغة العربية","دعم — اللغة الإنجليزية","دعم — الرياضيات",
    "دعم — الفيزياء","دعم — الكيمياء","دعم — الأحياء",
    "دعم — العلوم التجارية","دعم — المواد الاجتماعية","دعم — التربية الإسلامية",
    "دعم — التربية الأسرية","دعم — التربية الفنية","دعم — الحاسب الآلي",
    "دعم — التربية البدنية","دعم — كنترول فني","دعم — رصد الدرجات","دعم — ضبط مركزي",
]
TASKS_ALL = TASKS_MAIN + TASKS_SUPPORT
JOB_TITLES    = ["منسقة","معلمة أولى","معلمة","الهيئة الإدارية","مشرف تربوي","مديرة مدرسة","المديرة المساعدة","أخرى"]
reasons       = ["دوام مرن","موعد","مهمة رسمية","رعاية","الانتهاء من التصحيح","أخرى"]
abs_reasons   = ["مرض","إجازة اعتيادية","إجازة طارئة","بدون عذر","مهمة رسمية","أخرى"]

# ─── CSS ───────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700;900&display=swap');
html,body,[class*="css"]{direction:rtl!important;text-align:right!important;font-family:'Cairo',Tahoma,sans-serif!important;}
.stApp{direction:rtl!important;}
.block-container{max-width:680px;padding-top:0px;padding-bottom:40px;}
[data-testid="stVerticalBlock"],[data-testid="stHorizontalBlock"],.element-container,.stMarkdown,.stTextInput,.stSelectbox,.stRadio,.stButton,.stAlert{direction:rtl!important;text-align:right!important;}
.stTextInput input,.stSelectbox div{direction:rtl!important;text-align:right!important;}
label{direction:rtl!important;text-align:right!important;display:block!important;font-size:15px!important;font-weight:700!important;color:#0c3460!important;}
.app-header{background:linear-gradient(135deg,#0c3460 0%,#1a5276 60%,#1f6fa3 100%);border-radius:0 0 28px 28px;padding:28px 24px 32px;text-align:center!important;margin:-1rem -1rem 20px -1rem;}
.app-header .sub{color:rgba(255,255,255,0.78);font-size:12px;font-weight:600;margin-bottom:6px;}
.app-header .title{color:#fff;font-size:22px;font-weight:900;}
.app-header .date-pill{display:inline-block;background:rgba(255,255,255,0.15);border:1px solid rgba(255,255,255,0.2);border-radius:20px;padding:3px 14px;color:rgba(255,255,255,0.9);font-size:12px;font-weight:600;margin-top:10px;}
.card-title{color:#0c3460;font-size:19px;font-weight:900;margin-bottom:14px;}
.field-lbl{font-size:12px;font-weight:700;color:#888780;margin-bottom:4px;}
.field-val{background:#eaf3de;border:1px solid #c0dd97;border-radius:14px;padding:12px 14px;font-size:15px;font-weight:800;color:#27500A;margin-bottom:12px;}
.field-val.blue{background:#e6f1fb;border-color:#185FA5;color:#185FA5;}
.pro-card{background:#fff;border-radius:22px;padding:18px 20px;box-shadow:0 2px 14px rgba(12,52,96,0.07);margin-bottom:14px;}
.today-strip{display:flex;justify-content:space-around;background:#f0f4f8;border-radius:14px;padding:12px 8px;margin-bottom:14px;}
.stat-cell{text-align:center!important;}
.stat-val{font-size:17px;font-weight:900;color:#0c3460;display:block;text-align:center!important;}
.stat-lbl{font-size:10px;font-weight:600;color:#888780;text-align:center!important;}
.footer-bar{background:#0c3460;border-radius:14px;padding:12px 18px;display:flex;justify-content:space-between;align-items:center;margin-top:12px;}
.footer-bar span{font-size:11px;font-weight:600;color:rgba(255,255,255,.7);}
.footer-bar .hl{color:#fff;}
.stButton button{border-radius:14px!important;font-size:15px!important;font-weight:800!important;font-family:'Cairo',sans-serif!important;}
.audit-row{background:#f8fafc;border-radius:10px;padding:10px 14px;border-right:3px solid #378ADD;margin-bottom:6px;font-size:12px;color:#0c3460;}
.warn-row{background:#faeeda;border-radius:10px;padding:10px 14px;border-right:3px solid #BA7517;margin-bottom:6px;font-size:12px;color:#633806;font-weight:700;}
.absent-row{background:#fcebeb;border-radius:10px;padding:10px 14px;border-right:3px solid #E24B4A;margin-bottom:6px;font-size:12px;color:#791F1F;font-weight:700;}
/* كرت إرشادات الموقع */
.gps-steps-box{background:#fff8e8;border:1px solid #f0c36d;border-radius:18px;padding:16px 18px;margin:12px 0;color:#5b3a05;text-align:right!important;}
.gps-steps-title{font-size:19px;font-weight:900;color:#6b4300;margin-bottom:10px;text-align:center!important;}
.gps-step{display:flex;gap:10px;align-items:flex-start;margin:8px 0;font-size:15px;font-weight:800;color:#173763;}
.gps-step-num{background:#185FA5;color:white;border-radius:50%;min-width:26px;height:26px;display:inline-flex;align-items:center;justify-content:center;font-weight:900;}
.gps-click-hint{background:#edf8ef;border:1px solid #9bd0a8;border-radius:16px;padding:14px 16px;margin:10px 0 10px 0;color:#1f6b2a;font-weight:900;text-align:center!important;font-size:17px;}
.no-gps-approved{background:#fff6e7;border:1px solid #eba83a;border-radius:14px;padding:12px 14px;color:#7a4700;font-weight:800;margin-top:10px;}
</style>
""", unsafe_allow_html=True)

# ─── Google Sheets ──────────────────────────────────────────────
scope = ["https://spreadsheets.google.com/feeds","https://www.googleapis.com/auth/drive"]

@st.cache_resource(show_spinner=False)
def get_all_sheets():
    creds = ServiceAccountCredentials.from_json_keyfile_dict(st.secrets["gcp_service_account"], scope)
    client = gspread.authorize(creds)
    ss = client.open_by_key(SHEET_ID)

    def _get_or_create(name, headers):
        try:
            return ss.worksheet(name)
        except gspread.WorksheetNotFound:
            ws = ss.add_worksheet(title=name, rows=1000, cols=max(len(headers),10))
            ws.append_row(headers)
            return ws

    return {
        "spreadsheet": ss,
        "sheet":       ss.worksheet("sheet1"),
        "whitelist":   _get_or_create("القائمة_البيضاء",         ["الرقم الشخصي","الاسم","المدرسة","المهمة","دعم","رقم التواصل","البريد الإلكتروني","المسمى الوظيفي","نشط"]),
        "schedule":    _get_or_create("جدول_دوام_الأقسام",       ["المهمة","السبت","الأحد","الاثنين","الثلاثاء","الأربعاء","الخميس","الجمعة","نشط"]),
        "daily_schedule": _get_or_create("دوام_الأقسام_اليومي", ["التاريخ","المهمة","نشط","ملاحظات"]),
        "required_today": _get_or_create("مطلوبات_اليوم", ["التاريخ","المهمة","المدرسة","الرقم الشخصي","الاسم","نشط","ملاحظات"]),
        "device":      _get_or_create("device_lock",              ["التاريخ","بصمة الجهاز","الرقم الشخصي","الاسم","وقت_القفل"]),
        "device_exceptions": _get_or_create("استثناءات_قفل_الجهاز", ["الرقم الشخصي","الاسم","تاريخ_الانتهاء","نشط","ملاحظات","تاريخ_الإضافة"]),
        "trusted_devices": _get_or_create("الأجهزة_الموثوقة", ["بصمة الجهاز","اسم الجهاز","نشط","ملاحظات","تاريخ الاعتماد","آخر استخدام"]),
        "attempts":    _get_or_create("محاولات_تسجيل_باسم_آخر",  ["التاريخ","بصمة الجهاز","الرقم_المقفول_عليه","اسم_المقفول_عليه","الرقم_المحاول","اسم_المحاول","وقت_المحاولة","ملاحظات"]),
        "settings":    _get_or_create("إعدادات_النظام",           ["المفتاح","القيمة","تاريخ_الانتهاء","ملاحظات"]),
        "audit":       _get_or_create("سجل_التدقيق",              ["التاريخ","الوقت","المستخدم","الرقم الشخصي","نوع العملية","التفاصيل","بصمة الجهاز"]),
        "absence":     _get_or_create("سجل_الغياب",               ["التاريخ","اليوم","الرقم الشخصي","الاسم","المدرسة","المهمة","سبب الغياب","ملاحظات","سجّله"]),
        "manual_requests": _get_or_create("طلبات_التسجيل_اليدوي", ["تاريخ الطلب","وقت الطلب","الرقم الشخصي","الاسم","المدرسة","المهمة","نوع الطلب","وقت الحضور الفعلي","وقت الانصراف الفعلي","نوع المشكلة","ملاحظات","الحالة","بصمة الجهاز","وقت الاعتماد","اعتمده"]),
        "time_permits":    _get_or_create("تصاريح_الوقت_اليدوي",   ["تاريخ الإضافة","الرقم الشخصي","نوع التصريح","تاريخ البداية","تاريخ النهاية","وقت الفتح","وقت الإغلاق","نشط","ملاحظات","أضافه"]),
        "custom_schedules":_get_or_create("إعدادات_الدوام_المخصص",  ["تاريخ الإضافة","نوع النطاق","قيمة النطاق","نوع الدوام","تاريخ البداية","تاريخ النهاية","نشط","ملاحظات","وقت البداية","عدد الساعات"]),
    }

_sheets         = get_all_sheets()
spreadsheet     = _sheets["spreadsheet"]
sheet           = _sheets["sheet"]
whitelist_sheet = _sheets["whitelist"]
device_sheet    = _sheets["device"]
device_exceptions_sheet = _sheets["device_exceptions"]
trusted_devices_sheet = _sheets["trusted_devices"]
attempts_sheet  = _sheets["attempts"]
settings_sheet  = _sheets["settings"]
audit_sheet     = _sheets["audit"]
absence_sheet   = _sheets["absence"]
manual_requests_sheet = _sheets["manual_requests"]
time_permits_sheet     = _sheets["time_permits"]
custom_schedules_sheet = _sheets["custom_schedules"]
required_today_sheet = _sheets["required_today"]

@st.cache_data(ttl=60, show_spinner=False)
def get_required_today_records():
    try: return required_today_sheet.get_all_records()
    except: return []

@st.cache_data(ttl=120)
def get_custom_schedules():
    try: return custom_schedules_sheet.get_all_records()
    except: return []

def _split_ar_values(value):
    """يفصل القيم المكتوبة بفواصل عربية/إنجليزية أو أسطر."""
    txt = str(value or "").replace("\n", "،").replace(",", "،")
    return [v.strip() for v in txt.split("،") if v.strip()]


def _custom_scope_priority(scope_type):
    """الأولوية: شخصي > مهمة > مدرسة > الكل."""
    scope_type = str(scope_type or "").strip()
    if scope_type in ["رقم شخصي", "أرقام", "اسم", "أسماء"]:
        return 1
    if scope_type in ["مهمة", "مهام"]:
        return 2
    if scope_type in ["مدرسة", "مدارس"]:
        return 3
    if scope_type == "الكل":
        return 4
    return 9


def get_custom_schedule_rule_for_row(row, date_str):
    """يرجع قاعدة الدوام المخصصة الأعلى أولوية لهذه الموظفة في هذا التاريخ.
    الأولوية: إعداد شخصي > مهمة/قسم > مدرسة > الكل.
    يدعم الأعمدة القديمة والجديدة: نوع الدوام، وقت البداية، عدد الساعات.
    """
    schedules = get_custom_schedules()
    emp_id   = str(row.get("الرقم الشخصي","")).strip()
    school   = str(row.get("اسم المدرسة", row.get("المدرسة","")).strip()).strip()
    task     = str(row.get("المهمة","")).strip()
    name     = str(row.get("الاسم الثلاثي", row.get("الاسم","")).strip()).strip()

    matches = []
    for idx, s in enumerate(schedules):
        if str(s.get("نشط","")).strip() not in ["نعم","yes","Yes","1","TRUE","true","✅"]:
            continue
        d_from = str(s.get("تاريخ البداية","")).strip()
        d_to   = str(s.get("تاريخ النهاية","")).strip()
        if d_from and date_str < d_from: continue
        if d_to   and date_str > d_to:   continue

        scope_type = str(s.get("نوع النطاق","")).strip()
        scope_val  = str(s.get("قيمة النطاق","")).strip()
        vals = _split_ar_values(scope_val)

        match = False
        if scope_type == "مدرسة" and scope_val == school:
            match = True
        elif scope_type == "مدارس" and school in vals:
            match = True
        elif scope_type == "مهمة" and scope_val == task:
            match = True
        elif scope_type == "مهام" and task in vals:
            match = True
        elif scope_type == "رقم شخصي" and scope_val == emp_id:
            match = True
        elif scope_type == "أرقام" and emp_id in vals:
            match = True
        elif scope_type == "اسم" and scope_val == name:
            match = True
        elif scope_type == "أسماء" and name in vals:
            match = True
        elif scope_type == "الكل":
            match = True

        if match:
            priority = _custom_scope_priority(scope_type)
            # عند نفس الأولوية نأخذ آخر إعداد مضاف لأنه الأحدث غالبًا.
            matches.append((priority, -idx, s))

    if not matches:
        return None
    matches.sort(key=lambda x: (x[0], x[1]))
    return matches[0][2]


def get_custom_schedule_for_row(row, date_str):
    """للتوافق مع أجزاء قديمة من الكود: يرجع نوع الدوام فقط إذا وُجد إعداد مخصص."""
    rule = get_custom_schedule_rule_for_row(row, date_str)
    if not rule:
        return None
    return str(rule.get("نوع الدوام","")).strip()

@st.cache_data(ttl=60)
def get_time_permits():
    try: return time_permits_sheet.get_all_records()
    except: return []

def get_active_permit(emp_id):
    """يرجع التصريح النشط للموظفة إذا وجد، مع مراعاة نطاق التاريخ ونافذة الوقت."""
    now     = now_bh()
    today   = now.strftime("%Y-%m-%d")
    now_t   = now.strftime("%H:%M")
    permits = get_time_permits()
    for p in permits:
        if str(p.get("نشط","")).strip() not in ["نعم","yes","1","TRUE","true"]:
            continue
        p_id    = str(p.get("الرقم الشخصي","")).strip()
        # يطابق الموظفة أو كل الموظفات
        if p_id not in ["","الكل","*"] and p_id != str(emp_id).strip():
            continue
        # نطاق التاريخ
        d_from  = str(p.get("تاريخ البداية","")).strip()
        d_to    = str(p.get("تاريخ النهاية","")).strip()
        if d_from and today < d_from: continue
        if d_to   and today > d_to:   continue
        # نافذة الوقت (اختيارية)
        t_open  = str(p.get("وقت الفتح","")).strip()
        t_close = str(p.get("وقت الإغلاق","")).strip()
        if t_open  and now_t < t_open:  continue
        if t_close and now_t > t_close: continue
        return p
    return None

def has_time_permit(emp_id, permit_type="كليهما"):
    p = get_active_permit(emp_id)
    if not p: return False
    p_type = str(p.get("نوع التصريح","")).strip()
    return p_type in ["كليهما", permit_type] or permit_type == "كليهما"

def get_permit_dates(emp_id):
    """يرجع (تاريخ البداية, تاريخ النهاية) للتصريح النشط."""
    p = get_active_permit(emp_id)
    if not p: return None, None
    return str(p.get("تاريخ البداية","")).strip(), str(p.get("تاريخ النهاية","")).strip()

def add_time_permit(emp_id, permit_type, date_from, date_to, time_open="", time_close="", note=""):
    today = now_bh().strftime("%Y-%m-%d")
    time_permits_sheet.append_row([today, emp_id, permit_type, date_from, date_to, time_open, time_close, "نعم", note, "أدمن"])
    get_time_permits.clear()

def revoke_time_permit_row(row_num):
    time_permits_sheet.update_cell(row_num, 8, "لا")
    get_time_permits.clear()
schedule_sheet  = _sheets["schedule"]
daily_schedule_sheet = _sheets["daily_schedule"]

# ─── ضمان الأعمدة الجديدة بدون الرجوع إلى Google Sheet ───────────────
def ensure_headers(ws, headers):
    try:
        current = ws.row_values(1)
        for i, h in enumerate(headers, start=1):
            if len(current) < i or not str(current[i-1]).strip():
                ws.update_cell(1, i, h)
    except Exception:
        pass

SHEET1_HEADERS = [
    "التاريخ","اليوم","اسم المدرسة","المهمة","دعم","الاسم الثلاثي","الرقم الشخصي",
    "وقت الحضور","سبب التأخير","وقت الانصراف","سبب الانصراف","خروج استئذان","عودة","محاولة",
    "وقت البداية المحسوب","وقت النهاية المتوقع","ساعات العمل","الساعات الإضافية",
    "حالة الدوام","نوع الدوام اليومي","إغلاق تلقائي","تأكيد الرعاية","نوع التسجيل"
]
WHITELIST_HEADERS = [
    "الرقم الشخصي","الاسم","المدرسة","المهمة","دعم","رقم التواصل","البريد الإلكتروني",
    "المسمى الوظيفي","نشط","نوع الدوام الافتراضي","ملاحظات"
]
ensure_headers(sheet, SHEET1_HEADERS)
ensure_headers(whitelist_sheet, WHITELIST_HEADERS)
CUSTOM_SCHEDULE_HEADERS = ["تاريخ الإضافة","نوع النطاق","قيمة النطاق","نوع الدوام","تاريخ البداية","تاريخ النهاية","نشط","ملاحظات","وقت البداية","عدد الساعات"]
REQUIRED_TODAY_HEADERS = ["التاريخ","المهمة","المدرسة","الرقم الشخصي","الاسم","نشط","ملاحظات"]
ensure_headers(custom_schedules_sheet, CUSTOM_SCHEDULE_HEADERS)
ensure_headers(required_today_sheet, REQUIRED_TODAY_HEADERS)

# ─── دوال مساعدة ───────────────────────────────────────────────
def ar_to_en_digits(text):
    ar="٠١٢٣٤٥٦٧٨٩"; en="0123456789"
    result=str(text).strip()
    for a,e in zip(ar,en): result=result.replace(a,e)
    return result

def normalize_emp_id(value):
    """تنظيف الرقم الشخصي للمقارنة والبحث حتى لو رجع من Google Sheet كرقم أو .0."""
    txt = ar_to_en_digits(str(value or "")).strip()
    txt = txt.replace("'", "").replace('"', "").replace(" ", "")
    if txt.endswith(".0"):
        txt = txt[:-2]
    # إذا كان رقمًا عشريًا من Google Sheets مثل 830704019.0
    try:
        if "." in txt and txt.replace(".", "", 1).isdigit():
            txt = str(int(float(txt)))
    except Exception:
        pass
    return txt.strip()

def normalize_name(name):
    name=str(name).strip()
    for old,new in {"أ":"ا","إ":"ا","آ":"ا","ى":"ي","ة":"ه","ؤ":"و","ئ":"ي"}.items():
        name=name.replace(old,new)
    for ch in [".",  "،",",","-","_","ـ",":",";"] : name=name.replace(ch," ")
    return " ".join(name.split())

def normalize_school_name(value):
    """تطبيع اسم المدرسة للمقارنة حتى لو كان فيه مسافات/اختلافات بسيطة."""
    txt = str(value or "").strip()
    txt = txt.replace("ـ", "")
    txt = txt.replace("  ", " ")
    txt = " ".join(txt.split())
    return normalize_name(txt)


def get_emp_school(emp):
    """يدعم اختلاف اسم عمود المدرسة بين القائمة البيضاء والشيتات."""
    return str(emp.get("المدرسة", emp.get("اسم المدرسة", "")) or "").strip()


def get_emp_name(emp):
    """يدعم اختلاف اسم عمود الاسم بين القائمة البيضاء والشيتات."""
    return str(emp.get("الاسم", emp.get("الاسم الثلاثي", "")) or "").strip()

def distance_m(lat1,lon1,lat2,lon2):
    R=6371000
    p1,p2=math.radians(lat1),math.radians(lat2)
    dp=math.radians(lat2-lat1); dl=math.radians(lon2-lon1)
    a=math.sin(dp/2)**2+math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2*R*math.atan2(math.sqrt(a),math.sqrt(1-a))

def ls_get(key):
    if LOCAL_STORAGE_OK:
        try: return localS.getItem(key)
        except: pass
    return st.session_state.get(f"ls_{key}")

def ls_set(key, value, ls_key=None):
    if LOCAL_STORAGE_OK:
        try: localS.setItem(key, value, key=ls_key or f"set_{key}"); return
        except: pass
    st.session_state[f"ls_{key}"] = value

def ls_clear_emp_data():
    """يضع علامة مسح في LocalStorage تمنع قراءة البيانات القديمة."""
    ls_set("trusted_cleared", "yes", "set_trusted_cleared")

def get_device_fingerprint():
    if LOCAL_STORAGE_OK:
        try:
            fp=localS.getItem("device_fp")
            if not fp:
                fp=''.join(random.choices(string.ascii_letters+string.digits,k=32))
                localS.setItem("device_fp",fp,key="set_device_fp")
            return fp or "unknown"
        except: pass
    if "device_fp" not in st.session_state:
        st.session_state.device_fp=''.join(random.choices(string.ascii_letters+string.digits,k=32))
    return st.session_state.device_fp

def safe_append(ws, row, retries=3):
    for _ in range(retries):
        try: ws.append_row(row,value_input_option="USER_ENTERED"); return True
        except: time_module.sleep(1.5)
    return False

def safe_update(ws, row, col, value, retries=3):
    for _ in range(retries):
        try: ws.update_cell(row,col,value); return True
        except: time_module.sleep(1.5)
    return False

@st.cache_data(ttl=30, show_spinner=False)
def get_sheet_data():
    try: return sheet.get_all_records()
    except: return []

@st.cache_data(ttl=300, show_spinner=False)
def get_whitelist():
    try:
        records=whitelist_sheet.get_all_records(); result={}
        for r in records:
            active=str(r.get("نشط","")).strip()
            if active in ["نعم","yes","Yes","TRUE","true","1"]:
                eid=normalize_emp_id(r.get("الرقم الشخصي",""))
                if eid:
                    r["الرقم الشخصي"] = eid
                    result[eid]=r
        return result
    except: return {}

@st.cache_data(ttl=120, show_spinner=False)
def get_device_locks():
    try: return device_sheet.get_all_records()[-500:]
    except: return []

@st.cache_data(ttl=60, show_spinner=False)
def get_device_exceptions():
    try: return device_exceptions_sheet.get_all_records()
    except: return []

@st.cache_data(ttl=60, show_spinner=False)
def get_trusted_devices():
    try: return trusted_devices_sheet.get_all_records()
    except: return []

@st.cache_data(ttl=60, show_spinner=False)
def get_settings_records():
    try: return settings_sheet.get_all_records()
    except: return []

def get_system_setting(key, default=""):
    """قراءة قيمة من ورقة إعدادات_النظام."""
    try:
        for r in get_settings_records():
            if str(r.get("المفتاح", "")).strip() == str(key).strip():
                return str(r.get("القيمة", default)).strip()
    except Exception:
        pass
    return default


def set_system_setting(key, value, note=""):
    """حفظ/تحديث قيمة في ورقة إعدادات_النظام."""
    try:
        records = settings_sheet.get_all_records()
        row_found = None
        for i, r in enumerate(records):
            if str(r.get("المفتاح", "")).strip() == str(key).strip():
                row_found = i + 2
                break
        row_values = [str(key), str(value), "", note]
        if row_found:
            settings_sheet.update(f"A{row_found}:D{row_found}", [row_values], value_input_option="USER_ENTERED")
        else:
            safe_append(settings_sheet, row_values)
        get_settings_records.clear()
        return True
    except Exception as e:
        st.error(f"❌ تعذر حفظ الإعداد: {e}")
        return False


def manual_requests_enabled():
    """طلبات التسجيل اليدوي من واجهة الموظفة مفعلة افتراضيًا، ويمكن تعطيلها من الأدمن."""
    val = get_system_setting("manual_requests_enabled", "true").lower()
    return val in ["true", "1", "yes", "نعم", "مفعل", "on", ""]

@st.cache_data(ttl=60, show_spinner=False)
def get_manual_requests():
    try: return manual_requests_sheet.get_all_records()
    except: return []

def manual_request_exists_today(emp_id, req_type):
    """يمنع تكرار نفس نوع الطلب لنفس الرقم في نفس التاريخ فقط.
    يسمح بحضور ثم انصراف، لكنه يمنع حضورين أو انصرافين في نفس اليوم.
    """
    emp_id = ar_to_en_digits(emp_id).strip()
    req_type = str(req_type or "").strip()
    today = now_bh().strftime("%Y-%m-%d")
    try:
        for r in get_manual_requests():
            status = str(r.get("الحالة", "")).strip()
            if status == "مرفوض":
                continue
            if str(r.get("تاريخ الطلب", "")).strip() == today and \
               str(r.get("الرقم الشخصي", "")).strip() == emp_id and \
               str(r.get("نوع الطلب", "")).strip() == req_type:
                return True, r
    except Exception:
        pass
    return False, None

@st.cache_data(ttl=120, show_spinner=False)
def get_schedule_records():
    try: return schedule_sheet.get_all_records()
    except: return []

@st.cache_data(ttl=60, show_spinner=False)
def get_daily_schedule_records():
    try: return daily_schedule_sheet.get_all_records()
    except: return []

def clear_caches():
    get_sheet_data.clear(); get_device_locks.clear(); get_device_exceptions.clear(); get_trusted_devices.clear(); get_settings_records.clear(); get_schedule_records.clear(); get_daily_schedule_records.clear(); get_manual_requests.clear(); get_time_permits.clear(); get_custom_schedules.clear(); get_required_today_records.clear()


# ─── دوال احتساب الدوام والساعات والإغلاق التلقائي ───────────────
def parse_time_value(t):
    t = str(t or "").strip()
    if not t:
        return None
    for fmt in ["%H:%M:%S", "%H:%M"]:
        try:
            return datetime.strptime(t, fmt).time()
        except Exception:
            pass
    return None

def combine_date_time(date_str, t):
    if not t:
        return None
    try:
        d = datetime.strptime(str(date_str), "%Y-%m-%d").date()
        return datetime.combine(d, t).replace(tzinfo=BAHRAIN_TZ)
    except Exception:
        return None

def fmt_time_dt(dt):
    return dt.strftime("%H:%M:%S") if dt else ""

def fmt_hours_from_seconds(seconds):
    seconds = max(0, int(seconds))
    h = seconds // 3600
    m = (seconds % 3600) // 60
    return f"{h}:{m:02d}"

def is_care_day(row):
    txt = " ".join([str(row.get("سبب التأخير", "")), str(row.get("سبب الانصراف", "")), str(row.get("تأكيد الرعاية", "")), str(row.get("نوع الدوام اليومي", ""))])
    return "رعاية" in txt

def is_flexible_day(row):
    txt = " ".join([str(row.get("سبب التأخير", "")), str(row.get("سبب الانصراف", "")), str(row.get("نوع الدوام اليومي", ""))])
    return "دوام مرن" in txt

def is_official_mission_day(row):
    txt = " ".join([str(row.get("سبب التأخير", "")), str(row.get("سبب الانصراف", "")), str(row.get("نوع الدوام اليومي", ""))])
    return "مهمة رسمية" in txt

def is_correction_done_day(row):
    """حالة الانتهاء من التصحيح: لا تُحسب كتأخير ولا كساعات إضافية، وتظهر كحالة معفاة."""
    txt = " ".join([str(row.get("سبب التأخير", "")), str(row.get("سبب الانصراف", "")), str(row.get("حالة الدوام", "")), str(row.get("نوع الدوام اليومي", ""))])
    return "الانتهاء من التصحيح" in txt

def is_implicit_leave_late(row):
    """
    إذا كان الدوام عادي والموظفة حضرت بعد السماح وكتبت سببًا مثل موعد أو سبب آخر،
    فهذا يعتبر تأخيرًا موثقًا كاستئذان ضمني من بداية الدوام،
    ووقت الحضور يعتبر عودة من الاستئذان.
    لا ينطبق على الرعاية أو الدوام المرن أو المهمة الرسمية.
    """
    att = parse_time_value(row.get("وقت الحضور", ""))
    if not att or att <= time(7, 5, 30):
        return False
    reason = str(row.get("سبب التأخير", "")).strip()
    if not reason:
        return False
    if any(x in reason for x in ["رعاية", "دوام مرن", "مهمة رسمية", "الانتهاء من التصحيح"]):
        return False
    return True

def is_late_for_statistics(row):
    """التأخير الإحصائي: يحسب للدوام العادي فقط، ويستثني الرعاية/المرن/المهمة الرسمية."""
    att = parse_time_value(row.get("وقت الحضور", ""))
    if not att or att <= time(7, 5, 30):
        return False
    if is_care_day(row) or is_flexible_day(row) or is_official_mission_day(row) or is_correction_done_day(row):
        return False
    return True

def calculate_work_values(row):
    date_str = str(row.get("التاريخ", "")).strip()
    att = parse_time_value(row.get("وقت الحضور", ""))
    dep = parse_time_value(row.get("وقت الانصراف", ""))
    if not date_str or not att:
        return None
    care = is_care_day(row)
    flexible = is_flexible_day(row)
    att_dt = combine_date_time(date_str, att)
    dep_dt = combine_date_time(date_str, dep) if dep else None
    official_start = combine_date_time(date_str, time(7, 0, 0))
    grace_end = combine_date_time(date_str, time(7, 5, 30))
    official_mission = is_official_mission_day(row)
    correction_done = is_correction_done_day(row)
    implicit_leave = is_implicit_leave_late(row)

    # ── إعداد دوام مخصص من الأدمن ──
    # الأولوية: إعداد شخصي > المهمة > المدرسة > الكل.
    custom_rule = get_custom_schedule_rule_for_row(row, date_str)
    custom_required_hours = None
    if custom_rule and not correction_done:
        custom_type = str(custom_rule.get("نوع الدوام", "")).strip()
        custom_start_txt = str(custom_rule.get("وقت البداية", "") or "07:00").strip()
        custom_hours_txt = str(custom_rule.get("عدد الساعات", "") or "").strip()

        custom_start_time = parse_time_value(custom_start_txt) or time(7, 0, 0)
        official_start = combine_date_time(date_str, custom_start_time)
        grace_end = official_start + timedelta(minutes=5, seconds=30)

        try:
            custom_required_hours = float(str(custom_hours_txt).replace("٫", ".")) if custom_hours_txt else None
        except Exception:
            custom_required_hours = None

        if custom_type == "دوام مرن":
            flexible = True
            care = False
        elif custom_type == "رعاية":
            care = True
            flexible = False
        elif custom_type == "دوام عادي":
            flexible = False
            care = False

    if correction_done:
        daily_type = "انتهاء التصحيح"
        required_hours = 0
        calc_start = official_start if att_dt <= grace_end else att_dt
    elif care:
        daily_type = "رعاية"
        required_hours = custom_required_hours if custom_required_hours is not None else 5
        calc_start = max(att_dt, official_start)
    elif flexible:
        daily_type = "دوام مرن"
        required_hours = custom_required_hours if custom_required_hours is not None else 7
        calc_start = official_start  # يحسب دائماً من وقت البداية المعتمد حتى لو جاءت متأخرة
    elif official_mission:
        daily_type = "مهمة رسمية"
        required_hours = custom_required_hours if custom_required_hours is not None else 7
        calc_start = official_start
    elif implicit_leave:
        daily_type = "استئذان تأخير"
        required_hours = custom_required_hours if custom_required_hours is not None else 7
        calc_start = official_start
    else:
        daily_type = "دوام عادي"
        required_hours = custom_required_hours if custom_required_hours is not None else 7
        calc_start = official_start if att_dt <= grace_end else att_dt
    expected_end = calc_start + timedelta(hours=required_hours)
    work_seconds = 0
    extra_seconds = 0
    status = "لم يكتمل"
    if dep_dt:
        if dep_dt < att_dt:
            dep_dt += timedelta(days=1)
        work_seconds = max(0, int((dep_dt - calc_start).total_seconds()))
        if correction_done:
            # الانتهاء من التصحيح حالة معفاة: نحتفظ بالوقت الفعلي للعرض فقط، ولا نحسب نقصًا أو إضافيًا.
            expected_end = dep_dt
            extra_seconds = 0
            status = "معفى - انتهاء التصحيح"
        else:
            extra_seconds = max(0, int((dep_dt - expected_end).total_seconds()))
            if work_seconds < required_hours * 3600:
                status = "ناقص"
            elif extra_seconds > 0:
                status = "رعاية + إضافي" if care else "مكتمل + إضافي"
            else:
                status = "مكتمل رعاية" if care else "مكتمل"
    return {"calc_start": fmt_time_dt(calc_start), "expected_end": fmt_time_dt(expected_end), "work_hours": fmt_hours_from_seconds(work_seconds), "extra_hours": fmt_hours_from_seconds(extra_seconds), "status": status, "daily_type": daily_type}

def update_work_calculation(row_index, row_data=None):
    try:
        if row_data is None:
            # قراءة مباشرة بدون كاش حتى تكون إعادة الحساب مبنية على آخر بيانات في sheet1
            records = get_sheet_data_fresh()
            row_data = records[row_index - 2]
        vals = calculate_work_values(row_data)
        if not vals:
            return False
        sheet.update(f"O{row_index}:T{row_index}", [[vals["calc_start"], vals["expected_end"], vals["work_hours"], vals["extra_hours"], vals["status"], vals["daily_type"]]], value_input_option="USER_ENTERED")
        return True
    except Exception:
        return False

def auto_close_previous_open_records():
    try:
        today      = now_bh().strftime("%Y-%m-%d")
        now_t      = now_bh().time()
        # وقت الإغلاق التلقائي لليوم الحالي — قابل للتغيير من الإعدادات
        close_time_str = get_system_setting("auto_close_time", "22:00")
        try:
            ch, cm    = map(int, close_time_str.split(":"))
            close_time = time(ch, cm)
        except Exception:
            close_time = time(22, 0)

        # قراءة مباشرة بدون كاش حتى لا يغلق النظام سجلات بناءً على نسخة قديمة
        records = get_sheet_data_fresh()
        changed = 0
        for i, row in enumerate(records):
            row_num  = i + 2
            date_str = str(row.get("التاريخ", "")).strip()
            if not date_str:
                continue
            # أيام سابقة دائماً، أو اليوم بعد وقت الإغلاق
            is_prev_day  = date_str < today
            is_today_late = (date_str == today and now_t >= close_time)
            if not (is_prev_day or is_today_late):
                continue
            if not row.get("وقت الحضور") or row.get("وقت الانصراف"):
                continue
            if row.get("خروج استئذان") and not row.get("عودة"):
                dep_time   = str(row.get("خروج استئذان", "")).strip()
                dep_reason = "إغلاق تلقائي — استئذان مفتوح احتُسب انصرافًا"
                auto_note  = "نعم — استئذان مفتوح"
            else:
                vals       = calculate_work_values(row)
                dep_time   = vals["expected_end"] if vals else close_time_str + ":00"
                dep_reason = "إغلاق تلقائي — نسيان تسجيل الانصراف"
                auto_note  = "نعم — نسيان انصراف"
            if dep_time:
                safe_update(sheet, row_num, COL_DEPART, dep_time)
                safe_update(sheet, row_num, COL_DEPART_REASON, dep_reason)
                safe_update(sheet, row_num, COL_AUTO_CLOSE, auto_note)
                new_row = dict(row)
                new_row["وقت الانصراف"] = dep_time
                new_row["سبب الانصراف"] = dep_reason
                update_work_calculation(row_num, new_row)
                changed += 1
        if changed:
            get_sheet_data.clear()
    except Exception:
        pass


def show_previous_auto_close_notice(emp_id):
    """تنبيه الموظفة في اليوم الجديد إذا تم إغلاق يوم سابق تلقائياً بسبب نسيان الانصراف أو استئذان مفتوح."""
    try:
        emp_id = str(emp_id or "").strip()
        if not emp_id:
            return

        today = now_bh().date()
        yesterday = (today - timedelta(days=1)).strftime("%Y-%m-%d")
        today_key = today.strftime("%Y-%m-%d")
        dismissed_key = f"auto_close_notice_{emp_id}_{today_key}"

        if str(ls_get(dismissed_key) or "").strip() == "done":
            return

        data = get_sheet_data()
        matches = []
        for r in data:
            if str(r.get("الرقم الشخصي", "")).strip() != emp_id:
                continue
            if str(r.get("التاريخ", "")).strip() != yesterday:
                continue
            if str(r.get("إغلاق تلقائي", "")).strip():
                matches.append(r)

        if not matches:
            return

        r = matches[-1]
        st.warning(
            "⚠️ تنبيه: تم إغلاق سجل يوم أمس تلقائيًا لأن الانصراف لم يُسجَّل قبل نهاية اليوم. "
            "يرجى التأكد من تسجيل الانصراف يوميًا قبل مغادرة المركز."
        )
        st.markdown(f"""
        <div class="warn-row">
            <b>تفاصيل الإغلاق التلقائي:</b><br>
            التاريخ: {r.get('التاريخ','')}<br>
            وقت الحضور: {r.get('وقت الحضور','—') or '—'}<br>
            وقت الانصراف المسجّل تلقائيًا: {r.get('وقت الانصراف','—') or '—'}<br>
            السبب: {r.get('سبب الانصراف','—') or '—'}<br>
            الحالة: {r.get('إغلاق تلقائي','—') or '—'}
        </div>
        """, unsafe_allow_html=True)

        if st.button("✅ تم الاطلاع على التنبيه", use_container_width=True, key=f"dismiss_auto_close_{emp_id}_{today_key}"):
            ls_set(dismissed_key, "done", f"set_{dismissed_key}")
            st.success("✅ تم إخفاء التنبيه لهذا اليوم.")
            st.rerun()
    except Exception:
        pass


def show_previous_absence_notice(emp_id):
    """تنبيه الموظفة إذا احتُسب عليها غياب ليوم سابق لأنها لم تسجل حضوراً إلكترونياً.
    إذا لم يوجد سجل حضور ليوم أمس، وكان مطلوباً دوامها حسب جدول الأقسام، يسجل النظام الغياب مرة واحدة.
    """
    try:
        emp_id = str(emp_id or "").strip()
        if not emp_id:
            return

        today = now_bh().date()
        yesterday_date = today - timedelta(days=1)
        yesterday = yesterday_date.strftime("%Y-%m-%d")
        today_key = today.strftime("%Y-%m-%d")
        dismissed_key = f"absence_notice_{emp_id}_{today_key}"

        if str(ls_get(dismissed_key) or "").strip() == "done":
            return

        emp = validate_employee(emp_id) or st.session_state.get("emp_data") or {}
        if not emp:
            return

        # لا نحتسب الدعم تلقائياً كغياب لأن الدعم قد لا يكون مطلوباً يومياً
        task_txt = str(emp.get("المهمة", "")).strip()
        if is_yes(emp.get("دعم", "")) or "دعم" in task_txt:
            return

        scheduled_tasks, _schedule_source = scheduled_tasks_for_date(yesterday)
        if not emp_required_on_day(emp, scheduled_tasks):
            return

        # إذا يوجد حضور إلكتروني ليوم أمس، لا نسجل غياب
        data = get_sheet_data_fresh()
        had_attendance = any(
            str(r.get("التاريخ", "")).strip().replace("/", "-") == yesterday
            and str(r.get("الرقم الشخصي", "")).strip() == emp_id
            and str(r.get("وقت الحضور", "")).strip()
            for r in data
        )
        if had_attendance:
            return

        reason_txt = "عدم تسجيل الحضور والانصراف إلكترونيًا"
        abs_records = absence_sheet.get_all_records()
        existing_abs = None
        for r in abs_records:
            if str(r.get("التاريخ", "")).strip().replace("/", "-") == yesterday and str(r.get("الرقم الشخصي", "")).strip() == emp_id:
                existing_abs = r
                break

        full_name = normalize_name(emp.get("الاسم", ""))
        school = str(emp.get("المدرسة", "")).strip()
        task = str(emp.get("المهمة", "")).strip()
        yesterday_day = day_ar_from_date(yesterday_date)

        # يسجل الغياب مرة واحدة فقط إذا لم يكن مسجلاً سابقاً
        if not existing_abs:
            absence_sheet.append_row([
                yesterday, yesterday_day, emp_id, full_name, school, task,
                reason_txt, "احتساب تلقائي عند فتح النظام في اليوم التالي", "النظام"
            ], value_input_option="USER_ENTERED")
            log_audit(emp_id, full_name, "احتساب غياب تلقائي", f"التاريخ:{yesterday}|السبب:{reason_txt}")
            clear_caches()
            existing_abs = {"التاريخ": yesterday, "سبب الغياب": reason_txt}

        st.error(
            "⚠️ تنبيه: تم احتساب غياب ليوم أمس لعدم وجود تسجيل إلكتروني للحضور والانصراف. "
            "التسجيل الورقي وحده لا يُعتمد في النظام."
        )
        st.markdown(f"""
        <div class="absent-row">
            <b>تفاصيل الغياب المحتسب:</b><br>
            التاريخ: {yesterday}<br>
            السبب: {existing_abs.get('سبب الغياب', reason_txt) or reason_txt}<br>
            ملاحظة: يجب تسجيل الحضور والانصراف إلكترونيًا يوميًا حتى يتم احتساب الحضور.
        </div>
        """, unsafe_allow_html=True)

        if st.button("✅ تم الاطلاع على تنبيه الغياب", use_container_width=True, key=f"dismiss_absence_notice_{emp_id}_{today_key}"):
            ls_set(dismissed_key, "done", f"set_{dismissed_key}")
            st.success("✅ تم إخفاء تنبيه الغياب لهذا اليوم.")
            st.rerun()
    except Exception:
        pass

def mark_care_for_today(emp_id):
    data = get_sheet_data()
    idx, row = find_today_row(data, today_str, str(emp_id).strip())
    if not idx or not row:
        st.error("❌ لا يوجد سجل لهذا اليوم.")
        return False
    if not row.get("وقت الانصراف"):
        st.error("❌ يجب تسجيل الانصراف أولاً.")
        return False
    old_reason = str(row.get("سبب الانصراف", "")).strip()
    new_reason = old_reason if "رعاية" in old_reason else (old_reason + " | رعاية" if old_reason else "رعاية")
    safe_update(sheet, idx, COL_DEPART_REASON, new_reason)
    safe_update(sheet, idx, COL_CARE_CONF, "نعم")
    row["سبب الانصراف"] = new_reason
    row["تأكيد الرعاية"] = "نعم"
    update_work_calculation(idx, row)
    log_audit(emp_id, row.get("الاسم الثلاثي", ""), "تأكيد رعاية", "تأكيد الموظفة أن دوام اليوم رعاية بعد تسجيل الانصراف")
    clear_caches()
    st.success("✅ تم اعتماد الرعاية لهذا اليوم وإعادة احتساب الساعات على أساس 5 ساعات.")
    return True

def validate_employee(emp_id):
    return get_whitelist().get(normalize_emp_id(emp_id))


def validate_employee_fresh(emp_id):
    """قراءة مباشرة من ورقة القائمة_البيضاء بدون كاش.
    نستخدمها في واجهة الموظفة حتى لا تظهر مدرسة/مهمة قديمة محفوظة في المتصفح أو الكاش.
    """
    emp_id = ar_to_en_digits(str(emp_id or "")).strip()
    if not emp_id:
        return None
    try:
        records = whitelist_sheet.get_all_records()
        for r in records:
            eid = ar_to_en_digits(str(r.get("الرقم الشخصي", "")).strip())
            active = str(r.get("نشط", "")).strip()
            if eid == emp_id and active in ["نعم", "yes", "Yes", "TRUE", "true", "1", "✅"]:
                return r
    except Exception:
        pass
    return validate_employee(emp_id)


def is_support_employee_record(emp):
    """تحديد هل السجل دعم أو عضوة."""
    try:
        support_raw = str(emp.get("دعم", "")).strip()
        task_txt = str(emp.get("المهمة", "")).strip()
        reg_type = str(emp.get("نوع التسجيل", "")).strip()
        return is_yes(support_raw) or "دعم" in task_txt or "دعم" in reg_type
    except Exception:
        return False


def employee_category_label(emp):
    return "دعم" if is_support_employee_record(emp) else "عضوة"


def filter_rows_by_category(rows, category_choice):
    """فلترة التقارير حسب الأعضاء/الدعم."""
    if category_choice == "الأعضاء فقط":
        return [r for r in rows if not is_support_employee_record(r)]
    if category_choice == "الدعم فقط":
        return [r for r in rows if is_support_employee_record(r)]
    return rows


def whitelist_options_by_filters(school_filter="الكل", task_filter="الكل"):
    """إرجاع خيارات أسماء من القائمة البيضاء حسب المدرسة والمهمة مع تطبيع اسم المدرسة."""
    opts = []
    school_filter_norm = normalize_school_name(school_filter)
    task_filter_norm = normalize_name(task_filter)

    for eid, emp in get_whitelist().items():
        school = get_emp_school(emp)
        task = str(emp.get("المهمة", "") or "").strip()
        name = get_emp_name(emp)

        if school_filter and school_filter != "الكل" and normalize_school_name(school) != school_filter_norm:
            continue
        if task_filter and task_filter != "الكل" and normalize_name(task) != task_filter_norm:
            continue
        if eid and name:
            opts.append((eid, emp, f"{name} — #{eid} — {school} — {task}"))
    return sorted(opts, key=lambda x: x[2])


def find_whitelist_matches(term, limit=80):
    """بحث عام بالاسم أو الرقم الشخصي من كل القائمة البيضاء، بدون التقيد بالمدرسة أو المهمة."""
    term = ar_to_en_digits(str(term or "")).strip()
    if not term:
        return []

    norm_term = normalize_name(term)
    matches = []

    for eid, emp in get_whitelist().items():
        name = get_emp_name(emp)
        school = get_emp_school(emp)
        task = str(emp.get("المهمة", "") or "").strip()

        # البحث هنا عام من كل القائمة البيضاء، حتى لو كانت الموظفة مسجلة تحت مدرسة أخرى
        if term in str(eid).strip() or norm_term in normalize_name(name):
            matches.append((eid, emp, f"{name} — #{eid} — {school} — {task}"))

    return sorted(matches, key=lambda x: x[2])[:limit]

def day_ar_from_date(date_obj):
    return {"Saturday":"السبت","Sunday":"الأحد","Monday":"الاثنين","Tuesday":"الثلاثاء","Wednesday":"الأربعاء","Thursday":"الخميس","Friday":"الجمعة"}.get(date_obj.strftime("%A"), date_obj.strftime("%A"))

def is_yes(value):
    return str(value).strip() in ["نعم","yes","Yes","TRUE","true","1","✅","صح"]

def normalize_task_for_schedule(task):
    task=str(task or "").strip()
    # إذا كانت المهمة دعم — نفس القسم، نطابقها على اسم الدعم أو اسم المصححة عند الحاجة
    return task

def scheduled_tasks_for_day(day_ar):
    """الجدول الأسبوعي القديم حسب اليوم."""
    records=get_schedule_records()
    tasks=[]
    for r in records:
        task=str(r.get("المهمة","")).strip()
        active_raw=str(r.get("نشط","")).strip()
        active = active_raw=="" or is_yes(active_raw)
        if task and active and is_yes(r.get(day_ar,"")):
            tasks.append(task)
    # إزالة التكرار مع الحفاظ على الترتيب
    return list(dict.fromkeys(tasks)) if tasks else None

def scheduled_tasks_for_date(date_str):
    """يرجع مهام الدوام لتاريخ محدد.
    الأولوية لجدول دوام_الأقسام_اليومي إذا وُجدت مهام نشطة لذلك التاريخ.
    إذا لا يوجد جدول يومي، يرجع للجدول الأسبوعي حسب اليوم.
    """
    date_str = str(date_str).strip()
    daily_tasks=[]
    try:
        for r in get_daily_schedule_records():
            if str(r.get("التاريخ","")).strip()==date_str:
                task=str(r.get("المهمة","")).strip()
                active_raw=str(r.get("نشط","")).strip()
                active = active_raw=="" or is_yes(active_raw)
                if task and active:
                    daily_tasks.append(task)
    except Exception:
        pass
    if daily_tasks:
        return list(dict.fromkeys(daily_tasks)), "يومي"
    try:
        d=datetime.strptime(date_str,"%Y-%m-%d").date()
        day_ar=day_ar_from_date(d)
    except Exception:
        day_ar=day_arabic
    return scheduled_tasks_for_day(day_ar), "أسبوعي"

def emp_required_on_day(emp, scheduled_tasks):
    """يتحقق هل الموظفة ضمن أقسام/مهام الدوام المختارة لهذا اليوم."""
    if scheduled_tasks is None:
        return True
    task=str(emp.get("المهمة","")).strip()
    if task in scheduled_tasks:
        return True
    # دعم — الرياضيات يعتبر ضمن الرياضيات إذا تم اختيار مصححة — الرياضيات والعكس
    task_clean=task.replace("دعم —","").replace("مصححة —","").strip()
    for t in scheduled_tasks:
        t_clean=str(t).replace("دعم —","").replace("مصححة —","").strip()
        if task_clean and task_clean==t_clean:
            return True
    return False


def required_people_for_date(date_str):
    """يرجع قائمة مطلوبات اليوم من ورقة مطلوبات_اليوم.
    إذا رجعت None يعني لا توجد قائمة خاصة لهذا التاريخ، فيرجع النظام لجدول الأقسام.
    الترتيب المعتمد: المهمة ← المدرسة ← الاسم.
    """
    date_str = str(date_str or "").strip()
    wl_all = get_whitelist()
    result = {}

    for r in get_required_today_records():
        r_date = str(r.get("التاريخ", "")).strip().replace("/", "-")
        if r_date != date_str:
            continue
        active_raw = str(r.get("نشط", "نعم")).strip()
        if active_raw and not is_yes(active_raw):
            continue

        eid = ar_to_en_digits(str(r.get("الرقم الشخصي", "")).strip())
        if not eid:
            continue

        emp = dict(wl_all.get(eid, {}))
        emp["الرقم الشخصي"] = eid
        emp["الاسم"] = str(r.get("الاسم", "") or get_emp_name(emp)).strip()
        emp["المدرسة"] = str(r.get("المدرسة", "") or get_emp_school(emp)).strip()
        emp["المهمة"] = str(r.get("المهمة", "") or emp.get("المهمة", "")).strip()
        emp["نشط"] = "نعم"
        if not emp.get("دعم"):
            emp["دعم"] = "لا"
        result[eid] = emp

    if not result:
        return None

    return dict(sorted(result.items(), key=lambda item: (
        str(item[1].get("المهمة", "")),
        str(item[1].get("المدرسة", "")),
        str(item[1].get("الاسم", "")),
    )))


def required_people_source_label(date_str):
    return "قائمة مطلوبات اليوم" if required_people_for_date(date_str) is not None else "جدول الأقسام"

def find_today_row(data, today, emp_id):
    for i,row in enumerate(data):
        if str(row.get("التاريخ","")).strip()==str(today).strip() and \
           str(row.get("الرقم الشخصي","")).strip()==str(emp_id).strip():
            return i+2, row
    return None, None


def get_sheet_data_fresh():
    """قراءة مباشرة من sheet1 بدون كاش لاستخدامها في منع التكرار والتنظيف."""
    try:
        return sheet.get_all_records()
    except Exception:
        return []

def find_today_row_fresh(today, emp_id):
    data = get_sheet_data_fresh()
    return find_today_row(data, today, emp_id)

def find_duplicate_attendance_groups():
    """يرجع المجموعات المكررة حسب التاريخ + الرقم الشخصي في sheet1."""
    records = get_sheet_data_fresh()
    groups = {}
    for i, r in enumerate(records, start=2):
        date_val = str(r.get("التاريخ", "")).strip()
        emp_id_val = str(r.get("الرقم الشخصي", "")).strip()
        if not date_val or not emp_id_val:
            continue
        key = (date_val, emp_id_val)
        groups.setdefault(key, []).append((i, r))
    return {k: v for k, v in groups.items() if len(v) > 1}



def find_daily_rows_fresh(today, emp_id):
    """يرجع كل صفوف نفس الرقم الشخصي في نفس التاريخ من sheet1 مباشرة بدون كاش."""
    records = get_sheet_data_fresh()
    matches = []
    for i, r in enumerate(records, start=2):
        if str(r.get("التاريخ", "")).strip() == str(today).strip() and str(r.get("الرقم الشخصي", "")).strip() == str(emp_id).strip():
            matches.append((i, r))
    return matches

def pick_main_daily_row(matches):
    """اختيار السجل الأساسي عند وجود أكثر من صف: نفضّل أول صف فيه حضور، وإلا أول صف."""
    if not matches:
        return None, None
    with_attendance = [(i, r) for i, r in matches if str(r.get("وقت الحضور", "")).strip()]
    return (with_attendance[0] if with_attendance else matches[0])

def has_duplicate_daily_record(today, emp_id):
    return len(find_daily_rows_fresh(today, emp_id)) > 1

def parse_bahrain_datetime(dt_text):
    """تحويل التاريخ المحفوظ في الشيت إلى وقت البحرين حتى لا يفشل تجاوز الموقع."""
    dt_text = str(dt_text or "").strip()
    if not dt_text:
        return None
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"]:
        try:
            return datetime.strptime(dt_text, fmt).replace(tzinfo=BAHRAIN_TZ)
        except Exception:
            pass
    return None

def get_location_override():
    """يرجع True إذا كان الأدمن فعّل تجاوز الموقع وما زال الوقت ساريًا."""
    try:
        records = get_settings_records()
        for r in records:
            key = str(r.get("المفتاح", "")).strip()
            if key == "location_override":
                val = str(r.get("القيمة", "")).strip().lower()
                end_dt = parse_bahrain_datetime(r.get("تاريخ_الانتهاء", ""))

                if val in ["true", "1", "yes", "نعم"] and end_dt:
                    if now_bh() < end_dt:
                        return True, end_dt
                    else:
                        # انتهى الوقت، نطفئه تلقائياً حتى يظهر للأدمن أنه غير مفعّل
                        try:
                            disable_location_override()
                        except Exception:
                            pass
                        return False, None
    except Exception:
        pass
    return False, None

def set_location_override(minutes, note=""):
    """تفعيل تجاوز الموقع لمدة محددة وحفظه في شيت إعدادات_النظام."""
    end_dt = now_bh() + timedelta(minutes=int(minutes))
    end_str = end_dt.strftime("%Y-%m-%d %H:%M:%S")
    try:
        records = settings_sheet.get_all_records()
        row_found = None
        for i, r in enumerate(records):
            if str(r.get("المفتاح", "")).strip() == "location_override":
                row_found = i + 2
                break

        if row_found:
            settings_sheet.update(f"A{row_found}:D{row_found}", [["location_override", "true", end_str, note]])
        else:
            safe_append(settings_sheet, ["location_override", "true", end_str, note])

        get_settings_records.clear()
        return True, end_dt
    except Exception as e:
        st.error(f"❌ خطأ أثناء تفعيل تجاوز الموقع: {e}")
        return False, None

def disable_location_override():
    """إيقاف تجاوز الموقع من شيت إعدادات_النظام."""
    try:
        records = settings_sheet.get_all_records()
        found = False
        for i, r in enumerate(records):
            if str(r.get("المفتاح", "")).strip() == "location_override":
                row_num = i + 2
                settings_sheet.update(f"A{row_num}:D{row_num}", [["location_override", "false", "", "تم الإيقاف"]])
                found = True
                break

        if not found:
            safe_append(settings_sheet, ["location_override", "false", "", "تم الإيقاف"] )

        get_settings_records.clear()
        return True
    except Exception as e:
        st.error(f"❌ خطأ أثناء تعطيل تجاوز الموقع: {e}")
        return False

def log_audit(emp_id, emp_name, operation, details):
    now=now_bh(); fp=get_device_fingerprint()
    safe_append(audit_sheet,[now.strftime("%Y-%m-%d"),now.strftime("%H:%M:%S"),emp_name,str(emp_id),operation,details,fp])


def get_device_lock_global_override():
    """تعطيل قفل الجهاز للجميع لمدة محددة من إعدادات_النظام."""
    try:
        for r in get_settings_records():
            key = str(r.get("المفتاح", "")).strip()
            if key == "device_lock_override":
                val = str(r.get("القيمة", "")).strip().lower()
                end_dt = parse_bahrain_datetime(r.get("تاريخ_الانتهاء", ""))
                if val in ["true", "1", "yes", "نعم"] and end_dt and now_bh() < end_dt:
                    return True, end_dt
                if val in ["true", "1", "yes", "نعم"] and end_dt and now_bh() >= end_dt:
                    disable_device_lock_global_override()
                    return False, None
    except Exception:
        pass
    return False, None

def set_device_lock_global_override(minutes, note=""):
    end_dt = now_bh() + timedelta(minutes=int(minutes))
    end_str = end_dt.strftime("%Y-%m-%d %H:%M:%S")
    try:
        records = settings_sheet.get_all_records()
        row_found = None
        for i, r in enumerate(records):
            if str(r.get("المفتاح", "")).strip() == "device_lock_override":
                row_found = i + 2
                break
        if row_found:
            settings_sheet.update(f"A{row_found}:D{row_found}", [["device_lock_override", "true", end_str, note]])
        else:
            safe_append(settings_sheet, ["device_lock_override", "true", end_str, note])
        get_settings_records.clear()
        return True, end_dt
    except Exception as e:
        st.error(f"❌ خطأ أثناء تعطيل قفل الجهاز للجميع: {e}")
        return False, None

def disable_device_lock_global_override():
    try:
        records = settings_sheet.get_all_records()
        found = False
        for i, r in enumerate(records):
            if str(r.get("المفتاح", "")).strip() == "device_lock_override":
                settings_sheet.update(f"A{i+2}:D{i+2}", [["device_lock_override", "false", "", "تم الإيقاف"]])
                found = True
                break
        if not found:
            safe_append(settings_sheet, ["device_lock_override", "false", "", "تم الإيقاف"])
        get_settings_records.clear()
        return True
    except Exception as e:
        st.error(f"❌ خطأ أثناء تفعيل قفل الجهاز: {e}")
        return False

def employee_has_device_exception(emp_id):
    emp_id = str(emp_id or "").strip()
    if not emp_id:
        return False, None
    try:
        for r in get_device_exceptions():
            if str(r.get("الرقم الشخصي", "")).strip() != emp_id:
                continue
            if not is_yes(r.get("نشط", "")):
                continue
            end_dt = parse_bahrain_datetime(r.get("تاريخ_الانتهاء", ""))
            if end_dt and now_bh() < end_dt:
                return True, end_dt
            if end_dt and now_bh() >= end_dt:
                # الاستثناء منتهي، لا نعتمده
                continue
    except Exception:
        pass
    return False, None

def add_device_exception_for_employee(emp_id, emp_name, minutes, note=""):
    end_dt = now_bh() + timedelta(minutes=int(minutes))
    try:
        safe_append(device_exceptions_sheet, [
            str(emp_id).strip(),
            str(emp_name or "").strip(),
            end_dt.strftime("%Y-%m-%d %H:%M:%S"),
            "نعم",
            note,
            now_bh().strftime("%Y-%m-%d %H:%M:%S")
        ])
        get_device_exceptions.clear()
        return True, end_dt
    except Exception as e:
        st.error(f"❌ خطأ أثناء إضافة الاستثناء: {e}")
        return False, None

def disable_device_exception_for_employee(emp_id):
    emp_id = str(emp_id or "").strip()
    try:
        records = device_exceptions_sheet.get_all_records()
        changed = 0
        for i, r in enumerate(records):
            if str(r.get("الرقم الشخصي", "")).strip() == emp_id and is_yes(r.get("نشط", "")):
                device_exceptions_sheet.update_cell(i+2, 4, "لا")
                changed += 1
        get_device_exceptions.clear()
        return changed
    except Exception as e:
        st.error(f"❌ تعذر تعطيل الاستثناء: {e}")
        return 0

def auto_cleanup_duplicate_attendance_for_emp(today, emp_id, reason="تنظيف تلقائي"):
    """تنظيف ذكي لتكرار نفس الموظفة في نفس التاريخ.
    يحتفظ بالأكمل، ثم الأقدم حضوراً، ويحذف الباقي. يرجع عدد الصفوف المحذوفة.
    """
    matches = find_daily_rows_fresh(today, emp_id)
    if len(matches) <= 1:
        return 0

    def score(item):
        row_num, r = item
        has_att = 1 if str(r.get("وقت الحضور", "")).strip() else 0
        has_dep = 1 if str(r.get("وقت الانصراف", "")).strip() else 0
        has_exit = 1 if str(r.get("خروج استئذان", "")).strip() else 0
        has_return = 1 if str(r.get("عودة", "")).strip() else 0
        att_time = str(r.get("وقت الحضور", "") or "99:99:99")
        # نفضل السجل الأكمل، ثم أقدم حضور، ثم أقل رقم صف
        return (has_att + has_dep + has_exit + has_return, has_dep, has_att, -row_num, att_time)

    # اختيار السجل الذي نبقيه: الأكمل، وإذا تساووا نحتفظ بالأقدم في الشيت
    keep_item = sorted(matches, key=lambda x: (- (1 if str(x[1].get("وقت الحضور", "")).strip() else 0)
                                               - (1 if str(x[1].get("وقت الانصراف", "")).strip() else 0)
                                               - (1 if str(x[1].get("خروج استئذان", "")).strip() else 0)
                                               - (1 if str(x[1].get("عودة", "")).strip() else 0), x[0]))[0]
    keep_row_num = keep_item[0]
    delete_rows = sorted([rn for rn, _ in matches if rn != keep_row_num], reverse=True)
    emp_name = keep_item[1].get("الاسم الثلاثي", "") or keep_item[1].get("الاسم", "")
    for rn in delete_rows:
        try:
            sheet.delete_rows(rn)
        except Exception:
            pass
    if delete_rows:
        log_audit(emp_id, emp_name, reason, f"التاريخ:{today}|تم الاحتفاظ بالصف:{keep_row_num}|حذف الصفوف:{delete_rows}")
        clear_caches()
    return len(delete_rows)

def is_current_device_trusted():
    """يتحقق هل بصمة المتصفح الحالية معتمدة كجهاز موثوق في المركز."""
    fp = get_device_fingerprint()
    try:
        for r in get_trusted_devices():
            if str(r.get("بصمة الجهاز", "")).strip() == fp and is_yes(r.get("نشط", "")):
                return True, r
    except Exception:
        pass
    return False, None

def approve_current_device_as_trusted(device_name, note=""):
    """اعتماد الجهاز الحالي كجهاز موثوق يسمح بتسجيل أكثر من موظفة."""
    fp = get_device_fingerprint()
    now_txt = now_bh().strftime("%Y-%m-%d %H:%M:%S")
    try:
        records = trusted_devices_sheet.get_all_records()
        found = None
        for i, r in enumerate(records, start=2):
            if str(r.get("بصمة الجهاز", "")).strip() == fp:
                found = i
                break
        row = [fp, device_name or "جهاز موثوق", "نعم", note, now_txt, now_txt]
        if found:
            trusted_devices_sheet.update(f"A{found}:F{found}", [row], value_input_option="USER_ENTERED")
        else:
            trusted_devices_sheet.append_row(row, value_input_option="USER_ENTERED")
        get_trusted_devices.clear()
        return True
    except Exception as e:
        st.error(f"❌ تعذر اعتماد الجهاز: {e}")
        return False

def disable_trusted_device(fp):
    try:
        records = trusted_devices_sheet.get_all_records()
        for i, r in enumerate(records, start=2):
            if str(r.get("بصمة الجهاز", "")).strip() == str(fp).strip():
                trusted_devices_sheet.update_cell(i, 3, "لا")
                get_trusted_devices.clear()
                return True
    except Exception as e:
        st.error(f"❌ تعذر تعطيل الجهاز: {e}")
    return False

def touch_trusted_device_usage():
    fp = get_device_fingerprint()
    try:
        records = trusted_devices_sheet.get_all_records()
        for i, r in enumerate(records, start=2):
            if str(r.get("بصمة الجهاز", "")).strip() == fp and is_yes(r.get("نشط", "")):
                trusted_devices_sheet.update_cell(i, 6, now_bh().strftime("%Y-%m-%d %H:%M:%S"))
                get_trusted_devices.clear()
                return
    except Exception:
        pass

def check_device_lock(today, emp_id, emp_name):
    """قفل الجهاز طوال اليوم: نفس بصمة المتصفح لا تسجل لأكثر من رقم في نفس التاريخ.
    يمكن للأدمن تعطيله للجميع أو استثناء رقم شخصي محدد.
    """
    # تعطيل عام مؤقت من الأدمن
    global_off, _ = get_device_lock_global_override()
    if global_off:
        return True

    # استثناء شخصي مؤقت من الأدمن
    emp_off, _ = employee_has_device_exception(emp_id)
    if emp_off:
        return True

    # الجهاز الموثوق يسمح بتسجيل أكثر من موظفة من نفس الجهاز الاحتياطي داخل المركز
    trusted, _trusted_row = is_current_device_trusted()
    if trusted:
        touch_trusted_device_usage()
        return True

    fp = get_device_fingerprint()
    locks = get_device_locks()
    for r in locks:
        if str(r.get("التاريخ", "")).strip() == today and str(r.get("بصمة الجهاز", "")).strip() == fp:
            locked_id = str(r.get("الرقم الشخصي", "")).strip()
            locked_name = str(r.get("الاسم", "")).strip()
            if locked_id and locked_id != str(emp_id).strip():
                safe_append(attempts_sheet, [
                    today, fp, locked_id, locked_name, emp_id, emp_name,
                    now_bh().strftime("%H:%M:%S"),
                    "محاولة تسجيل رقم آخر من نفس الجهاز — قفل اليوم الكامل"
                ])
                try:
                    data = get_sheet_data_fresh()
                    row_index, _ = find_today_row(data, today, locked_id)
                    if row_index:
                        safe_update(sheet, row_index, COL_ATTEMPT, "⚠️ محاولة تسجيل باسم آخر من نفس الجهاز")
                except Exception:
                    pass
                st.error("🚫 هذا الجهاز مسجّل اليوم باسم موظفة أخرى، ولا يمكن التسجيل برقم مختلف من نفس الجهاز.")
                st.warning("يمكن للأدمن فتح قفل الجهاز، أو تعطيل قفل الجهاز مؤقتًا للجميع، أو إضافة استثناء لهذا الرقم عند وجود سبب رسمي.")
                return False
    return True

def lock_device(today, emp_id, emp_name):
    fp=get_device_fingerprint(); locks=get_device_locks()
    for r in locks:
        if str(r.get("التاريخ","")).strip()==today and \
           str(r.get("بصمة الجهاز","")).strip()==fp and \
           str(r.get("الرقم الشخصي","")).strip()==str(emp_id).strip(): return True
    ok=safe_append(device_sheet,[today,fp,str(emp_id),emp_name,now_bh().strftime("%H:%M:%S")])
    get_device_locks.clear(); return ok

def register_operation(operation, emp_id, note=""):
    override_active,_=get_location_override()
    no_gps_allowed = bool(st.session_state.get("allow_no_gps_today", False))
    gps_status_note = ""
    if not st.session_state.get("location_allowed",False) and not override_active:
        if no_gps_allowed:
            gps_status_note = " | ⚠️ بدون تحقق GPS"
        else:
            st.error("❌ يجب تحديد الموقع أولاً، أو اختيار تعذر التحقق من الموقع، أو تفعيل تجاوز الموقع من الأدمن."); return False

    emp_id=ar_to_en_digits(emp_id).strip()
    emp=validate_employee(emp_id) or st.session_state.get("emp_data")
    if not emp or str(emp.get("الرقم الشخصي",emp_id)).strip()!=emp_id:
        st.error("❌ بيانات الموظفة غير مكتملة."); return False

    # حفظ موظفة جديدة في القائمة البيضاء
    if not validate_employee(emp_id) and operation=="تسجيل حضور":
        is_support_new="دعم" in str(emp.get("المهمة",""))
        if not is_support_new and not emp.get("دعم"):
            try:
                whitelist_sheet.append_row([
                    emp_id,
                    emp.get("الاسم",""),
                    emp.get("المدرسة",""),
                    emp.get("المهمة",""),
                    "لا",
                    emp.get("رقم التواصل",""),
                    emp.get("البريد الإلكتروني",""),
                    emp.get("المسمى الوظيفي","موظفة"),
                    "نعم"
                ])
                get_whitelist.clear()
            except: pass

    full_name=normalize_name(emp.get("الاسم",""))
    school=emp.get("المدرسة",schools[0])
    task=emp.get("المهمة",TASKS_MAIN[0])
    is_support="نعم" if ("دعم" in str(task) or emp.get("دعم")) else "لا"

    now=now_bh(); today=now.strftime("%Y-%m-%d")
    day_name=now.strftime("%A"); time_now=now.strftime("%H:%M:%S")
    if gps_status_note and gps_status_note not in str(note or ""):
        note = (str(note or "").strip() + gps_status_note).strip()

    # تنظيف أي تكرار سابق لنفس الموظفة قبل العملية حتى لا يعتمد النظام آخر ضغطة بالخطأ
    try:
        auto_cleanup_duplicate_attendance_for_emp(today, emp_id, "تنظيف تلقائي قبل التسجيل")
    except Exception:
        pass

    if not check_device_lock(today,emp_id,full_name): return False

    # قراءة مباشرة بدون كاش حتى لا تتكرر العمليات بسبب تأخر تحديث Google Sheet
    daily_matches = find_daily_rows_fresh(today, emp_id)
    row_index, row = pick_main_daily_row(daily_matches)

    if len(daily_matches) > 1:
        st.warning("⚠️ يوجد أكثر من سجل لنفس الموظفة في نفس التاريخ. يرجى تنظيف التكرارات من لوحة الأدمن. لن يتم إنشاء سجل جديد.")

    if operation=="تسجيل حضور":
        # حماية مباشرة من التكرار: أي سجل لنفس الرقم في نفس التاريخ يمنع إنشاء سجل حضور جديد
        daily_matches = find_daily_rows_fresh(today, emp_id)
        row_index, row = pick_main_daily_row(daily_matches)
        if row and str(row.get("وقت الحضور", "")).strip():
            st.error(f"❌ تم تسجيل حضورك مسبقاً لهذا اليوم الساعة {row.get('وقت الحضور','')}. لا يمكن تسجيل حضور مكرر.")
            return False
        if len(daily_matches) > 1:
            st.error("❌ يوجد تكرار سابق لهذا الرقم في نفس التاريخ. افتحي لوحة الأدمن > تنظيف التكرارات، ثم حاولي مرة أخرى.")
            return False

        # إذا حضرت متأخرة بسبب موعد/سبب آخر، يوثق النظام ذلك كاستئذان ضمني:
        # خروج الاستئذان من 7:00، ووقت الحضور هو العودة.
        att_is_late = now.time() > time(7, 5, 30)
        note_txt = str(note or "").strip()
        implicit_leave = att_is_late and note_txt and not any(x in note_txt for x in ["رعاية", "دوام مرن", "مهمة رسمية"])
        implicit_exit_time = "07:00:00" if implicit_leave else ""
        implicit_return_time = time_now if implicit_leave else ""

        if row_index:
            safe_update(sheet,row_index,COL_ATTEND,time_now)
            safe_update(sheet,row_index,COL_LATE_REASON,note)
            if implicit_leave:
                safe_update(sheet,row_index,COL_EXIT,implicit_exit_time)
                safe_update(sheet,row_index,COL_RETURN,implicit_return_time)
        else:
            # فحص أخير مباشرة قبل الإضافة لمنع التكرار عند الضغط السريع أو فتح أكثر من نافذة
            final_matches = find_daily_rows_fresh(today, emp_id)
            final_row_index, final_row = pick_main_daily_row(final_matches)
            if final_row and str(final_row.get("وقت الحضور", "")).strip():
                st.error(f"❌ تم تسجيل حضورك مسبقاً لهذا اليوم الساعة {final_row.get('وقت الحضور','')}. لا يمكن تسجيل حضور مكرر.")
                return False
            if final_row_index:
                safe_update(sheet,final_row_index,COL_ATTEND,time_now)
                safe_update(sheet,final_row_index,COL_LATE_REASON,note)
                if implicit_leave:
                    safe_update(sheet,final_row_index,COL_EXIT,implicit_exit_time)
                    safe_update(sheet,final_row_index,COL_RETURN,implicit_return_time)
                row_index = final_row_index
            else:
                ok=safe_append(sheet,[today,day_name,school,task,is_support,full_name,emp_id,time_now,note,"","",implicit_exit_time,implicit_return_time,"","","","","","","","","","دعم مباشر" if is_support == "نعم" else "تسجيل ذاتي"])
                if not ok: st.error("❌ تعذر الحفظ، حاولي بعد قليل."); return False
        lock_device(today,emp_id,full_name)
        log_audit(emp_id,full_name,"تسجيل حضور",f"الوقت:{time_now}|السبب:{note or 'بدون'}" + ("|جهاز موثوق" if is_current_device_trusted()[0] else ""))
        # ثبّت البيانات
        st.session_state.data_locked_today=True
        st.session_state.locked_emp={"الرقم الشخصي":emp_id,"الاسم":full_name,"المدرسة":school,"المهمة":task,"دعم":is_support=="نعم","نشط":"نعم"}
        st.session_state.locked_date=today
        ls_set("saved_date",today,"sv_date"); ls_set("saved_id",emp_id,"sv_id")
        ls_set("saved_name",full_name,"sv_name"); ls_set("saved_school",school,"sv_school")
        ls_set("saved_section",task,"sv_section"); ls_set("saved_support",is_support,"sv_support")
        try:
            deleted_count = auto_cleanup_duplicate_attendance_for_emp(today, emp_id, "تنظيف تلقائي بعد تسجيل الحضور")
            if deleted_count:
                st.info(f"🧹 تم تنظيف {deleted_count} سجل مكرر تلقائيًا لهذا اليوم.")
        except Exception:
            pass
        clear_caches()
        st.session_state.force_fresh_today_row = True
        st.session_state.attendance_done_now = time_now
        st.success(f"✅ تم تسجيل الحضور بنجاح الساعة {time_now}")
        return True

    if operation=="تسجيل انصراف":
        if not row_index or not row or not row.get("وقت الحضور"): st.error("❌ يجب تسجيل الحضور أولاً."); return False
        if row.get("وقت الانصراف"): st.error("❌ تم تسجيل الانصراف مسبقاً."); return False
        safe_update(sheet,row_index,COL_DEPART,time_now); safe_update(sheet,row_index,COL_DEPART_REASON,note)
        row["وقت الانصراف"] = time_now
        row["سبب الانصراف"] = note
        update_work_calculation(row_index, row)
        log_audit(emp_id,full_name,"تسجيل انصراف",f"الوقت:{time_now}|السبب:{note or 'بدون'}")
        clear_caches(); st.success("✅ تم تسجيل الانصراف بنجاح."); return True

    if operation=="خروج استئذان":
        if not row_index or not row or not row.get("وقت الحضور"): st.error("❌ يجب تسجيل الحضور أولاً."); return False
        if row.get("خروج استئذان") and not row.get("عودة"): st.error("❌ يوجد خروج استئذان مفتوح."); return False
        if row.get("خروج استئذان"): st.error("❌ تم تسجيل خروج الاستئذان مسبقاً."); return False
        safe_update(sheet,row_index,COL_EXIT,time_now); safe_update(sheet,row_index,COL_DEPART_REASON,note)
        if "استئذان انصراف" in str(note):
            safe_update(sheet,row_index,COL_DEPART,time_now)
            row["خروج استئذان"] = time_now
            row["وقت الانصراف"] = time_now
            row["سبب الانصراف"] = note
            update_work_calculation(row_index, row)
            log_audit(emp_id,full_name,"استئذان انصراف",f"الوقت:{time_now}|السبب:{note}")
            clear_caches(); st.success("✅ تم تسجيل استئذان انصراف بنجاح. تم احتساب وقت الاستئذان كوقت انصراف."); return True
        log_audit(emp_id,full_name,"خروج استئذان",f"الوقت:{time_now}|السبب:{note}")
        clear_caches(); st.success("✅ تم تسجيل خروج الاستئذان بنجاح. عند الرجوع اضغطي زر عودة من استئذان."); return True

    if operation=="عودة من استئذان":
        if not row_index or not row or not row.get("وقت الحضور"): st.error("❌ يجب تسجيل الحضور أولاً."); return False
        if not row.get("خروج استئذان"): st.error("❌ لا يوجد خروج استئذان مفتوح."); return False
        if row.get("عودة"): st.error("❌ تم تسجيل العودة مسبقاً."); return False
        safe_update(sheet,row_index,COL_RETURN,time_now)
        log_audit(emp_id,full_name,"عودة من استئذان",f"الوقت:{time_now}")
        clear_caches(); st.success("✅ تم تسجيل العودة من الاستئذان بنجاح."); return True

    st.error("❌ عملية غير معروفة."); return False


# ─── طلبات التسجيل اليدوي / مشاكل المتصفح والموقع ─────────────────────
def submit_manual_request(emp_id, emp_name, school, task, req_type, actual_att="", actual_dep="", problem_type="", notes=""):
    """إرسال طلب تسجيل يدوي للأدمن.
    الوقت يعتمد تلقائيًا على وقت إرسال الطلب، ولا يُطلب من الموظفة إدخال وقت يدوي.
    يمنع تكرار نفس نوع الطلب لنفس الرقم في نفس اليوم.
    """
    try:
        emp_id = ar_to_en_digits(emp_id).strip()
        req_type = str(req_type or "").strip()
        exists, old_req = manual_request_exists_today(emp_id, req_type)
        if exists:
            old_time = str(old_req.get("وقت الطلب", "") or "")
            st.warning(f"⚠️ تم إرسال طلب {req_type} سابق لهذا اليوم{(' الساعة ' + old_time) if old_time else ''}. لا يمكن تكرار نفس الطلب.")
            return "duplicate"

        now = now_bh()
        request_time = now.strftime("%H:%M:%S")
        fp = get_device_fingerprint()

        # نحفظ وقت الإرسال في خانة الحضور أو الانصراف حسب نوع الطلب، للتوثيق فقط.
        request_att = request_time if req_type == "حضور" else ""
        request_dep = request_time if req_type == "انصراف" else ""

        row = [
            now.strftime("%Y-%m-%d"),
            request_time,
            emp_id,
            normalize_name(emp_name),
            str(school or "").strip(),
            str(task or "").strip(),
            req_type,
            request_att,
            request_dep,
            str(problem_type or "").strip(),
            str(notes or "").strip(),
            "بانتظار الاعتماد",
            fp,
            "",
            ""
        ]
        ok = safe_append(manual_requests_sheet, row)
        get_manual_requests.clear()
        return True if ok else False
    except Exception:
        return False

def approve_manual_request(req_row_num, req, approve_type="حضور", use_actual_time=True):
    eid = str(req.get("الرقم الشخصي", "")).strip()
    if not eid:
        st.error("❌ لا يوجد رقم شخصي في الطلب.")
        return False
    emp = validate_employee(eid) or {}
    full_name = normalize_name(req.get("الاسم", "") or emp.get("الاسم", ""))
    school = req.get("المدرسة", "") or emp.get("المدرسة", "")
    task = req.get("المهمة", "") or emp.get("المهمة", "")
    support_value = "نعم" if ("دعم" in str(task) or is_yes(emp.get("دعم", ""))) else "لا"
    date_str = str(req.get("تاريخ الطلب", "") or now_bh().strftime("%Y-%m-%d"))
    try:
        d_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        day_name = d_obj.strftime("%A")
    except Exception:
        day_name = day_arabic
    request_time = str(req.get("وقت الطلب", "") or now_bh().strftime("%H:%M:%S"))
    # الاعتماد يكون دائمًا حسب وقت إرسال الطلب، وليس وقتًا تكتبه الموظفة يدويًا.
    att_time = request_time
    dep_time = request_time
    problem = str(req.get("نوع المشكلة", "") or "مشكلة تقنية").strip()
    notes = str(req.get("ملاحظات", "") or "").strip()
    reason_note = f"[طلب يدوي] {problem} | وقت إرسال الطلب: {request_time} | {notes}".strip()
    existing_matches = find_daily_rows_fresh(date_str, eid)
    existing_idx, existing_row = pick_main_daily_row(existing_matches)
    if approve_type == "حضور":
        if existing_idx:
            safe_update(sheet, existing_idx, COL_ATTEND, att_time)
            safe_update(sheet, existing_idx, COL_LATE_REASON, reason_note)
            existing_row["وقت الحضور"] = att_time
            existing_row["سبب التأخير"] = reason_note
            update_work_calculation(existing_idx, existing_row)
        else:
            safe_append(sheet, [date_str, day_name, school, task, support_value, full_name, eid, att_time, reason_note, "", "", "", "", "⚠️ طلب يدوي بدون تحقق GPS", "", "", "", "", "", "", "", "", "طلب يدوي معتمد" if support_value != "نعم" else "طلب دعم معتمد"])
            idx_after, row_after = find_today_row_fresh(date_str, eid)
            if idx_after:
                update_work_calculation(idx_after, row_after)
        log_audit(eid, full_name, "اعتماد طلب حضور يدوي", f"التاريخ:{date_str}|حضور:{att_time}|السبب:{reason_note}")
    elif approve_type == "انصراف":
        if not existing_idx:
            st.error("❌ لا يوجد سجل حضور لاعتماد الانصراف.")
            return False
        dep_to_use = dep_time or request_time
        safe_update(sheet, existing_idx, COL_DEPART, dep_to_use)
        safe_update(sheet, existing_idx, COL_DEPART_REASON, reason_note)
        existing_row["وقت الانصراف"] = dep_to_use
        existing_row["سبب الانصراف"] = reason_note
        update_work_calculation(existing_idx, existing_row)
        log_audit(eid, full_name, "اعتماد طلب انصراف يدوي", f"التاريخ:{date_str}|انصراف:{dep_to_use}|السبب:{reason_note}")
    else:
        return False
    try:
        manual_requests_sheet.update_cell(req_row_num, 12, "تم الاعتماد")
        manual_requests_sheet.update_cell(req_row_num, 14, now_bh().strftime("%Y-%m-%d %H:%M:%S"))
        manual_requests_sheet.update_cell(req_row_num, 15, "أدمن")
    except Exception:
        pass
    clear_caches()
    # ── إعادة حساب الساعات بعد الاعتماد بقراءة مباشرة من الشيت ──
    try:
        fresh_matches = find_daily_rows_fresh(date_str, eid)
        fresh_idx, fresh_row = pick_main_daily_row(fresh_matches)
        if fresh_idx and fresh_row:
            update_work_calculation(fresh_idx, fresh_row)
    except Exception:
        pass
    return True


# ─── وضع الطوارئ الخفيف لنفس الرابط ─────────────────────────────
def _get_query_param_value(name, default=""):
    try:
        val = st.query_params.get(name, default)
        if isinstance(val, list):
            return val[0] if val else default
        return val
    except Exception:
        try:
            val = st.experimental_get_query_params().get(name, [default])
            return val[0] if isinstance(val, list) and val else val
        except Exception:
            return default


def is_lite_emergency_mode():
    mode_val = str(_get_query_param_value("lite", "")).strip().lower()
    emergency_val = str(_get_query_param_value("emergency", "")).strip().lower()
    return mode_val in ["1", "true", "yes", "نعم"] or emergency_val in ["1", "true", "yes", "نعم"]


def render_lite_emergency_mode():
    """واجهة خفيفة جدًا للأجهزة القديمة: بدون GPS وبدون LocalStorage وبدون عناصر ثقيلة."""
    st.markdown("""
    <style>
    html,body,[class*="css"]{direction:rtl!important;text-align:right!important;font-family:Tahoma,Arial,sans-serif!important;}
    .block-container{max-width:560px;padding-top:1rem;padding-bottom:2rem;}
    .lite-title{background:#0c3460;color:white;border-radius:18px;padding:18px;text-align:center!important;font-size:22px;font-weight:800;margin-bottom:14px;}
    .lite-card{border:1px solid #ddd;border-radius:16px;padding:14px;background:#fff;margin-bottom:12px;}
    .lite-note{background:#fff7e6;border-right:4px solid #BA7517;border-radius:12px;padding:10px;color:#5f3908;font-weight:700;margin-bottom:12px;}
    label{font-weight:700!important;color:#0c3460!important;}
    .stButton button{border-radius:14px!important;font-weight:800!important;}
    </style>
    """, unsafe_allow_html=True)
    st.markdown('<div class="lite-title">🆘 طلب تسجيل يدوي — وضع الطوارئ الخفيف</div>', unsafe_allow_html=True)
    st.markdown('<div class="lite-note">هذه الصفحة مخصصة للأجهزة القديمة أو عند ظهور صفحة بيضاء في النظام الكامل. الطلب لا يُسجل مباشرة إلا بعد اعتماد الأدمن.</div>', unsafe_allow_html=True)

    with st.container(border=True):
        emp_id = ar_to_en_digits(st.text_input("الرقم الشخصي", max_chars=20, key="lite_emp_id")).strip()
        emp = validate_employee(emp_id) if emp_id else None

        if emp:
            # لا نعرض الاسم تلقائيًا للموظفة في وضع الطوارئ؛ نستخدمه داخليًا للأدمن فقط.
            emp_name = str(emp.get("الاسم", "")).strip()
            emp_school = str(emp.get("المدرسة", "")).strip()
            emp_task = str(emp.get("المهمة", "")).strip()
            st.success("✅ تم التعرف على الرقم من القائمة البيضاء. أكملي نوع الطلب والوقت فقط.")
        else:
            if emp_id:
                st.warning("⚠️ الرقم غير موجود في القائمة البيضاء، اكتبي البيانات ليتمكن الأدمن من مراجعة الطلب.")
            emp_name = st.text_input("الاسم الثلاثي", key="lite_emp_name")
            school_choice = st.selectbox("المدرسة", schools + ["أخرى"], key="lite_school_choice")
            if school_choice == "أخرى":
                emp_school = st.text_input("اكتبي اسم المدرسة", key="lite_school_other").strip()
            else:
                emp_school = school_choice
            emp_task = st.selectbox("المهمة", TASKS_ALL, key="lite_task")

        req_type = st.selectbox("نوع الطلب", ["حضور", "انصراف"], key="lite_req_type")
        actual_att = st.text_input("وقت الحضور الفعلي", value="07:00:00", key="lite_actual_att")
        actual_dep = st.text_input("وقت الانصراف الفعلي (اختياري)", key="lite_actual_dep")
        problem_type = st.selectbox("نوع المشكلة", ["صفحة بيضاء", "تعذر تحديد الموقع", "الموقع لا يعمل", "زر لا يستجيب", "مشكلة في المتصفح", "أخرى"], key="lite_problem_type")
        notes = st.text_area("ملاحظات اختيارية", key="lite_notes")

        if st.button("📨 إرسال الطلب للأدمن", use_container_width=True, type="primary", key="lite_submit_request"):
            if not emp_id:
                st.error("❌ الرقم الشخصي مطلوب.")
            elif not str(emp_name).strip() or not str(emp_school).strip() or not str(emp_task).strip():
                st.error("❌ الاسم والمدرسة والمهمة مطلوبة إذا لم يكن الرقم موجودًا في القائمة البيضاء.")
            elif not actual_att.strip() and req_type == "حضور":
                st.error("❌ وقت الحضور الفعلي مطلوب.")
            else:
                ok = submit_manual_request(emp_id, emp_name, emp_school, emp_task, req_type, actual_att, actual_dep, problem_type + " — وضع الطوارئ الخفيف", notes)
                if ok:
                    st.success("✅ تم إرسال الطلب للأدمن. لا تضغطي مرة ثانية.")
                    st.info("سيظهر طلبك في لوحة الأدمن ليتم اعتماده يدويًا.")
                else:
                    st.error("❌ تعذر إرسال الطلب. تواصلي مع الأدمن عبر واتساب.")

    st.markdown("---")
    st.markdown('[الرجوع للنظام الكامل](?lite=0)')


if is_lite_emergency_mode():
    render_lite_emergency_mode()
    st.stop()

# ─── Session State ──────────────────────────────────────────────
default_state={
    "pending_operation":None,"admin_logged_in":False,"admin_last_active":None,
    "location_allowed":False,"emp_verified":False,"emp_data":None,
    "data_locked_today":False,"locked_emp":None,"locked_date":None,"operation_saving":False,
    "location_check_requested":False,
    "allow_no_gps_today":False,
    "no_gps_option_available":False,
    "emp_step":"login",
    "nogps_saving":False,
    "_queued_op":"",
    "_queued_note":"",
    "_loc_step_initialized": False,
    "_emp_session_active": False,
    "_trusted_cleared": False,
    "edit_results": None,
}
for k,v in default_state.items():
    if k not in st.session_state: st.session_state[k]=v

today_str=now_bh().strftime("%Y-%m-%d")
auto_close_previous_open_records()

_saved_date=ls_get("saved_date"); _saved_id=ls_get("saved_id")
_saved_name=ls_get("saved_name"); _saved_school=ls_get("saved_school")
_saved_section=ls_get("saved_section"); _saved_support=ls_get("saved_support")
_trusted_cleared_ls = str(ls_get("trusted_cleared") or "").strip() == "yes"

# نظّف القيم الفارغة
_saved_id   = str(_saved_id   or "").strip()
_saved_date = str(_saved_date or "").strip()

# مهم: لا نعتمد على المدرسة/المهمة المحفوظة في المتصفح لأنها قد تكون قديمة.
# إذا كان الرقم محفوظًا لهذا اليوم، نقرأ بياناته مباشرة من القائمة البيضاء في Google Sheet.
if _saved_id:
    fresh_emp = validate_employee_fresh(_saved_id)
    if fresh_emp:
        _saved_name = get_emp_name(fresh_emp) or _saved_name
        _saved_school = get_emp_school(fresh_emp) or _saved_school
        _saved_section = str(fresh_emp.get("المهمة", "") or _saved_section).strip()
        _saved_support = "نعم" if is_yes(fresh_emp.get("دعم", "")) else "لا"

# الجهاز الموثوق لا يستخدم التثبيت التلقائي — كل موظفة تدخل رقمها من جديد
_is_trusted_device = is_current_device_trusted()[0]

_data_locked=(
    not _is_trusted_device
    and not _trusted_cleared_ls
    and (
        (st.session_state.get("data_locked_today",False) and st.session_state.get("locked_date")==today_str)
        or (_saved_date==today_str and len(_saved_id) > 3)
    )
)

if _data_locked and not st.session_state.emp_verified:
    st.session_state.emp_verified=True
    st.session_state.emp_data=st.session_state.get("locked_emp") or {
        "الرقم الشخصي":_saved_id,"الاسم":str(_saved_name or "").strip(),"المدرسة":str(_saved_school or "").strip(),
        "المهمة":str(_saved_section or "").strip(),"دعم":_saved_support=="نعم","نشط":"نعم"
    }

if st.session_state.admin_logged_in and st.session_state.admin_last_active:
    idle=(now_bh()-st.session_state.admin_last_active).seconds//60
    if idle>=30:
        st.session_state.admin_logged_in=False; st.session_state.admin_last_active=None
        st.warning("⏱️ انتهت جلسة الأدمن بسبب الخمول.")

# ─── Header ────────────────────────────────────────────────────
day_arabic={"Saturday":"السبت","Sunday":"الأحد","Monday":"الاثنين","Tuesday":"الثلاثاء","Wednesday":"الأربعاء","Thursday":"الخميس","Friday":"الجمعة"}.get(now_bh().strftime("%A"),now_bh().strftime("%A"))

st.image("logo.png", use_container_width=True)

st.markdown(f"""
<div class="app-header">
    <div class="sub">مركز مدرسة جدحفص للتصحيح المركزي — المنطقة التعليمية (2)</div>
    <div class="title">نظام الحضور والانصراف</div>
    <div class="date-pill">{day_arabic} — {today_str}</div>
</div>
""", unsafe_allow_html=True)

mode=st.radio("",["👤 موظفة","🛡️ أدمن"],horizontal=True,label_visibility="collapsed")

# ══════════════════════════════════════════════════════════════════
# ══ واجهة الموظفة ══
# ══════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════
# ══ واجهة الموظفة ══
# ══════════════════════════════════════════════════════════════════
if mode=="👤 موظفة":

    # ── شاشة تحميل تمنع الضغط المزدوج ──
    st.markdown("""
    <style>
    .loading-overlay {
        display: none;
        position: fixed;
        top: 0; left: 0;
        width: 100%; height: 100%;
        background: rgba(255,255,255,0.88);
        z-index: 9999;
        justify-content: center;
        align-items: center;
        flex-direction: column;
        font-size: 18px;
        font-weight: 700;
        color: #0c3460;
        direction: rtl;
        cursor: not-allowed;
        pointer-events: all;
    }
    .loading-overlay.show { display: flex; }
    </style>
    <div class="loading-overlay" id="loadingOverlay">
        <div>⏳ جارٍ التحميل… يرجى الانتظار</div>
        <div style="font-size:13px;font-weight:400;margin-top:8px;color:#666;">لا تضغطي مرة أخرى</div>
    </div>
    <script>
    document.addEventListener('click', function(e) {
        var btn = e.target.tagName === 'BUTTON' || e.target.closest('button');
        if (btn) {
            var overlay = document.getElementById('loadingOverlay');
            overlay.classList.add('show');
            // يختفي تلقائياً بعد 5 ثواني كحماية
            setTimeout(function() {
                overlay.classList.remove('show');
            }, 5000);
        }
    }, true);
    </script>
    """, unsafe_allow_html=True)

    # ── تهيئة الـ session state ──
    if "emp_verified" not in st.session_state:
        st.session_state.emp_verified = False
    if "emp_data" not in st.session_state:
        st.session_state.emp_data = None

    # ── مسح البيانات القديمة إذا ما في رقم مدخل في هذه الجلسة ──
    if not st.session_state.get("_emp_session_active"):
        st.session_state.emp_verified        = False
        st.session_state.emp_data            = None
        st.session_state.pending_operation   = None
        st.session_state._queued_op          = ""
        st.session_state._queued_note        = ""
        st.session_state._emp_session_active = True

    trusted = _is_trusted_device
    _trusted_rec = None

    # ══════════════════════════════════
    # كرت 1: البيانات الشخصية
    # ══════════════════════════════════
    with st.container(border=True):
        st.markdown('<div class="card-title">🪪 البيانات الشخصية</div>', unsafe_allow_html=True)

        if _data_locked and st.session_state.emp_data:
            emp = st.session_state.emp_data
            st.markdown(f"""
            <div class="field-lbl">الاسم</div><div class="field-val">{emp.get("الاسم","")}</div>
            <div class="field-lbl">المدرسة</div><div class="field-val">{emp.get("المدرسة","")}</div>
            <div class="field-lbl">المهمة</div><div class="field-val blue">{emp.get("المهمة","")}</div>
            <div style="font-size:12px;color:#3B6D11;font-weight:700;">🔒 بياناتك محفوظة</div>
            """, unsafe_allow_html=True)
            show_previous_auto_close_notice(emp.get("الرقم الشخصي",""))
            show_previous_absence_notice(emp.get("الرقم الشخصي",""))
        else:
            emp_id_raw = st.text_input("الرقم الشخصي", placeholder="أدخلي رقمك الشخصي", max_chars=20, key="main_emp_id")
            emp_id_input = ar_to_en_digits(emp_id_raw).strip()

            # عند تغيير الرقم الشخصي نمسح بيانات الموظفة السابقة من الجلسة حتى لا تظهر مدرسة/مهمة قديمة.
            if st.session_state.get("last_emp_id_input", "") != emp_id_input:
                st.session_state.last_emp_id_input = emp_id_input
                st.session_state.emp_verified = False
                st.session_state.emp_data = None
                st.session_state.pending_operation = None
                st.session_state.location_allowed = False
                st.session_state.location_check_requested = False
                st.session_state.allow_no_gps_today = False

            if emp_id_input:
                # قراءة مباشرة من القائمة البيضاء لضمان أن الاسم/المدرسة/المهمة مثل Google Sheet تمامًا.
                existing = validate_employee_fresh(emp_id_input)
                if existing:
                    is_sup = str(existing.get("دعم","")).strip() in ["نعم","yes","Yes","TRUE","true","1"]
                    is_sup_pending = is_sup

                    if is_sup_pending:
                        st.markdown(f"""
                        <div class="field-lbl">الاسم</div><div class="field-val">{existing.get("الاسم","")}</div>
                        <div style="background:#faeeda;border-radius:12px;padding:10px 14px;font-size:13px;font-weight:700;color:#633806;">
                        🔄 أنتِ مسجّلة كدعم مؤقت
                        </div>
                        """, unsafe_allow_html=True)
                        still_sup = st.radio("ما زلتِ دعم أم صرتِ عضوة؟",
                                             ["🔄 لا زلت دعم","🏫 صرت عضوة في المركز"],
                                             horizontal=True, key="sup_upgrade")
                        if still_sup == "🔄 لا زلت دعم":
                            st.session_state.emp_data = {"الرقم الشخصي":emp_id_input,"الاسم":existing.get("الاسم",""),"المدرسة":existing.get("المدرسة",""),"المهمة":existing.get("المهمة",""),"نشط":"نعم","دعم":True,"_existing":True}
                            st.session_state.emp_verified = True
                            st.warning("🔄 سيتم تسجيل حضورك لهذا اليوم فقط كدعم")
                        else:
                            emp_task_new = st.selectbox("المهمة الجديدة", TASKS_MAIN, key="upgrade_task")
                            emp_job_new  = st.selectbox("المسمى الوظيفي", JOB_TITLES, key="upgrade_job")
                            emp_phone_new= st.text_input("رقم التواصل", value=existing.get("رقم التواصل",""), key="upgrade_phone")
                            emp_email_new= st.text_input("البريد الإلكتروني", value=existing.get("البريد الإلكتروني",""), key="upgrade_email")
                            if st.button("💾 حفظ كعضوة", use_container_width=True, type="primary", key="btn_upgrade"):
                                try:
                                    wl_records = whitelist_sheet.get_all_records()
                                    for i, r in enumerate(wl_records):
                                        if str(r.get("الرقم الشخصي","")).strip() == emp_id_input:
                                            rn = i+2
                                            whitelist_sheet.update_cell(rn,4,emp_task_new)
                                            whitelist_sheet.update_cell(rn,5,"لا")
                                            whitelist_sheet.update_cell(rn,6,emp_phone_new)
                                            whitelist_sheet.update_cell(rn,7,emp_email_new)
                                            whitelist_sheet.update_cell(rn,8,emp_job_new)
                                            break
                                    get_whitelist.clear()
                                    st.session_state.emp_data = {"الرقم الشخصي":emp_id_input,"الاسم":existing.get("الاسم",""),"المدرسة":existing.get("المدرسة",""),"المهمة":emp_task_new,"نشط":"نعم","دعم":False,"_existing":True}
                                    st.session_state.emp_verified = True
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"خطأ: {e}")
                    else:
                        st.session_state.emp_data = {"الرقم الشخصي":emp_id_input,"الاسم":existing.get("الاسم",""),"المدرسة":existing.get("المدرسة",""),"المهمة":existing.get("المهمة",""),"نشط":"نعم","دعم":False,"_existing":True}
                        st.session_state.emp_verified = True
                        ls_set("trusted_cleared","no","clear_trusted_cleared")
                        st.markdown(f"""
                        <div class="field-lbl">الاسم</div><div class="field-val">{existing.get("الاسم","")}</div>
                        <div class="field-lbl">المدرسة</div><div class="field-val">{existing.get("المدرسة","")}</div>
                        <div class="field-lbl">المهمة</div><div class="field-val blue">{existing.get("المهمة","")}</div>
                        """, unsafe_allow_html=True)
                        show_previous_auto_close_notice(emp_id_input)
                        show_previous_absence_notice(emp_id_input)
                else:
                    # موظفة جديدة
                    emp_type = st.radio("نوع التسجيل",["🏫 عضوة في المركز","🔄 دعم"],horizontal=True,key="emp_type")
                    is_sup = emp_type == "🔄 دعم"
                    emp_name = st.text_input("الاسم الثلاثي", placeholder="اكتبي اسمك الثلاثي", key="new_name")
                    if is_sup:
                        sch_choice = st.selectbox("المدرسة", schools+["أخرى"], key="new_school_sup")
                        emp_school = st.text_input("اكتبي اسم المدرسة", key="new_school_other").strip() if sch_choice=="أخرى" else sch_choice
                    else:
                        emp_school = st.selectbox("المدرسة", schools, key="new_school")
                    emp_task = st.selectbox("المهمة", TASKS_SUPPORT if is_sup else TASKS_MAIN, key="new_task")
                    if not is_sup:
                        emp_job   = st.selectbox("المسمى الوظيفي", JOB_TITLES, key="new_job")
                        emp_phone = st.text_input("رقم التواصل", key="new_phone")
                        emp_email = st.text_input("البريد الإلكتروني", key="new_email")
                    else:
                        emp_job="دعم"; emp_phone=""; emp_email=""
                        st.warning("🔄 سيتم تسجيل حضورك اليوم فقط كدعم")

                    if emp_name.strip() and st.button("💾 حفظ البيانات", use_container_width=True, type="primary", key="btn_save_new"):
                        if not str(emp_school).strip():
                            st.error("❌ اسم المدرسة مطلوب.")
                        else:
                            new_data = {"الرقم الشخصي":emp_id_input,"الاسم":normalize_name(emp_name),"المدرسة":emp_school,"المهمة":emp_task,"المسمى الوظيفي":emp_job,"رقم التواصل":emp_phone,"البريد الإلكتروني":emp_email,"نشط":"نعم","دعم":is_sup,"_existing":False}
                            if not is_sup:
                                try:
                                    whitelist_sheet.append_row([emp_id_input,normalize_name(emp_name),emp_school,emp_task,"لا",emp_phone,emp_email,emp_job,"نعم"])
                                    get_whitelist.clear()
                                except Exception as e:
                                    st.warning(f"⚠️ تعذّر الحفظ: {e}")
                            st.session_state.emp_data = new_data
                            st.session_state.emp_verified = True
                            st.rerun()

    # ══════════════════════════════════
    # كرت 2: التحقق من الموقع
    # ══════════════════════════════════
    # بعد ظهور بيانات الموظفة تظهر أيقونة الموقع مباشرة بدون زر إضافي وبدون expander.
    if trusted:
        st.session_state.location_allowed = True
        st.success("✅ جهاز موثوق — تم تجاوز التحقق من الموقع.")

    elif st.session_state.get("location_allowed"):
        st.success("✅ تم التحقق من الموقع بنجاح.")

    elif st.session_state.emp_verified and st.session_state.emp_data:
        with st.container(border=True):
            st.markdown('<div class="card-title">📍 يرجى التحقق من الموقع</div>', unsafe_allow_html=True)
            st.info("اضغطي أيقونة الموقع التي تظهر بالأسفل واختاري سماح / Allow")

            try:
                location = streamlit_geolocation()
            except Exception:
                location = None
                st.session_state.no_gps_option_available = True

            if location:
                lat = location.get("latitude")
                lon = location.get("longitude")
                error = location.get("error", "")

                if error:
                    st.session_state.no_gps_option_available = True
                    st.warning("⚠️ الموقع غير مفعّل أو تم رفض السماح.")

                elif lat is not None and lon is not None:
                    try:
                        dist_val = distance_m(float(lat), float(lon), SCHOOL_LAT, SCHOOL_LON)
                        if dist_val <= ALLOWED_RADIUS:
                            st.session_state.location_allowed = True
                            st.session_state.no_gps_option_available = False
                            st.success(f"✅ داخل نطاق المدرسة — {int(dist_val)} م")
                            st.rerun()
                        else:
                            st.session_state.no_gps_option_available = True
                            st.error(f"❌ خارج النطاق — {int(dist_val)} م")
                    except Exception:
                        st.session_state.no_gps_option_available = True
                        st.error("❌ خطأ في قراءة الموقع.")

                else:
                    st.session_state.no_gps_option_available = True
                    st.warning("⚠️ لم يتم استلام إحداثيات.")

    # ══════════════════════════════════
    # كرت 3: تصريح الوقت اليدوي
    # ══════════════════════════════════
    if st.session_state.emp_verified and st.session_state.emp_data:
        _emp_permit = st.session_state.emp_data
        _permit_id  = str(_emp_permit.get("الرقم الشخصي","")).strip()
        _active_p   = get_active_permit(_permit_id) if _permit_id else None

        if _active_p:
            p_type  = str(_active_p.get("نوع التصريح","")).strip()
            d_from  = str(_active_p.get("تاريخ البداية","")).strip()
            d_to    = str(_active_p.get("تاريخ النهاية","")).strip()
            t_open  = str(_active_p.get("وقت الفتح","")).strip()
            t_close = str(_active_p.get("وقت الإغلاق","")).strip()
            time_w  = f"من {t_open} إلى {t_close}" if t_open else "يوم كامل"

            with st.container(border=True):
                st.markdown('<div class="card-title">⏰ تعديل وقت يدوي — مصرّح من الأدمن</div>', unsafe_allow_html=True)
                st.markdown(f"""
                <div style="font-size:12px;color:#3B6D11;margin-bottom:10px;">
                صالح من {d_from} إلى {d_to} — {time_w}
                </div>
                """, unsafe_allow_html=True)

                # اختيار التاريخ ضمن نطاق التصريح
                try:
                    min_d = datetime.strptime(d_from, "%Y-%m-%d").date() if d_from else now_bh().date()
                    max_d = datetime.strptime(d_to,   "%Y-%m-%d").date() if d_to   else now_bh().date()
                except:
                    min_d = max_d = now_bh().date()

                sel_date = st.date_input("اختاري التاريخ", value=max_d,
                                          min_value=min_d, max_value=max_d, key="permit_sel_date")
                sel_date_str = sel_date.strftime("%Y-%m-%d")

                # جلب السجل الموجود لهذا التاريخ — قراءة مباشرة بدون كاش
                _idx_p, _row_p = find_today_row_fresh(sel_date_str, _permit_id)

                if _row_p:
                    st.markdown(f"""
                    <div style="background:#f0f4f8;border-radius:10px;padding:8px 12px;font-size:13px;margin-bottom:10px;">
                    السجل الحالي — حضور: <b>{_row_p.get("وقت الحضور","—")}</b> |
                    انصراف: <b>{_row_p.get("وقت الانصراف","—")}</b>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.info("لا يوجد سجل لهذا التاريخ — سيتم إنشاء سجل جديد.")

                col_p1, col_p2 = st.columns(2)

                with col_p1:
                    if p_type in ["حضور","كليهما"]:
                        st.markdown("**وقت الحضور الجديد**")
                        try:
                            curr_att = datetime.strptime(_row_p.get("وقت الحضور","07:00:00"),"%H:%M:%S").time() if _row_p and _row_p.get("وقت الحضور") else time(7,0)
                        except: curr_att = time(7,0)
                        p_att = st.time_input("الوقت", value=curr_att, key="permit_att_time")
                        p_att_str = p_att.strftime("%H:%M:%S")
                        if st.button("✅ حفظ الحضور", use_container_width=True, type="primary", key="btn_permit_att"):
                            try:
                                emp_d = validate_employee(_permit_id) or _emp_permit
                                if _idx_p:
                                    safe_update(sheet, _idx_p, COL_ATTEND, p_att_str)
                                    safe_update(sheet, _idx_p, COL_LATE_REASON, "تعديل يدوي بتصريح الأدمن")
                                    update_work_calculation(_idx_p, {**_row_p, "وقت الحضور": p_att_str})
                                else:
                                    day_n = datetime.strptime(sel_date_str,"%Y-%m-%d").strftime("%A")
                                    safe_append(sheet,[sel_date_str,day_n,
                                        emp_d.get("المدرسة",""),emp_d.get("المهمة",""),
                                        "نعم" if is_yes(emp_d.get("دعم","")) else "لا",
                                        emp_d.get("الاسم",""),_permit_id,
                                        p_att_str,"تعديل يدوي بتصريح الأدمن",
                                        "","","","","","","","","","","","","",""])
                                    _idx2,_row2 = find_today_row_fresh(sel_date_str,_permit_id)
                                    if _idx2: update_work_calculation(_idx2,_row2)
                                log_audit(_permit_id,_emp_permit.get("الاسم",""),"تعديل وقت يدوي بتصريح",f"تاريخ:{sel_date_str}|حضور:{p_att_str}")
                                clear_caches()
                                st.success(f"✅ تم حفظ الحضور: {p_att_str}")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ خطأ: {e}")

                with col_p2:
                    if p_type in ["انصراف","كليهما"]:
                        st.markdown("**وقت الانصراف الجديد**")
                        try:
                            curr_dep = datetime.strptime(_row_p.get("وقت الانصراف","14:00:00"),"%H:%M:%S").time() if _row_p and _row_p.get("وقت الانصراف") else time(14,0)
                        except: curr_dep = time(14,0)
                        p_dep = st.time_input("الوقت", value=curr_dep, key="permit_dep_time")
                        p_dep_str = p_dep.strftime("%H:%M:%S")
                        if st.button("🔵 حفظ الانصراف", use_container_width=True, key="btn_permit_dep"):
                            try:
                                if not _idx_p:
                                    st.error("❌ يجب تسجيل الحضور أولاً.")
                                else:
                                    safe_update(sheet, _idx_p, COL_DEPART, p_dep_str)
                                    safe_update(sheet, _idx_p, COL_DEPART_REASON, "تعديل يدوي بتصريح الأدمن")
                                    update_work_calculation(_idx_p,{**_row_p,"وقت الانصراف":p_dep_str})
                                    log_audit(_permit_id,_emp_permit.get("الاسم",""),"تعديل وقت يدوي بتصريح",f"تاريخ:{sel_date_str}|انصراف:{p_dep_str}")
                                    clear_caches()
                                    st.success(f"✅ تم حفظ الانصراف: {p_dep_str}")
                                    st.rerun()
                            except Exception as e:
                                st.error(f"❌ خطأ: {e}")

    # ══════════════════════════════════
    # كرت 4: العمليات (حضور/انصراف)
    # ══════════════════════════════════
    _permit_active = bool(get_active_permit(str((st.session_state.emp_data or {}).get("الرقم الشخصي","")).strip())) if st.session_state.emp_data else False
    if st.session_state.emp_verified and st.session_state.emp_data and (st.session_state.get("location_allowed") or trusted or _permit_active):
        emp    = st.session_state.emp_data
        emp_id = str(emp.get("الرقم الشخصي","")).strip()

        # إصلاح التسجيل: لا نؤجل العملية بعد الضغط حتى لا تضيع بسبب rerun.
        # أي زر عملية ينفّذ register_operation مباشرة في نفس الضغط.
        st.session_state.operation_saving = False
        st.session_state._queued_op   = ""
        st.session_state._queued_note = ""

        # قراءة مباشرة بدون كاش حتى يظهر تسجيل الحضور/الانصراف فورًا ولا تضغط الموظفة أكثر من مرة.
        data = get_sheet_data_fresh()
        _, today_row = find_today_row(data, today_str, emp_id)

        # إذا تأخر ظهور السجل من Google Sheet، نعرض وقت الحضور فوراً ونمنع الضغط المتكرر
        if st.session_state.get("attendance_done_now") and not today_row:
            today_row = {
                "وقت الحضور": st.session_state.attendance_done_now,
                "وقت الانصراف": "",
                "خروج استئذان": "",
                "عودة": "",
                "سبب التأخير": "",
                "سبب الانصراف": "",
            }

        att_time    = today_row.get("وقت الحضور","—")   if today_row else "—"
        dep_time    = today_row.get("وقت الانصراف","—") if today_row else "—"
        exit_time   = today_row.get("خروج استئذان","—") if today_row else "—"
        return_time = today_row.get("عودة","—")          if today_row else "—"
        has_att     = bool(today_row and today_row.get("وقت الحضور","").strip())
        has_dep     = bool(today_row and today_row.get("وقت الانصراف","").strip())
        has_exit    = bool(today_row and today_row.get("خروج استئذان","").strip())
        has_return  = bool(today_row and today_row.get("عودة","").strip())

        if has_exit and not has_return and not has_dep:   status,sc = "استئذان مفتوح","#BA7517"
        elif has_exit and has_dep and "استئذان انصراف" in str(today_row.get("سبب الانصراف","")): status,sc = "انصراف باستئذان","#185FA5"
        elif has_exit and has_return and has_dep:         status,sc = "منصرف بعد الاستئذان ✓","#185FA5"
        elif has_dep:                                     status,sc = "منصرف ✓","#185FA5"
        elif has_att:                                     status,sc = "حاضر ✓","#3B6D11"
        else:                                             status,sc = "لم يُسجَّل","#A32D2D"

        if today_row and has_att and not has_dep and now_bh().time() >= time(13,30):
            st.warning("⚠️ لم يتم تسجيل الانصراف. يرجى تسجيله قبل المغادرة.")

        with st.container(border=True):
            if has_exit:
                st.markdown(f"""<div class="today-strip">
                    <div class="stat-cell"><span class="stat-val">{att_time}</span><span class="stat-lbl">حضور</span></div>
                    <div style="width:1px;background:#d3d1c7;margin:4px 0;"></div>
                    <div class="stat-cell"><span class="stat-val">{exit_time}</span><span class="stat-lbl">استئذان</span></div>
                    <div style="width:1px;background:#d3d1c7;margin:4px 0;"></div>
                    <div class="stat-cell"><span class="stat-val">{return_time if has_return else '—'}</span><span class="stat-lbl">عودة</span></div>
                    <div style="width:1px;background:#d3d1c7;margin:4px 0;"></div>
                    <div class="stat-cell"><span class="stat-val">{dep_time if has_dep else '—'}</span><span class="stat-lbl">انصراف</span></div>
                    <div style="width:1px;background:#d3d1c7;margin:4px 0;"></div>
                    <div class="stat-cell"><span class="stat-val" style="color:{sc};">{status}</span><span class="stat-lbl">الحالة</span></div>
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown(f"""<div class="today-strip">
                    <div class="stat-cell"><span class="stat-val">{att_time}</span><span class="stat-lbl">حضور</span></div>
                    <div style="width:1px;background:#d3d1c7;margin:4px 0;"></div>
                    <div class="stat-cell"><span class="stat-val">{dep_time}</span><span class="stat-lbl">انصراف</span></div>
                    <div style="width:1px;background:#d3d1c7;margin:4px 0;"></div>
                    <div class="stat-cell"><span class="stat-val" style="color:{sc};">{status}</span><span class="stat-lbl">الحالة</span></div>
                </div>""", unsafe_allow_html=True)

            if today_row and has_dep and "رعاية" not in str(today_row.get("سبب الانصراف","")) and "رعاية" not in str(today_row.get("سبب التأخير","")):
                with st.expander("هل لديكِ رعاية لهذا اليوم؟", expanded=False):
                    st.info("اضغطي نعم ليُحتسب دوامك 5 ساعات رعاية.")
                    if st.button("نعم، لدي رعاية", use_container_width=True, key="btn_care"):
                        if mark_care_for_today(emp_id): st.rerun()

            col1,col2 = st.columns(2)
            with col1:
                if st.button("✅ تسجيل حضور", use_container_width=True, disabled=has_att, key="btn_att"):
                    if now_bh().time() > time(7,30):
                        st.session_state.pending_operation = "تسجيل حضور"
                        st.rerun()
                    else:
                        register_operation("تسجيل حضور", emp_id, "")
                        st.rerun()
            with col2:
                if st.button("🔵 تسجيل انصراف", use_container_width=True, disabled=has_dep, key="btn_dep"):
                    if now_bh().time() < time(14,0):
                        st.session_state.pending_operation = "تسجيل انصراف"
                        st.rerun()
                    else:
                        register_operation("تسجيل انصراف", emp_id, "")
                        st.rerun()
            col3,col4 = st.columns(2)
            with col3:
                if st.button("📤 خروج استئذان", use_container_width=True, key="btn_exit"):
                    st.session_state.pending_operation = "خروج استئذان"
                    st.rerun()
            with col4:
                if st.button("🔁 عودة من استئذان", use_container_width=True, key="btn_return"):
                    register_operation("عودة من استئذان", emp_id, "")
                    st.rerun()

            if st.session_state.get("pending_operation") == "تسجيل حضور":
                with st.container(border=True):
                    st.markdown('<div class="card-title">سبب التأخير — اختياري</div>', unsafe_allow_html=True)
                    late_reason = st.selectbox("السبب",["اختاري السبب (اختياري)"]+reasons,key="late_reason")
                    late_other  = st.text_input("اكتبي السبب",key="late_other") if late_reason=="أخرى" else ""
                    final = "" if late_reason=="اختاري السبب (اختياري)" else (late_other.strip() if late_reason=="أخرى" else late_reason)
                    if st.button("تأكيد تسجيل الحضور", use_container_width=True, type="primary", key="btn_confirm_att"):
                        st.session_state.pending_operation = None
                        register_operation("تسجيل حضور", emp_id, final)
                        st.rerun()

            if st.session_state.get("pending_operation") == "تسجيل انصراف":
                with st.container(border=True):
                    st.markdown('<div class="card-title">سبب الانصراف قبل 2:00</div>', unsafe_allow_html=True)
                    reason = st.selectbox("السبب",reasons,key="early_reason")
                    other  = st.text_input("اكتبي السبب",key="early_other") if reason=="أخرى" else ""
                    final  = other.strip() if reason=="أخرى" else reason
                    if st.button("تأكيد تسجيل الانصراف", use_container_width=True, type="primary", key="btn_confirm_dep"):
                        if not final: st.error("السبب مطلوب")
                        else:
                            st.session_state.pending_operation = None
                            register_operation("تسجيل انصراف", emp_id, final)
                            st.rerun()

            if st.session_state.get("pending_operation") == "خروج استئذان":
                with st.container(border=True):
                    st.markdown('<div class="card-title">نوع وسبب الاستئذان</div>', unsafe_allow_html=True)
                    leave_kind   = st.radio("نوع الاستئذان",["استئذان مع عودة","استئذان انصراف"],horizontal=True,key="leave_kind")
                    reason       = st.selectbox("السبب",reasons,key="exit_reason")
                    other        = st.text_input("اكتبي السبب",key="exit_other") if reason=="أخرى" else ""
                    reason_final = other.strip() if reason=="أخرى" else reason
                    final        = f"{leave_kind} — {reason_final}" if reason_final else leave_kind
                    if st.button("تأكيد خروج الاستئذان", use_container_width=True, type="primary", key="btn_confirm_exit"):
                        if not reason_final: st.error("السبب مطلوب")
                        else:
                            st.session_state.pending_operation = None
                            register_operation("خروج استئذان", emp_id, final)
                            st.rerun()

    # ── زر "موظفة أخرى" للأجهزة الموثوقة ──
    if trusted and st.session_state.emp_verified and st.session_state.emp_data:
        st.markdown("---")
        if st.button("🚪 تسجيل موظفة أخرى", use_container_width=True, key="btn_next_emp"):
            _loc = st.session_state.get("location_allowed", True)
            _fp  = get_device_fingerprint()
            ls_clear_emp_data()
            st.session_state.clear()
            st.session_state.location_allowed = _loc
            st.session_state.device_fp        = _fp
            st.session_state._trusted_cleared = True
            import streamlit.components.v1 as components
            components.html("<script>window.parent.location.reload();</script>", height=0)
    # ══════════════════════════════════
    _show_support = manual_requests_enabled()
    with st.expander("💬 الدعم الفني" if _show_support else "💬 الدعم الفني", expanded=False) if _show_support else st.empty():

        # ── بيانات الموظفة ──
        _sup_emp  = st.session_state.get("emp_data") or {}
        _sup_id   = str(_sup_emp.get("الرقم الشخصي","")).strip()
        _sup_name = str(_sup_emp.get("الاسم","")).strip()
        _sup_sch  = str(_sup_emp.get("المدرسة","")).strip()
        _sup_task = str(_sup_emp.get("المهمة","")).strip()

        if not _sup_id:
            _sup_raw = st.text_input("الرقم الشخصي", placeholder="أدخلي رقمك الشخصي", key="sup_emp_id")
            _sup_id  = ar_to_en_digits(_sup_raw).strip()
            if _sup_id:
                _sup_found = validate_employee(_sup_id)
                if _sup_found:
                    _sup_name = str(_sup_found.get("الاسم","")).strip()
                    _sup_sch  = str(_sup_found.get("المدرسة","")).strip()
                    _sup_task = str(_sup_found.get("المهمة","")).strip()

        # ── القسم 1: تسجيل بدون موقع ──
        st.markdown("#### 📋 تسجيل بدون موقع")
        st.caption("سيُعتمد وقت الإرسال حضوراً أو انصرافاً بعد موافقة الأدمن.")

        ex_att,_ = manual_request_exists_today(_sup_id,"حضور")  if _sup_id else (False,None)
        ex_dep,_ = manual_request_exists_today(_sup_id,"انصراف") if _sup_id else (False,None)

        col_a, col_b = st.columns(2)
        with col_a:
            if st.button(
                "✅ تم إرسال طلب الحضور" if ex_att else "📋 تسجيل حضور بدون موقع",
                use_container_width=True, key="btn_nogps_att",
                disabled=ex_att or st.session_state.get("nogps_saving"),
            ):
                if not _sup_id:
                    st.error("❌ أدخلي رقمك الشخصي أولاً.")
                else:
                    st.session_state.nogps_saving = True
                    ok = submit_manual_request(_sup_id,_sup_name,_sup_sch,_sup_task,"حضور","","","تعذر تحديد الموقع — GPS",f"تسجيل بدون GPS | {today_str}")
                    st.session_state.nogps_saving = False
                    if ok and ok != "duplicate":
                        log_audit(_sup_id,_sup_name,"تسجيل بدون GPS",f"حضور | {today_str}")
                        st.success("✅ تم إرسال طلب الحضور — بانتظار اعتماد الأدمن.")
                        st.rerun()
                    elif ok == "duplicate":
                        st.warning("⚠️ تم إرسال طلب سابق لهذا اليوم.")
                    else:
                        st.error("❌ تعذّر الإرسال.")
        with col_b:
            if st.button(
                "✅ تم إرسال طلب الانصراف" if ex_dep else "📋 تسجيل انصراف بدون موقع",
                use_container_width=True, key="btn_nogps_dep",
                disabled=ex_dep or st.session_state.get("nogps_saving"),
            ):
                if not _sup_id:
                    st.error("❌ أدخلي رقمك الشخصي أولاً.")
                else:
                    st.session_state.nogps_saving = True
                    ok = submit_manual_request(_sup_id,_sup_name,_sup_sch,_sup_task,"انصراف","","","تعذر تحديد الموقع — GPS",f"تسجيل بدون GPS | {today_str}")
                    st.session_state.nogps_saving = False
                    if ok and ok != "duplicate":
                        log_audit(_sup_id,_sup_name,"تسجيل بدون GPS",f"انصراف | {today_str}")
                        st.success("✅ تم إرسال طلب الانصراف — بانتظار اعتماد الأدمن.")
                        st.rerun()
                    elif ok == "duplicate":
                        st.warning("⚠️ تم إرسال طلب سابق لهذا اليوم.")
                    else:
                        st.error("❌ تعذّر الإرسال.")

        # ── القسم 2: تواصل مع الأدمن للتعديل ──
        st.markdown("---")
        st.markdown("#### 💬 تواصل مع الأدمن للتعديل")
        st.caption("لتصحيح وقت الحضور أو الانصراف أو أي تعديل.")

        data_wa  = get_sheet_data()
        _,row_wa = find_today_row(data_wa,today_str,_sup_id) if _sup_id else (None,None)
        att_wa   = row_wa.get("وقت الحضور","—")   if row_wa else "—"
        dep_wa   = row_wa.get("وقت الانصراف","—") if row_wa else "—"

        issue_choice = st.selectbox("نوع المشكلة",[
            "تصحيح وقت الحضور","تصحيح وقت الانصراف",
            "تعديل سبب التأخير أو الانصراف","مشكلة تقنية",
            "تعديل بيانات شخصية","أخرى"],key="wa_issue")
        issue_notes = st.text_input("تفاصيل إضافية (اختياري)", key="wa_notes")

        wa_msg = f"""مرحباً 👋
لدي طلب في نظام الحضور:
الاسم: {_sup_name}
الرقم الشخصي: {_sup_id}
المدرسة: {_sup_sch}
المهمة: {_sup_task}
التاريخ: {today_str} | الوقت: {now_bh().strftime('%H:%M:%S')}
وقت الحضور: {att_wa} | وقت الانصراف: {dep_wa}
نوع الطلب: {issue_choice}
{('التفاصيل: '+issue_notes) if issue_notes.strip() else ''}"""
        wa_link = "https://wa.me/97333738668?text=" + urllib.parse.quote(wa_msg)
        st.link_button("📞 فتح واتساب برسالة جاهزة", wa_link, use_container_width=True)

# ══════════════════════════════════════════════════════════════════
# ══ واجهة الأدمن ══
# ══════════════════════════════════════════════════════════════════
else:
    if not st.session_state.admin_logged_in:
        with st.container(border=True):
            st.markdown('<div class="card-title">🛡️ دخول الأدمن</div>',unsafe_allow_html=True)
            pw=st.text_input("كلمة المرور",type="password")
            if st.button("دخول",use_container_width=True):
                if pw.strip()==ADMIN_PASSWORD:
                    st.session_state.admin_logged_in=True; st.session_state.admin_last_active=now_bh(); st.rerun()
                else: st.error("❌ كلمة المرور غير صحيحة.")
    else:
        st.session_state.admin_last_active=now_bh()
        st.markdown("## 🛡️ لوحة الأدمن")

        admin_tab = st.selectbox("القسم",[
            "📊 إحصائيات اليوم",
            "✏️ حضور اليوم والتعديل المباشر",
            "📑 التقارير",
            "🛠️ إصلاح شامل",
            "🆘 طلبات التسجيل اليدوي",
            "📞 التواصل والمتابعة",
            "⚙️ إعدادات التسجيل اليدوي",
            "🔴 تسجيل الغياب",
            "📋 مطلوبات اليوم",
            "📅 دوام الأقسام",
            "🧹 تنظيف التكرارات",
            "✏️ تعديل سجل",
            "➕ تسجيل يدوي",
            "📋 القائمة البيضاء",
            "🚫 محاولات التسجيل",
            "📡 تجاوز الموقع",
            "⚙️ قفل الجهاز",
            "📱 الأجهزة الموثوقة",
            "🔓 فتح قفل جهاز",
            "🔍 سجل التدقيق",
            "⚠️ تقرير الأجهزة",
            "⏰ تصاريح الوقت اليدوي",
            "📅 إعدادات الدوام المخصص",
        ], key="admin_tab")


        # ── حضور اليوم والتعديل المباشر ─────────────────────────────
        if admin_tab=="✏️ حضور اليوم والتعديل المباشر":
            st.markdown("#### ✏️ حضور اليوم والتعديل المباشر")
            st.caption("تشوفين الحاضرات وتعدلين مباشرة، وتضيفين حضور للي ما سجلوا مع فلتر بالاسم أو الرقم الشخصي.")

            # رسالة نجاح ثابتة بدون إخراجك من نفس الصفحة أو نفس الفلتر
            if st.session_state.get("today_direct_msg"):
                st.success(st.session_state.get("today_direct_msg"))

            # اختيار تاريخ العرض؛ الافتراضي اليوم
            direct_view_date = st.date_input("تاريخ العرض", value=now_bh().date(), key="direct_view_date")
            direct_view_date_str = direct_view_date.strftime("%Y-%m-%d")

            # ── تحديد دوام معلمات اليوم من الأقسام مباشرة من نفس الصفحة ──
            with st.container(border=True):
                st.markdown("##### 📌 تحديد دوام المعلمات من الأقسام")
                st.caption("اختاري الأقسام/المهام المطلوبة لهذا التاريخ. بعدها قائمة غير المسجلات تتحدث بناءً على الاختيار.")

                direct_schedule_date_str = direct_view_date_str
                current_direct_daily = [
                    str(r.get("المهمة", "")).strip()
                    for r in get_daily_schedule_records()
                    if str(r.get("التاريخ", "")).strip().replace("/", "-") == direct_schedule_date_str
                    and (str(r.get("نشط", "")).strip() == "" or is_yes(r.get("نشط", "")))
                    and str(r.get("المهمة", "")).strip()
                ]
                current_direct_daily = list(dict.fromkeys(current_direct_daily))

                selected_direct_tasks = st.multiselect(
                    "الأقسام/المهام التي دوامها في هذا التاريخ",
                    TASKS_ALL,
                    default=[t for t in current_direct_daily if t in TASKS_ALL],
                    key="direct_schedule_tasks"
                )
                direct_schedule_note = st.text_input(
                    "ملاحظة اختيارية",
                    value="تحديد دوام من صفحة حضور اليوم",
                    key="direct_schedule_note"
                )

                c_ds1, c_ds2 = st.columns(2)
                with c_ds1:
                    if st.button("💾 اعتماد دوام الأقسام لهذا التاريخ", use_container_width=True, type="primary", key="direct_save_schedule_tasks"):
                        if not selected_direct_tasks:
                            st.error("❌ اختاري قسم/مهمة واحدة على الأقل.")
                        else:
                            try:
                                records = daily_schedule_sheet.get_all_records()
                                batch_updates = []
                                for i, r in enumerate(records):
                                    if str(r.get("التاريخ", "")).strip().replace("/", "-") == direct_schedule_date_str:
                                        batch_updates.append({"range": f"C{i+2}", "values": [["لا"]]})
                                if batch_updates:
                                    daily_schedule_sheet.batch_update(batch_updates)
                                for task_name in selected_direct_tasks:
                                    daily_schedule_sheet.append_row(
                                        [direct_schedule_date_str, task_name, "نعم", direct_schedule_note],
                                        value_input_option="USER_ENTERED"
                                    )
                                get_daily_schedule_records.clear()
                                clear_caches()
                                st.session_state.today_direct_msg = "✅ تم اعتماد دوام الأقسام لهذا التاريخ."
                                st.success(st.session_state.today_direct_msg)
                            except Exception as e:
                                st.error(f"❌ تعذر حفظ دوام الأقسام: {e}")

                with c_ds2:
                    if st.button("🗑️ إلغاء دوام الأقسام الخاص بهذا التاريخ", use_container_width=True, key="direct_disable_schedule_tasks"):
                        try:
                            records = daily_schedule_sheet.get_all_records()
                            batch_updates = []
                            for i, r in enumerate(records):
                                if str(r.get("التاريخ", "")).strip().replace("/", "-") == direct_schedule_date_str:
                                    batch_updates.append({"range": f"C{i+2}", "values": [["لا"]]})
                            if batch_updates:
                                daily_schedule_sheet.batch_update(batch_updates)
                            get_daily_schedule_records.clear()
                            clear_caches()
                            st.session_state.today_direct_msg = "✅ تم إلغاء دوام الأقسام الخاص بهذا التاريخ."
                            st.success(st.session_state.today_direct_msg)
                        except Exception as e:
                            st.error(f"❌ تعذر الإلغاء: {e}")

                if current_direct_daily:
                    st.success("دوام هذا التاريخ محدد حالياً: " + "، ".join(current_direct_daily))
                else:
                    st.info("لا يوجد دوام خاص محفوظ لهذا التاريخ. النظام سيعتمد على مطلوبات اليوم أو الجدول الأسبوعي أو القائمة البيضاء.")

            col_ref1, col_ref2 = st.columns([3,1])
            with col_ref2:
                if st.button("🔄 تحديث", key="btn_refresh_today_direct", use_container_width=True):
                    clear_caches()
                    get_whitelist.clear()
                    st.session_state.today_direct_msg = "✅ تم تحديث البيانات من Google Sheet."
                    st.rerun()
            with col_ref1:
                st.info("هذه الصفحة تقرأ من Google Sheet مباشرة بدون كاش.")

            # قراءة سجلات التاريخ المختار مباشرة من sheet1 مع رقم الصف الحقيقي
            data_direct = get_sheet_data_fresh()

            def _direct_normalize_date(value):
                txt = ar_to_en_digits(str(value or "")).strip()
                if not txt:
                    return ""
                txt = txt.replace("/", "-")
                txt = txt.split(" ")[0].strip()
                # تاريخ بصيغة ISO
                if len(txt) >= 10 and txt[4:5] == "-" and txt[7:8] == "-":
                    return txt[:10]
                # تاريخ كسيري أو يوم/شهر/سنة
                for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y", "%Y-%m-%dT%H:%M:%S"]:
                    try:
                        return datetime.strptime(txt[:10], fmt).strftime("%Y-%m-%d")
                    except Exception:
                        pass
                # لو Google Sheet رجع رقم تسلسلي لتاريخ Excel
                try:
                    if txt.replace(".", "", 1).isdigit():
                        base = datetime(1899, 12, 30)
                        return (base + timedelta(days=float(txt))).strftime("%Y-%m-%d")
                except Exception:
                    pass
                return txt[:10]

            today_rows_direct = []
            seen_dates_direct = {}
            for row_num_direct, rr in enumerate(data_direct, start=2):
                raw_date = rr.get("التاريخ", rr.get("Date", ""))
                row_date = _direct_normalize_date(raw_date)
                if row_date:
                    seen_dates_direct[row_date] = seen_dates_direct.get(row_date, 0) + 1
                if row_date != direct_view_date_str:
                    continue

                r = dict(rr)
                school_name = str(r.get("اسم المدرسة", "") or r.get("المدرسة", "") or r.get("School", "")).strip()
                task_name   = str(r.get("المهمة", "") or r.get("القسم", "") or r.get("Task", "")).strip()
                name_value  = str(r.get("الاسم الثلاثي", "") or r.get("الاسم", "") or r.get("Name", "")).strip()
                eid_value   = normalize_emp_id(r.get("الرقم الشخصي", "") or r.get("ID", ""))

                r["_row_num"] = row_num_direct
                r["التاريخ"] = row_date
                r["اسم المدرسة"] = school_name
                r["المدرسة"] = school_name
                r["المهمة"] = task_name
                r["الاسم الثلاثي"] = name_value
                r["الاسم"] = name_value
                r["الرقم الشخصي"] = eid_value
                today_rows_direct.append(r)

            wl_all_direct = get_whitelist()

            # تحديد المطلوبات حسب التاريخ المختار
            required_people_direct = required_people_for_date(direct_view_date_str)
            if required_people_direct is not None:
                required_wl_direct = required_people_direct
                source_label_direct = "مطلوبات اليوم"
            else:
                scheduled_tasks_direct, schedule_source_direct = scheduled_tasks_for_date(direct_view_date_str)
                if scheduled_tasks_direct is None:
                    required_wl_direct = wl_all_direct
                    source_label_direct = "كل القائمة البيضاء"
                else:
                    required_wl_direct = {eid: emp for eid, emp in wl_all_direct.items() if emp_required_on_day(emp, scheduled_tasks_direct)}
                    source_label_direct = f"جدول الأقسام — {schedule_source_direct}"

            attended_today_direct = [r for r in today_rows_direct if str(r.get("وقت الحضور","")).strip()]
            attended_ids_direct = set(normalize_emp_id(r.get("الرقم الشخصي","")) for r in attended_today_direct)
            missing_today_direct = {normalize_emp_id(eid): emp for eid, emp in required_wl_direct.items() if normalize_emp_id(eid) not in attended_ids_direct}

            m1, m2, m3 = st.columns(3)
            m1.metric("مصدر القائمة", source_label_direct)
            m2.metric("سجلوا حضور", len(attended_today_direct))
            m3.metric("لم يسجلوا من المطلوبات", len(missing_today_direct))

            if not today_rows_direct:
                with st.expander("🔎 تشخيص سريع إذا لم تظهر سجلات الحضور", expanded=False):
                    st.caption("هذه آخر التواريخ التي قرأها النظام من sheet1. إذا تاريخ اليوم غير موجود هنا، فالمشكلة من عمود التاريخ أو من أن البيانات ليست في sheet1.")
                    if seen_dates_direct:
                        for d, cnt in sorted(seen_dates_direct.items(), reverse=True)[:10]:
                            st.write(f"{d} — {cnt} سجل")
                    else:
                        st.write("لم يتمكن النظام من قراءة أي تاريخ من عمود التاريخ.")

            direct_tab = st.radio("اختاري العملية", ["✅ الحاضرات اليوم", "➕ إضافة لمن لم تسجل"], horizontal=True, key="direct_today_tab")

            if direct_tab == "✅ الحاضرات اليوم":
                st.markdown("##### ✅ الحاضرات — تعديل مباشر")

                # فلتر عام بالاسم أو الرقم الشخصي حتى لا تحتاجين تبحثين يدوياً
                search_direct = st.text_input("فلتر بالاسم أو الرقم الشخصي", key="direct_att_search", placeholder="اكتبي جزء من الاسم أو الرقم")
                search_direct_norm = normalize_name(search_direct) if search_direct else ""

                if not attended_today_direct:
                    st.warning("لا توجد سجلات حضور لهذا التاريخ حتى الآن.")
                else:
                    attended_sorted = sorted(attended_today_direct, key=lambda r: (
                        str(r.get("المهمة", "")),
                        str(r.get("اسم المدرسة", r.get("المدرسة", ""))),
                        str(r.get("الاسم الثلاثي", r.get("الاسم", ""))),
                    ))

                    school_filter_direct = st.selectbox(
                        "فلترة بالمدرسة",
                        ["الكل"] + sorted(list(set(str(r.get("اسم المدرسة", r.get("المدرسة", ""))).strip() for r in attended_sorted if str(r.get("اسم المدرسة", r.get("المدرسة", ""))).strip()))),
                        key="direct_school_filter"
                    )
                    task_filter_direct = st.selectbox(
                        "فلترة بالمهمة",
                        ["الكل"] + sorted(list(set(str(r.get("المهمة","")).strip() for r in attended_sorted if str(r.get("المهمة","")).strip()))),
                        key="direct_task_filter"
                    )

                    filtered_attended = attended_sorted
                    if school_filter_direct != "الكل":
                        filtered_attended = [r for r in filtered_attended if str(r.get("اسم المدرسة", r.get("المدرسة", ""))).strip() == school_filter_direct]
                    if task_filter_direct != "الكل":
                        filtered_attended = [r for r in filtered_attended if str(r.get("المهمة","")).strip() == task_filter_direct]
                    if search_direct.strip():
                        filtered_attended = [r for r in filtered_attended if search_direct.strip() in str(r.get("الرقم الشخصي","")).strip() or search_direct_norm in normalize_name(r.get("الاسم الثلاثي", r.get("الاسم", "")))]

                    st.caption(f"المعروض حالياً: {len(filtered_attended)} سجل")

                    for idx_direct, r in enumerate(filtered_attended):
                        eid = str(r.get("الرقم الشخصي","")).strip()
                        date_val = str(r.get("التاريخ","")).strip().replace("/","-")
                        row_num = int(r.get("_row_num", 0) or 0)
                        if not row_num:
                            continue
                        name_val = str(r.get("الاسم الثلاثي", r.get("الاسم", ""))).strip()
                        school_val = str(r.get("اسم المدرسة", r.get("المدرسة", ""))).strip()
                        task_val = str(r.get("المهمة","")).strip()
                        att_val = str(r.get("وقت الحضور","")).strip()
                        dep_val = str(r.get("وقت الانصراف","")).strip()

                        with st.expander(f"{name_val} — #{eid} — حضور: {att_val or '—'} — انصراف: {dep_val or '—'}", expanded=False):
                            st.markdown(f"""
                            <div class="audit-row">
                                <b>{name_val}</b><br>
                                الرقم الشخصي: {eid}<br>
                                المدرسة: {school_val}<br>
                                المهمة: {task_val}<br>
                                رقم الصف في Sheet1: {row_num}
                            </div>
                            """, unsafe_allow_html=True)

                            c1, c2 = st.columns(2)
                            with c1:
                                new_att = st.text_input("وقت الحضور", value=att_val, key=f"direct_att_{idx_direct}_{row_num}", placeholder="07:00:00")
                                new_late_reason = st.text_input("سبب التأخير", value=str(r.get("سبب التأخير","")).strip(), key=f"direct_late_{idx_direct}_{row_num}")
                            with c2:
                                new_dep = st.text_input("وقت الانصراف", value=dep_val, key=f"direct_dep_{idx_direct}_{row_num}", placeholder="14:00:00")
                                new_dep_reason = st.text_input("سبب الانصراف", value=str(r.get("سبب الانصراف","")).strip(), key=f"direct_depr_{idx_direct}_{row_num}")

                            c3, c4 = st.columns(2)
                            with c3:
                                care_choice = st.selectbox("رعاية؟", ["لا", "نعم"], index=1 if is_care_day(r) else 0, key=f"direct_care_{idx_direct}_{row_num}")
                            with c4:
                                daily_type_edit = st.text_input("نوع الدوام اليومي", value=str(r.get("نوع الدوام اليومي","")).strip(), key=f"direct_dtype_{idx_direct}_{row_num}")

                            col_save, col_delete = st.columns(2)
                            with col_save:
                                if st.button("💾 حفظ التعديل وإعادة الحساب", use_container_width=True, type="primary", key=f"direct_save_{idx_direct}_{row_num}"):
                                    try:
                                        safe_update(sheet, row_num, COL_ATTEND, new_att.strip())
                                        safe_update(sheet, row_num, COL_LATE_REASON, new_late_reason.strip())
                                        safe_update(sheet, row_num, COL_DEPART, new_dep.strip())
                                        safe_update(sheet, row_num, COL_DEPART_REASON, new_dep_reason.strip())
                                        safe_update(sheet, row_num, COL_CARE_CONF, "نعم" if care_choice == "نعم" else "")
                                        safe_update(sheet, row_num, COL_DAILY_TYPE, daily_type_edit.strip())
                                        updated_row = dict(r)
                                        updated_row.update({
                                            "وقت الحضور": new_att.strip(),
                                            "سبب التأخير": new_late_reason.strip(),
                                            "وقت الانصراف": new_dep.strip(),
                                            "سبب الانصراف": new_dep_reason.strip(),
                                            "تأكيد الرعاية": "نعم" if care_choice == "نعم" else "",
                                            "نوع الدوام اليومي": daily_type_edit.strip(),
                                        })
                                        update_work_calculation(row_num, updated_row)
                                        log_audit(eid, name_val, "تعديل مباشر لحضور اليوم", f"تاريخ:{date_val}|صف:{row_num}")
                                        clear_caches()
                                        st.session_state.today_direct_msg = f"✅ تم تعديل {name_val} وحفظ التغييرات بنجاح."
                                        st.success(st.session_state.today_direct_msg)
                                    except Exception as e:
                                        st.error(f"❌ خطأ أثناء الحفظ: {e}")
                            with col_delete:
                                if st.button("🗑️ حذف السجل", use_container_width=True, key=f"direct_delete_{idx_direct}_{row_num}"):
                                    st.session_state[f"direct_confirm_delete_{row_num}"] = True
                                if st.session_state.get(f"direct_confirm_delete_{row_num}"):
                                    st.warning("تأكيد حذف السجل؟")
                                    if st.button("نعم، احذفيه", use_container_width=True, key=f"direct_yes_delete_{idx_direct}_{row_num}"):
                                        try:
                                            sheet.delete_rows(row_num)
                                            log_audit(eid, name_val, "حذف سجل حضور اليوم", f"تاريخ:{date_val}|صف:{row_num}")
                                            clear_caches()
                                            st.session_state.today_direct_msg = f"✅ تم حذف سجل {name_val} بنجاح."
                                            st.success(st.session_state.today_direct_msg)
                                        except Exception as e:
                                            st.error(f"❌ خطأ أثناء الحذف: {e}")

            else:
                st.markdown("##### ➕ إضافة حضور لمن لم تسجل")
                st.caption("استخدمي الفلتر بالاسم أو الرقم الشخصي، ثم اختاري الموظفة وأضيفي الحضور.")

                missing_search = st.text_input("فلتر بالاسم أو الرقم الشخصي", key="direct_missing_search", placeholder="اكتبي الاسم أو الرقم")

                # مصدر الإضافة هنا هو القائمة البيضاء كاملة، وليس مطلوبات اليوم.
                # نُخفي فقط من لديها حضور فعلي في التاريخ المختار حتى لا ننشئ تكرار.
                whitelist_for_add = {}
                for eid, emp in wl_all_direct.items():
                    eid_clean = normalize_emp_id(eid)
                    if not eid_clean:
                        continue
                    if eid_clean in attended_ids_direct:
                        continue
                    emp_copy = dict(emp)
                    emp_copy["الاسم"] = get_emp_name(emp_copy)
                    emp_copy["المدرسة"] = get_emp_school(emp_copy)
                    emp_copy["المهمة"] = str(emp_copy.get("المهمة", "")).strip()
                    whitelist_for_add[eid_clean] = emp_copy

                missing_items_all = sorted(whitelist_for_add.items(), key=lambda item: (
                    str(item[1].get("المهمة", "")),
                    str(item[1].get("المدرسة", "")),
                    str(item[1].get("الاسم", "")),
                ))

                search_value = normalize_emp_id(missing_search)
                search_name = normalize_name(str(missing_search or "")) if str(missing_search or "").strip() else ""

                if search_value or search_name:
                    missing_items = []
                    for eid, emp in missing_items_all:
                        emp_id_clean = normalize_emp_id(eid)
                        emp_name_norm = normalize_name(str(emp.get("الاسم", "")))
                        if (search_value and search_value in emp_id_clean) or (search_name and search_name in emp_name_norm):
                            missing_items.append((eid, emp))
                else:
                    missing_items = missing_items_all[:80]

                st.caption(f"المعروض حالياً: {len(missing_items)} من أصل {len(missing_items_all)} من القائمة البيضاء بدون حضور لهذا التاريخ")

                if not missing_items:
                    st.warning("لا توجد نتيجة مطابقة للفلتر في القائمة البيضاء.")
                else:
                    missing_labels = [f"{emp.get('الاسم','')} — #{eid} — {emp.get('المدرسة','')} — {emp.get('المهمة','')}" for eid, emp in missing_items]
                    selected_missing = st.selectbox("اختاري الموظفة لإضافة حضور", missing_labels, key="direct_missing_select")
                    selected_idx = missing_labels.index(selected_missing)
                    add_eid, add_emp = missing_items[selected_idx]

                    st.markdown(f"""
                    <div class="audit-row">
                        <b>{add_emp.get('الاسم','')}</b><br>
                        الرقم الشخصي: {add_eid}<br>
                        المدرسة: {add_emp.get('المدرسة','')}<br>
                        المهمة: {add_emp.get('المهمة','')}
                    </div>
                    """, unsafe_allow_html=True)

                    cadd1, cadd2 = st.columns(2)
                    with cadd1:
                        add_att_time = st.text_input("وقت الحضور", value="07:00:00", key="direct_add_att")
                        add_late_reason = st.selectbox("سبب التأخير / الملاحظة", [""] + reasons, key="direct_add_late_reason")
                    with cadd2:
                        add_dep_time = st.text_input("وقت الانصراف اختياري", value="", key="direct_add_dep")
                        add_dep_reason = st.selectbox("سبب الانصراف اختياري", [""] + reasons, key="direct_add_dep_reason")

                    add_care_choice = st.selectbox("هل لديها رعاية؟", ["لا", "نعم"], key="direct_add_care")
                    add_note = st.text_input("ملاحظة", value="إضافة حضور يدوي من صفحة حضور اليوم", key="direct_add_note")

                    if st.button("➕ إضافة السجل الآن", use_container_width=True, type="primary", key="direct_add_record_btn"):
                        if not str(add_att_time).strip():
                            st.error("❌ وقت الحضور مطلوب.")
                        else:
                            try:
                                existing_idx, existing_row = find_today_row_fresh(direct_view_date_str, str(add_eid).strip())
                                if existing_row and str(existing_row.get("وقت الحضور","")).strip():
                                    st.warning("⚠️ هذه الموظفة أصبح لديها سجل حضور، لم يتم إنشاء سجل مكرر.")
                                else:
                                    add_name = str(add_emp.get("الاسم","")).strip()
                                    add_school = str(add_emp.get("المدرسة","")).strip()
                                    add_task = str(add_emp.get("المهمة","")).strip()
                                    add_support = "نعم" if is_support_employee_record(add_emp) else "لا"
                                    late_reason_to_save = "رعاية" if add_care_choice == "نعم" else str(add_late_reason or "").strip()
                                    daily_type_to_save = "رعاية" if add_care_choice == "نعم" else ""
                                    care_confirm_to_save = "نعم" if add_care_choice == "نعم" else ""
                                    row_values = [
                                        direct_view_date_str, direct_view_date.strftime("%A"), add_school, add_task, add_support,
                                        add_name, str(add_eid).strip(), str(add_att_time).strip(), late_reason_to_save,
                                        str(add_dep_time).strip(), str(add_dep_reason or "").strip(), "", "", "",
                                        "", "", "", "", "", daily_type_to_save, "", care_confirm_to_save,
                                        "إضافة يدوية من حضور اليوم"
                                    ]
                                    safe_append(sheet, row_values)
                                    idx_new, row_new = find_today_row_fresh(direct_view_date_str, str(add_eid).strip())
                                    if idx_new and row_new:
                                        update_work_calculation(idx_new, row_new)
                                    log_audit(str(add_eid).strip(), add_name, "إضافة حضور من صفحة حضور اليوم", f"تاريخ:{direct_view_date_str}|{add_note}")
                                    clear_caches()
                                    st.session_state.today_direct_msg = f"✅ تم إضافة سجل الحضور لـ {add_name} بنجاح."
                                    st.success(st.session_state.today_direct_msg)
                            except Exception as e:
                                st.error(f"❌ خطأ أثناء الإضافة: {e}")

        # ── إحصائيات اليوم ──────────────────────────────────────
        elif admin_tab=="📊 إحصائيات اليوم":
            # زر تحديث يدوي
            col_ref1, col_ref2 = st.columns([3,1])
            with col_ref2:
                if st.button("🔄 تحديث البيانات", key="btn_refresh_stats", use_container_width=True):
                    clear_caches()
                    get_whitelist.clear()
                    st.rerun()
            with col_ref1:
                st.caption("⏱️ اضغطي تحديث للحصول على أحدث البيانات من الشيت فوراً.")

            # قراءة مباشرة بدون كاش لضمان دقة البيانات
            data = get_sheet_data_fresh()
            today_rows=[r for r in data if str(r.get("التاريخ","")).strip().replace("/","-")==today_str]

            # ربط إحصائيات اليوم بقائمة مطلوبات اليوم أولاً، ثم جدول دوام الأقسام إذا لا توجد قائمة
            today_day_ar = day_arabic
            wl_all = get_whitelist()
            required_from_people = required_people_for_date(today_str)

            if required_from_people is not None:
                required_wl = required_from_people
                st.success("✅ الإحصائيات تعتمد على ورقة مطلوبات_اليوم لهذا التاريخ.")
                with st.expander(f"📋 مطلوبات اليوم مرتبة حسب المهمة ← المدرسة ← المعلمات — {today_str}", expanded=False):
                    last_task = None
                    last_school = None
                    for eid, emp in required_wl.items():
                        task = str(emp.get("المهمة", "")).strip() or "غير محدد"
                        school = str(emp.get("المدرسة", "")).strip() or "غير محدد"
                        name = str(emp.get("الاسم", "")).strip()
                        if task != last_task:
                            st.markdown(f"### 📌 {task}")
                            last_task = task
                            last_school = None
                        if school != last_school:
                            st.markdown(f"**🏫 {school}**")
                            last_school = school
                        st.markdown(f"- {name} — #{eid}")
            else:
                scheduled_tasks, schedule_source = scheduled_tasks_for_date(today_str)
                if scheduled_tasks is None:
                    required_wl = wl_all
                    st.warning("⚠️ لا توجد قائمة مطلوبات اليوم ولم يتم تحديد دوام أقسام لهذا اليوم، لذلك ستعرض الإحصائيات على جميع القائمة البيضاء.")
                else:
                    required_wl = {eid: emp for eid, emp in wl_all.items() if emp_required_on_day(emp, scheduled_tasks)}
                    with st.expander(f"📅 الأقسام المعتمدة في إحصائيات اليوم ({today_day_ar}) — مصدر الجدول: {schedule_source}", expanded=False):
                        for t in scheduled_tasks:
                            st.markdown(f"- {t}")

            required_ids = set(str(eid).strip() for eid in required_wl.keys())

            # سجلات اليوم للأقسام المطلوبة فقط
            today_required_rows = [r for r in today_rows if str(r.get("الرقم الشخصي","")).strip() in required_ids]

            # دعم خارجي: حضر اليوم لكنه غير موجود في القائمة البيضاء
            external_support_rows = []
            for r in today_rows:
                eid = str(r.get("الرقم الشخصي","")).strip()
                support_raw = str(r.get("دعم","")).strip()
                task_txt = str(r.get("المهمة","")).strip()
                is_external_support = eid and eid not in wl_all and (is_yes(support_raw) or "دعم" in task_txt)
                if is_external_support:
                    external_support_rows.append(r)

            try:
                abs_records_all = absence_sheet.get_all_records()
                abs_today=[r for r in abs_records_all if r.get("التاريخ")==today_str and str(r.get("الرقم الشخصي","")).strip() in required_ids]
            except:
                abs_today=[]

            # طلبات الحضور اليدوية المعلقة تعتبر "وصلت للنظام" فلا نرسل لها تذكير عدم حضور بالخطأ
            try:
                manual_reqs_all = manual_requests_sheet.get_all_records()
                pending_manual_att_ids = set(
                    str(r.get("الرقم الشخصي","")).strip()
                    for r in manual_reqs_all
                    if str(r.get("تاريخ الطلب","")).strip() == today_str
                    and str(r.get("نوع الطلب","")).strip() == "حضور"
                    and str(r.get("الحالة","")).strip() in ["", "بانتظار الاعتماد"]
                )
            except:
                pending_manual_att_ids = set()

            attended=[r for r in today_required_rows if r.get("وقت الحضور")]
            late_list=[r for r in today_required_rows if is_late_for_statistics(r)]
            early_dep=[r for r in today_required_rows if r.get("وقت الانصراف","") and r.get("وقت الانصراف","")< "14:00:00"]
            on_leave=[r for r in today_required_rows if r.get("خروج استئذان") and not r.get("عودة") and not r.get("وقت الانصراف")]
            missing_depart=[r for r in today_required_rows if r.get("وقت الحضور") and not r.get("وقت الانصراف")]
            auto_closed=[r for r in today_required_rows if str(r.get("إغلاق تلقائي","")).strip()]
            correction_done_rows=[r for r in today_required_rows if is_correction_done_day(r)]
            no_gps_rows=[r for r in today_rows if "بدون تحقق GPS" in str(r.get("سبب التأخير", "")) or "بدون تحقق GPS" in str(r.get("سبب الانصراف", "")) or "بدون تحقق GPS" in str(r.get("محاولة", ""))]

            c1,c2,c3=st.columns(3)
            c1.metric("المتوقع حضورهم اليوم",len(required_wl))
            c2.metric("حاضرون من الأقسام",len(attended))
            c3.metric("غائبات",len(abs_today))
            c4,c5,c6=st.columns(3)
            c4.metric("متأخرون",len(late_list))
            c5.metric("لم يسجلن انصراف",len(missing_depart))
            c6.metric("استئذان مفتوح",len(on_leave))
            c7,c8,c9=st.columns(3)
            c7.metric("الدعم الخارجي",len(external_support_rows))
            c8.metric("إجمالي حضور اليوم",len([r for r in today_rows if r.get("وقت الحضور")]))
            c9.metric("خارج الأقسام المحددة",max(0, len([r for r in today_rows if r.get("وقت الحضور")])-len(attended)-len(external_support_rows)))
            c10,c11,c12=st.columns(3)
            c10.metric("معفى - انتهاء التصحيح",len(correction_done_rows))
            c11.metric("إغلاق تلقائي",len(auto_closed))
            c12.metric("بدون تحقق GPS",len(no_gps_rows))

            st.info("✅ الإحصائيات الأساسية أعلاه تعتمد على دوام الأقسام المحدد لهذا اليوم، مع فصل الدعم الخارجي غير الموجود في القائمة البيضاء.")

            # تنبيه بعد الساعة 8:00 للموظفات المطلوب دوامهن ولم يسجلن حضور (باستثناء الدعم)
            if now_bh().time() >= time(8, 0):
                attended_ids = set(str(r.get("الرقم الشخصي", "")).strip() for r in today_required_rows if r.get("وقت الحضور"))
                absent_ids   = set(str(r.get("الرقم الشخصي", "")).strip() for r in abs_today)
                # استثناء الدعم من القائمة
                not_checked_in = {
                    eid: emp for eid, emp in required_wl.items()
                    if str(eid).strip() not in attended_ids
                    and str(eid).strip() not in absent_ids
                    and str(eid).strip() not in pending_manual_att_ids
                    and not is_yes(str(emp.get("دعم","")).strip())
                    and "دعم" not in str(emp.get("المهمة","")).strip()
                }

                if not_checked_in:
                    st.markdown(f"#### 🚨 لم يسجلن حضور حتى الآن — {len(not_checked_in)} موظفة")
                    st.caption("باستثناء الدعم. تستثني من حضرن أو تم تسجيل غيابهن.")

                    # ── زر نسخ الكل ──────────────────────────────
                    bulk_msg = f"""السلام عليكم 🌷

لم يُسجَّل حضوركِ في نظام الحضور الإلكتروني لهذا اليوم.
يُرجى التسجيل فوراً — النظام الإلكتروني هو المرجع الرسمي المعتمد.

🔗 {APP_URL}

📍 في حال عدم عمل التطبيق:
- لابتوب في الصالة الرياضية
- لابتوب في مقر الكنترول الخارجي
"""
                    phones_list = []
                    for eid, emp in not_checked_in.items():
                        ph = str(emp.get("رقم التواصل","") or "").strip().replace(" ","")
                        if ph:
                            if not ph.startswith("973"):
                                ph = "973" + ph.lstrip("0")
                            phones_list.append(f"{emp.get('الاسم','')} — {ph}")

                    col_bulk1, col_bulk2 = st.columns(2)
                    with col_bulk1:
                        if phones_list:
                            st.code("\n".join(phones_list), language=None)
                            st.caption(f"📋 {len(phones_list)} رقم — انسخيهم وأرسلي الرسالة يدوياً")
                    with col_bulk2:
                        st.code(bulk_msg, language=None)
                        st.caption("📋 انسخي الرسالة")

                    st.markdown("---")

                    # ── قائمة الموظفات الفردية ──────────────────
                    for eid, emp in not_checked_in.items():
                        # هل تم تذكيرها مسبقاً؟
                        reminded_key   = f"reminded_{eid}_{today_str}"
                        not_req_key    = f"not_required_{eid}_{today_str}"
                        already_reminded = st.session_state.get(reminded_key, False)
                        not_required     = st.session_state.get(not_req_key, False)

                        if not_required:
                            st.markdown(f'<div style="background:#e2e3e5;border-radius:10px;padding:8px 14px;margin-bottom:6px;font-size:12px;color:#555;font-weight:700;">🚫 غير مطلوبة اليوم — {emp.get("الاسم", "")} — #{eid}</div>', unsafe_allow_html=True)
                        elif already_reminded:
                            col_rem1, col_rem2 = st.columns(2)
                            with col_rem1:
                                st.markdown(f'<div style="background:#d4edda;border-radius:10px;padding:8px 14px;font-size:12px;color:#155724;font-weight:700;">✅ تم التذكير — {emp.get("الاسم", "")} — #{eid}</div>', unsafe_allow_html=True)
                            with col_rem2:
                                if st.button("↩️ تراجع", key=f"undo_reminded_{eid}_{today_str}", use_container_width=True):
                                    st.session_state[reminded_key] = False
                                    st.rerun()
                        else:
                            st.markdown(f'<div class="warn-row">🚨 {emp.get("الاسم", "")} — #{eid} — {emp.get("المدرسة", "")} — {emp.get("المهمة", "")}</div>', unsafe_allow_html=True)
                            phone_raw = str(emp.get("رقم التواصل", "") or "").strip().replace(" ", "")
                            msg = f"""السلام عليكم أ. {emp.get("الاسم","").split()[0]} 🌷
(الرقم الشخصي: {eid})

لم يُسجَّل حضوركِ في نظام الحضور الإلكتروني لهذا اليوم.
يُرجى التسجيل فوراً — النظام الإلكتروني هو المرجع الرسمي المعتمد.

🔗 {APP_URL}

📍 في حال عدم عمل التطبيق:
- لابتوب في الصالة الرياضية
- لابتوب في مقر الكنترول الخارجي
"""
                            cwa1, cwa2, cwa3 = st.columns(3)
                            if phone_raw:
                                if not phone_raw.startswith("973"):
                                    phone_raw = "973" + phone_raw.lstrip("0")
                                wa_url = "https://wa.me/" + phone_raw + "?text=" + urllib.parse.quote(msg)
                                with cwa1:
                                    st.link_button("📩 واتساب", wa_url, use_container_width=True)
                            else:
                                with cwa1:
                                    st.caption("لا يوجد رقم")
                            with cwa2:
                                if st.button("✅ تم التذكير", key=f"reminder_sent_{eid}_{today_str}", use_container_width=True):
                                    st.session_state[reminded_key] = True
                                    log_audit(eid, emp.get("الاسم", ""), "إرسال تذكير عدم تسجيل", "تم التذكير من الداشبورد")
                                    st.rerun()
                            with cwa3:
                                if st.button("🚫 غير مطلوبة اليوم", key=f"not_req_{eid}_{today_str}", use_container_width=True):
                                    st.session_state[not_req_key] = True
                                    log_audit(eid, emp.get("الاسم", ""), "غير مطلوبة اليوم", "تم التحديد من الداشبورد")
                                    st.rerun()
                else:
                    st.success("✅ جميع الموظفات المطلوبات سجّلن حضورهن أو تم تسجيل غيابهن.")

            if no_gps_rows:
                st.markdown("#### ⚠️ عمليات بدون تحقق GPS")
                for r in no_gps_rows[-50:]:
                    st.markdown(f'<div class="warn-row">⚠️ {r.get("الاسم الثلاثي","")} — #{r.get("الرقم الشخصي","")} — حضور: {r.get("وقت الحضور","") or "—"} — انصراف: {r.get("وقت الانصراف","") or "—"}</div>', unsafe_allow_html=True)

            if correction_done_rows:
                st.markdown("#### ✅ معفى - الانتهاء من التصحيح")
                for r in correction_done_rows:
                    st.markdown(f'<div class="audit-row">✅ {r.get("الاسم الثلاثي","")} — {r.get("المهمة","")} — حضور: {r.get("وقت الحضور","")} — انصراف: {r.get("وقت الانصراف","") or "لم يسجل"}</div>',unsafe_allow_html=True)

            if external_support_rows:
                st.markdown("#### 🟣 الدعم الخارجي غير الموجود في القائمة البيضاء")
                for r in external_support_rows:
                    st.markdown(f'<div class="audit-row">🟣 {r.get("الاسم الثلاثي","")} — #{r.get("الرقم الشخصي","")} — {r.get("اسم المدرسة",r.get("المدرسة",""))} — {r.get("المهمة","")} — حضور: {r.get("وقت الحضور","")} — انصراف: {r.get("وقت الانصراف","") or "لم يسجل"}</div>',unsafe_allow_html=True)

            if missing_depart:
                # تنبيه قبل الساعة 2 بـ 30 دقيقة
                now_t = now_bh().time()
                if time(13, 30) <= now_t <= time(14, 30):
                    st.markdown("#### ⏰ تذكير الانصراف — الوقت يقترب من الساعة 2")
                    st.warning(f"⚠️ {len(missing_depart)} موظفة لم تسجّل انصرافها — يرجى التذكير قبل انتهاء الدوام.")

                    # نسخ الكل
                    dep_reminder_msg = f"""السلام عليكم 🌷

يُرجى تسجيل الانصراف في النظام الإلكتروني قبل مغادرة المركز.
⚠️ في حال عدم التسجيل سيتم إغلاق سجلك تلقائياً بوقت الانصراف الرسمي.

🔗 {APP_URL}

📍 في حال عدم عمل التطبيق:
- لابتوب في الصالة الرياضية
- لابتوب في مقر الكنترول الخارجي"""
                    dep_phones = []
                    for r in missing_depart:
                        eid = str(r.get("الرقم الشخصي","")).strip()
                        emp_wl = get_whitelist().get(eid, {})
                        ph = str(emp_wl.get("رقم التواصل","") or "").strip().replace(" ","")
                        if ph:
                            if not ph.startswith("973"):
                                ph = "973" + ph.lstrip("0")
                            dep_phones.append(f"{r.get('الاسم الثلاثي','')} — {ph}")

                    col_dep1, col_dep2 = st.columns(2)
                    with col_dep1:
                        if dep_phones:
                            st.code("\n".join(dep_phones), language=None)
                            st.caption(f"📋 {len(dep_phones)} رقم")
                    with col_dep2:
                        st.code(dep_reminder_msg, language=None)
                        st.caption("📋 الرسالة")

                    st.markdown("---")

                st.markdown("#### 🚨 لم يسجلن الانصراف — تذكير فردي")
                for ri_dep, r in enumerate(missing_depart):
                    eid = str(r.get("الرقم الشخصي","")).strip()
                    emp_wl = get_whitelist().get(eid, {})
                    ph = str(emp_wl.get("رقم التواصل","") or "").strip().replace(" ","")
                    reminded_dep_key = f"reminded_dep_{eid}_{today_str}"
                    not_req_dep_key  = f"not_req_dep_{eid}_{today_str}"
                    already_reminded_dep = st.session_state.get(reminded_dep_key, False)
                    not_required_dep     = st.session_state.get(not_req_dep_key, False)

                    if not_required_dep:
                        st.markdown(f'<div style="background:#e2e3e5;border-radius:10px;padding:8px 14px;margin-bottom:6px;font-size:12px;color:#555;font-weight:700;">🚫 غير مطلوبة اليوم — {r.get("الاسم الثلاثي","")} — حضور: {r.get("وقت الحضور","")}</div>', unsafe_allow_html=True)
                    elif already_reminded_dep:
                        st.markdown(f'<div style="background:#d4edda;border-radius:10px;padding:8px 14px;margin-bottom:6px;font-size:12px;color:#155724;font-weight:700;">✅ تم التذكير — {r.get("الاسم الثلاثي","")} — حضور: {r.get("وقت الحضور","")}</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<div class="warn-row">🚨 {r.get("الاسم الثلاثي","")} — {r.get("اسم المدرسة",r.get("المدرسة",""))} — {r.get("المهمة","")} — حضور: {r.get("وقت الحضور","")}</div>', unsafe_allow_html=True)
                        c_dep1, c_dep2, c_dep3 = st.columns(3)
                        if ph:
                            if not ph.startswith("973"):
                                ph = "973" + ph.lstrip("0")
                            dep_wa_msg = f"""السلام عليكم أ. {str(r.get('الاسم الثلاثي','')).split()[0]} 🌷
(الرقم الشخصي: {eid})

يُرجى تسجيل الانصراف في النظام الإلكتروني قبل مغادرة المركز.
⚠️ في حال عدم التسجيل سيتم إغلاق سجلك تلقائياً بوقت الانصراف الرسمي.

🔗 {APP_URL}

📍 في حال عدم عمل التطبيق:
- لابتوب في الصالة الرياضية
- لابتوب في مقر الكنترول الخارجي"""
                            dep_wa_url = "https://wa.me/" + ph + "?text=" + urllib.parse.quote(dep_wa_msg)
                            with c_dep1:
                                st.link_button("📩 واتساب", dep_wa_url, use_container_width=True)
                        else:
                            with c_dep1:
                                st.caption("لا يوجد رقم")
                        with c_dep2:
                            if st.button("✅ تم التذكير", key=f"reminded_dep_btn_{ri_dep}_{eid}_{today_str}", use_container_width=True):
                                st.session_state[reminded_dep_key] = True
                                log_audit(eid, r.get("الاسم الثلاثي",""), "تذكير انصراف", "تم التذكير من الداشبورد")
                                st.rerun()
            # ── موظفات تكرر إغلاق سجلهن تلقائياً مرتين أو أكثر ──
            with st.container(border=True):
                st.markdown("##### 🔁 تكرار الإغلاق التلقائي")
                all_data = data  # نفس القراءة المباشرة من sheet1 بدون كاش
                auto_close_count = {}
                for r in all_data:
                    if str(r.get("إغلاق تلقائي","")).strip():
                        eid  = str(r.get("الرقم الشخصي","")).strip()
                        name = str(r.get("الاسم الثلاثي","")).strip()
                        sch  = str(r.get("اسم المدرسة","")).strip()
                        if eid:
                            auto_close_count.setdefault(eid, {"الاسم": name, "المدرسة": sch, "عدد": 0})
                            auto_close_count[eid]["عدد"] += 1

                repeated = {k:v for k,v in auto_close_count.items() if v["عدد"] >= 2}
                if not repeated:
                    st.success("✅ لا يوجد تكرار في الإغلاق التلقائي.")
                else:
                    st.warning(f"⚠️ {len(repeated)} موظفة تكرر إغلاق سجلها تلقائياً مرتين أو أكثر.")
                    for eid, info in sorted(repeated.items(), key=lambda x: -x[1]["عدد"]):
                        emp_wl = get_whitelist().get(eid, {})
                        ph = str(emp_wl.get("رقم التواصل","") or "").strip().replace(" ","")
                        st.markdown(f'<div class="warn-row">🔁 {info["الاسم"]} — #{eid} — {info["المدرسة"]} — تكرر: <b>{info["عدد"]}</b> مرة</div>', unsafe_allow_html=True)
                        if ph:
                            if not ph.startswith("973"):
                                ph = "973" + ph.lstrip("0")
                            repeat_msg = f"""السلام عليكم أ. {info["الاسم"].split()[0]} 🌷
(الرقم الشخصي: {eid})

نود إعلامكِ بأنه تم إغلاق سجل انصرافكِ تلقائياً {info["عدد"]} مرات نظراً لعدم تسجيل الانصراف يدوياً في نظام الحضور.

يُرجى الحرص على تسجيل الانصراف قبل مغادرة المركز يومياً لتجنب احتساب وقت الانصراف الرسمي تلقائياً.

🔗 {APP_URL}"""
                            repeat_url = "https://wa.me/" + ph + "?text=" + urllib.parse.quote(repeat_msg)
                            st.link_button(f"📩 إرسال تنبيه لـ {info['الاسم']}", repeat_url, use_container_width=True)
                        else:
                            st.caption("لا يوجد رقم تواصل")

            if on_leave:
                st.markdown("#### 📤 استئذان مفتوح — من الأقسام المطلوبة")
                for r in on_leave:
                    st.markdown(f'<div class="warn-row">📤 {r.get("الاسم الثلاثي","")} — خروج: {r.get("خروج استئذان","")} — لم تسجل عودة أو انصراف</div>',unsafe_allow_html=True)
            if abs_today:
                st.markdown("#### الغائبات — حسب دوام الأقسام")
                for r in abs_today:
                    st.markdown(f'<div class="absent-row">🔴 {r.get("الاسم","")} — {r.get("المدرسة","")} — {r.get("المهمة","")} — سبب: {r.get("سبب الغياب","")}</div>',unsafe_allow_html=True)
            if late_list:
                st.markdown("#### المتأخرون — حسب القواعد المعتمدة")
                for r in late_list:
                    st.markdown(f'<div class="warn-row">⏰ {r.get("الاسم الثلاثي","")} — وصل {r.get("وقت الحضور","")} — السبب: {r.get("سبب التأخير","") or "بدون"}</div>',unsafe_allow_html=True)

            # ── أسباب التأخير "أخرى" مع تعديل مباشر ──
            other_reasons = [r for r in today_rows
                             if str(r.get("سبب التأخير","")).strip()
                             and str(r.get("سبب التأخير","")).strip() not in ["دوام مرن","موعد","مهمة رسمية","رعاية","الانتهاء من التصحيح",""]
                             and "أخرى" not in str(r.get("سبب التأخير","")).strip() == False
                             or (str(r.get("سبب التأخير","")).strip().startswith("أخرى") or
                                 (str(r.get("سبب التأخير","")).strip() and
                                  str(r.get("سبب التأخير","")).strip() not in ["دوام مرن","موعد","مهمة رسمية","رعاية","الانتهاء من التصحيح"]))]

            # فلتر أبسط — كل سبب مو من القائمة الرسمية
            official_reasons = {"دوام مرن","موعد","مهمة رسمية","رعاية","الانتهاء من التصحيح",""}
            other_reasons = [r for r in today_rows
                             if r.get("وقت الحضور","")
                             and str(r.get("سبب التأخير","")).strip() not in official_reasons]

            if other_reasons:
                st.markdown("#### 📝 أسباب تأخير تحتاج مراجعة")
                st.caption("موظفات كتبن سبباً غير رسمي — يمكنك تعديل وقت حضورهن مباشرة.")
                for idx_or, r in enumerate(other_reasons):
                    eid_or  = str(r.get("الرقم الشخصي","")).strip()
                    rn_or   = None
                    for ii, rd in enumerate(data):
                        if str(rd.get("الرقم الشخصي","")).strip()==eid_or and str(rd.get("التاريخ","")).strip().replace("/","-")==today_str:
                            rn_or = ii+2; break

                    with st.expander(f"📝 {r.get('الاسم الثلاثي','')} — وصل: {r.get('وقت الحضور','')} — السبب: {r.get('سبب التأخير','')}", expanded=False):
                        col_or1, col_or2 = st.columns(2)
                        with col_or1:
                            new_att_or = st.text_input("وقت الحضور الجديد", value=str(r.get("وقت الحضور","")).strip(), key=f"or_att_{idx_or}_{eid_or}", placeholder="مثال: 07:00:00")
                        with col_or2:
                            new_rsn_or = st.selectbox("تغيير السبب", ["أبقي كما هو"] + reasons, key=f"or_rsn_{idx_or}_{eid_or}")

                        if st.button("💾 حفظ التعديل", key=f"or_save_{idx_or}_{eid_or}", use_container_width=True, type="primary"):
                            try:
                                if rn_or:
                                    if new_att_or.strip() != str(r.get("وقت الحضور","")).strip():
                                        safe_update(sheet, rn_or, COL_ATTEND, new_att_or.strip())
                                    if new_rsn_or != "أبقي كما هو":
                                        safe_update(sheet, rn_or, COL_LATE_REASON, new_rsn_or)
                                    new_row = dict(r); new_row["وقت الحضور"] = new_att_or.strip()
                                    update_work_calculation(rn_or, new_row)
                                    clear_caches()
                                    log_audit(eid_or, r.get("الاسم الثلاثي",""), "تعديل وقت حضور من الداشبورد", f"من:{r.get('وقت الحضور','')} إلى:{new_att_or}")
                                    st.success("✅ تم الحفظ.")
                                    st.rerun()
                            except Exception as e:
                                st.error(f"❌ خطأ: {e}")
            if auto_closed:
                with st.expander("🔄 سجلات أُغلقت تلقائيًا", expanded=False):
                    for r in reversed(auto_closed[-50:]):
                        st.markdown(f'<div class="audit-row">{r.get("التاريخ","")} — {r.get("الاسم الثلاثي","")} — انصراف: {r.get("وقت الانصراف","")} — {r.get("إغلاق تلقائي","")}</div>',unsafe_allow_html=True)


        # ── التقارير ──────────────────────────────────────────────
        elif admin_tab=="📑 التقارير":
            st.markdown("#### 📑 التقارير والتعديل")

            rpt_main_tab = st.selectbox("القسم", [
                "📊 التقارير",
                "✏️ بحث وتعديل سجل",
                "➕ إضافة سجل جديد",
            ], key="rpt_main_tab")

            # ══════════════════════════════════════════════
            if rpt_main_tab == "📊 التقارير":
                col_rref1, col_rref2 = st.columns([3,1])
                with col_rref2:
                    if st.button("🔄 تحديث", key="btn_refresh_rpt", use_container_width=True):
                        get_sheet_data.clear()
                        st.rerun()
                with col_rref1:
                    st.caption("⏱️ اضغطي تحديث للحصول على أحدث البيانات من الشيت.")

                with st.container(border=True):
                    st.markdown("##### 🔍 تحديد نطاق التقرير")

                    col_r1, col_r2 = st.columns(2)
                    with col_r1:
                        rpt_date_from = st.date_input("من تاريخ",
                            value=now_bh().date().replace(day=1), key="rpt_from")
                    with col_r2:
                        rpt_date_to = st.date_input("إلى تاريخ",
                            value=now_bh().date(), key="rpt_to")
                    date_from = rpt_date_from.strftime("%Y-%m-%d")
                    date_to   = rpt_date_to.strftime("%Y-%m-%d")

                    rpt_school_filter = st.multiselect("فلترة بالمدرسة (اتركي فارغاً للكل)", schools, key="rpt_sch_filter")
                    rpt_task_filter   = st.multiselect("فلترة بالمهمة (اتركي فارغاً للكل)", TASKS_ALL, key="rpt_task_filter")
                    rpt_category_filter = st.selectbox("نوع الفئة", ["الأعضاء فقط", "الدعم فقط", "الأعضاء + الدعم"], key="rpt_category_filter")

                if st.button("📊 إنشاء التقرير", use_container_width=True, type="primary", key="btn_gen_report"):
                    try:
                        data = get_sheet_data()
                        def norm_d(d): return str(d).strip().replace("/","-")
                        def parse_mins(val):
                            try:
                                if not val or str(val).strip() in ["","0:00","00:00"]: return 0
                                p = str(val).strip().split(":")
                                return int(p[0])*60 + int(p[1])
                            except: return 0
                        def fmt_m(m):
                            if m <= 0: return "—"
                            return f"{m//60}:{m%60:02d}"

                        # فلترة
                        rows = [r for r in data
                                if date_from <= norm_d(r.get("التاريخ","")) <= date_to
                                and r.get("وقت الحضور","")]
                        if rpt_school_filter:
                            rows = [r for r in rows if str(r.get("اسم المدرسة","")).strip() in rpt_school_filter]
                        if rpt_task_filter:
                            rows = [r for r in rows if str(r.get("المهمة","")).strip() in rpt_task_filter]
                        rows = filter_rows_by_category(rows, rpt_category_filter)

                        if not rows:
                            st.warning("⚠️ لا توجد بيانات للنطاق المحدد.")
                        else:
                            # بناء هيكل: مدرسة → مهمة → موظفة → أيام
                            structure = {}
                            for r in rows:
                                sch  = str(r.get("اسم المدرسة","")).strip() or "غير محدد"
                                task = str(r.get("المهمة","")).strip() or "غير محدد"
                                eid  = str(r.get("الرقم الشخصي","")).strip()
                                name = str(r.get("الاسم الثلاثي","")).strip()

                                structure.setdefault(sch, {})
                                structure[sch].setdefault(task, {})
                                structure[sch][task].setdefault(eid, {"الاسم": name, "أيام": []})
                                structure[sch][task][eid]["أيام"].append(r)

                            grand_work = 0; grand_extra = 0
                            excel_rows = []  # لتصدير Excel

                            for sch in sorted(structure.keys()):
                                st.markdown(f"""
                                <div style="background:#0c3460;color:#fff;border-radius:10px;
                                padding:10px 16px;margin:20px 0 8px 0;direction:rtl;font-size:15px;font-weight:700;">
                                🏫 {sch}
                                </div>""", unsafe_allow_html=True)

                                sch_work = 0; sch_extra = 0

                                for task in sorted(structure[sch].keys()):
                                    st.markdown(f'<div style="background:#185FA5;color:#fff;border-radius:8px;padding:6px 14px;margin:8px 0 4px 0;direction:rtl;font-size:13px;font-weight:700;">📋 {task}</div>', unsafe_allow_html=True)

                                    task_work = 0; task_extra = 0

                                    for eid, emp_data in sorted(structure[sch][task].items(), key=lambda x: x[1]["الاسم"]):
                                        emp_work = 0; emp_extra = 0
                                        day_rows = []

                                        for r in sorted(emp_data["أيام"], key=lambda x: norm_d(x.get("التاريخ",""))):
                                            wm = parse_mins(r.get("ساعات العمل",""))
                                            em = parse_mins(r.get("الساعات الإضافية",""))
                                            emp_work  += wm
                                            emp_extra += em
                                            day_rows.append({
                                                "التاريخ":    norm_d(r.get("التاريخ","")),
                                                "اليوم":      r.get("اليوم",""),
                                                "حضور":       r.get("وقت الحضور","—"),
                                                "انصراف":     r.get("وقت الانصراف","—"),
                                                "ساعات":      r.get("ساعات العمل","—"),
                                                "إضافي":      r.get("الساعات الإضافية","—") or "—",
                                                "الفئة":      employee_category_label(r),
                                                "رعاية":      "نعم" if is_care_day(r) else "لا",
                                                "نوع الدوام": r.get("نوع الدوام اليومي",""),
                                            })
                                            excel_rows.append({
                                                "المدرسة": sch, "المهمة": task,
                                                "الاسم": emp_data["الاسم"], "الرقم الشخصي": eid,
                                                "الفئة": employee_category_label(r),
                                                "التاريخ": norm_d(r.get("التاريخ","")),
                                                "اليوم": r.get("اليوم",""),
                                                "وقت الحضور": r.get("وقت الحضور",""),
                                                "وقت الانصراف": r.get("وقت الانصراف",""),
                                                "ساعات العمل": r.get("ساعات العمل",""),
                                                "الساعات الإضافية": r.get("الساعات الإضافية","") or "",
                                                "رعاية": "نعم" if is_care_day(r) else "لا",
                                                "نوع الدوام": r.get("نوع الدوام اليومي",""),
                                            })

                                        task_work  += emp_work
                                        task_extra += emp_extra

                                        with st.expander(f"👤 {emp_data['الاسم']} — {fmt_m(emp_work)} عمل | {fmt_m(emp_extra)} إضافي", expanded=False):
                                            df_emp = pd.DataFrame(day_rows)
                                            st.dataframe(df_emp, use_container_width=True, hide_index=True)
                                            c1,c2,c3 = st.columns(3)
                                            c1.metric("عدد الأيام", len(day_rows))
                                            c2.metric("إجمالي ساعات العمل", fmt_m(emp_work))
                                            c3.metric("إجمالي الإضافي", fmt_m(emp_extra))

                                    sch_work  += task_work
                                    sch_extra += task_extra
                                    st.markdown(f'<div style="background:#f0f4f8;border-radius:8px;padding:6px 14px;font-size:12px;direction:rtl;color:#444;">إجمالي {task}: ساعات عمل <b>{fmt_m(task_work)}</b> | إضافي <b>{fmt_m(task_extra)}</b></div>', unsafe_allow_html=True)

                                grand_work  += sch_work
                                grand_extra += sch_extra
                                st.markdown(f'<div style="background:#0c3460;color:#fff;border-radius:8px;padding:8px 14px;font-size:13px;direction:rtl;margin-bottom:8px;">إجمالي {sch}: ساعات عمل <b>{fmt_m(sch_work)}</b> | إضافي <b>{fmt_m(sch_extra)}</b></div>', unsafe_allow_html=True)

                            # الإجمالي الكلي
                            st.markdown("---")
                            c1,c2,c3 = st.columns(3)
                            c1.metric("إجمالي السجلات", len(rows))
                            c2.metric("إجمالي ساعات العمل الكلي", fmt_m(grand_work))
                            c3.metric("إجمالي الساعات الإضافية الكلي", fmt_m(grand_extra))

                            # ── تصدير Excel ──
                            st.markdown("---")
                            try:
                                from openpyxl import load_workbook
                                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                                from openpyxl.utils import get_column_letter

                                buf = BytesIO()
                                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                                    pd.DataFrame(excel_rows).to_excel(writer, index=False, sheet_name="التقرير التفصيلي")

                                buf.seek(0)
                                wb = load_workbook(buf)
                                ws = wb["التقرير التفصيلي"]
                                ws.sheet_view.rightToLeft = True
                                ws.freeze_panes = "A2"

                                hf = PatternFill("solid", fgColor="0C3460")
                                hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                                bfont = Font(name="Arial", size=10)
                                af = PatternFill("solid", fgColor="F5F5F5")
                                wf = PatternFill("solid", fgColor="FFFFFF")
                                thin = Side(style="thin", color="CCCCCC")
                                brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
                                ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
                                rgt  = Alignment(horizontal="right",  vertical="center", wrap_text=True, readingOrder=2)

                                col_widths = {"المدرسة":28,"المهمة":30,"الاسم":22,"الرقم الشخصي":16,
                                              "التاريخ":14,"اليوم":10,"وقت الحضور":13,"وقت الانصراف":13,
                                              "ساعات العمل":13,"الساعات الإضافية":16,"رعاية":10,"نوع الدوام":20}

                                for cell in ws[1]:
                                    cell.font=hfont; cell.fill=hf; cell.alignment=ctr; cell.border=brd
                                    hv = str(cell.value or "")
                                    ws.column_dimensions[get_column_letter(cell.column)].width = col_widths.get(hv, 16)
                                ws.row_dimensions[1].height = 28

                                prev_sch = prev_task = prev_name = ""
                                for ri, row_cells in enumerate(ws.iter_rows(min_row=2), 2):
                                    ws.row_dimensions[ri].height = 18
                                    cur_sch  = str(ws.cell(ri,1).value or "")
                                    cur_task = str(ws.cell(ri,2).value or "")
                                    cur_name = str(ws.cell(ri,3).value or "")

                                    # لون حسب المجموعة
                                    if cur_sch != prev_sch:
                                        row_fill = PatternFill("solid", fgColor="D6E4F0")
                                    elif cur_task != prev_task:
                                        row_fill = PatternFill("solid", fgColor="EAF3DE")
                                    elif cur_name != prev_name:
                                        row_fill = wf if ri%2==0 else af
                                    else:
                                        row_fill = wf if ri%2==0 else af

                                    prev_sch = cur_sch; prev_task = cur_task; prev_name = cur_name

                                    for cell in row_cells:
                                        cell.font=bfont; cell.fill=row_fill; cell.border=brd
                                        hv = str(ws.cell(1,cell.column).value or "")
                                        cell.alignment = ctr if hv in ["التاريخ","اليوم","وقت الحضور","وقت الانصراف","ساعات العمل","الساعات الإضافية","رعاية"] else rgt
                                        if hv == "الرقم الشخصي":
                                            cell.value=str(cell.value or ""); cell.number_format="@"

                                ws.page_setup.orientation="landscape"; ws.page_setup.paperSize=9
                                ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1
                                ws.print_title_rows="1:1"
                                ws.oddHeader.center.text = f"مركز جدحفص الثانوية للتصحيح المركزي\nنظام الحضور والانصراف — {date_from} إلى {date_to}"
                                ws.oddHeader.center.font = "Arial,Bold"
                                ws.oddHeader.right.text  = "تصميم وبرمجة: أ. عفاف حسين"
                                ws.oddFooter.right.text  = "صفحة &P من &N"
                                ws.oddFooter.left.text   = "رئيسة المركز: أ. خلود يعقوب بدو"
                                ws.oddFooter.left.font   = "Arial,Bold"

                                buf2 = BytesIO(); wb.save(buf2); buf2.seek(0)
                                st.download_button("📥 تحميل Excel — منسق وجاهز للطباعة",
                                    data=buf2,
                                    file_name=f"تقرير_تفصيلي_{date_from}_{date_to}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True)
                            except Exception as e:
                                st.warning(f"⚠️ تعذّر إنشاء Excel: {e}")

                    except Exception as e:
                        st.error(f"❌ خطأ: {e}")

                rpt_view = st.radio("نوع العرض", [
                    "📊 تقرير مدرسة", "👤 تقرير معلمة", "🏫 كل المدارس"
                ], horizontal=True, key="rpt_view")

                col_r1, col_r2 = st.columns(2)
                with col_r1:
                    rpt_date_from = st.date_input("من تاريخ",
                        value=now_bh().date().replace(day=1), key="rpt_from_legacy")
                with col_r2:
                    rpt_date_to = st.date_input("إلى تاريخ",
                        value=now_bh().date(), key="rpt_to_legacy")
                date_from = rpt_date_from.strftime("%Y-%m-%d")
                date_to   = rpt_date_to.strftime("%Y-%m-%d")
                rpt_category_filter2 = st.selectbox("نوع الفئة", ["الأعضاء فقط", "الدعم فقط", "الأعضاء + الدعم"], key="rpt_category_filter2")

                if rpt_view == "📊 تقرير مدرسة":
                    rpt_school = st.selectbox("المدرسة", schools, key="rpt_school_sel")
                    rpt_emp_id = None
                elif rpt_view == "👤 تقرير معلمة":
                    rpt_emp_raw = st.text_input("الرقم الشخصي أو الاسم", key="rpt_emp_input")
                    rpt_emp_id  = ar_to_en_digits(rpt_emp_raw).strip()
                    rpt_school  = None
                    if rpt_emp_id:
                        f_emp = validate_employee(rpt_emp_id)
                        if f_emp:
                            st.success(f"✅ {f_emp.get('الاسم','')} — {f_emp.get('المدرسة','')}")
                        else:
                            # بحث بالاسم
                            wl_all = get_whitelist()
                            matches = [(eid,e) for eid,e in wl_all.items()
                                       if rpt_emp_raw.strip() in str(e.get("الاسم",""))]
                            if matches:
                                names = [f"{e.get('الاسم','')} (#{eid})" for eid,e in matches[:5]]
                                sel   = st.selectbox("اختاري الموظفة", names, key="rpt_emp_match")
                                rpt_emp_id = sel.split("(#")[-1].rstrip(")")
                else:
                    rpt_school = "الكل"
                    rpt_emp_id = None

            if st.button("📊 إنشاء التقرير", use_container_width=True, type="primary", key="btn_gen_report_legacy"):
                try:
                    data = get_sheet_data()
                    def norm_date(d): return str(d).strip().replace("/","-")

                    # فلترة البيانات
                    rows = [r for r in data
                            if date_from <= norm_date(r.get("التاريخ","")) <= date_to
                            and r.get("وقت الحضور","")]

                    if rpt_view == "📊 تقرير مدرسة":
                        rows = [r for r in rows if str(r.get("اسم المدرسة","")).strip() == rpt_school]
                    elif rpt_view == "👤 تقرير معلمة" and rpt_emp_id:
                        rows = [r for r in rows if str(r.get("الرقم الشخصي","")).strip() == rpt_emp_id]
                    rows = filter_rows_by_category(rows, rpt_category_filter2)

                    if not rows:
                        st.warning("⚠️ لا توجد بيانات للنطاق المحدد.")
                    else:
                        def parse_hours(val):
                            try:
                                if not val or str(val).strip() in ["","0:00","00:00"]: return 0
                                parts = str(val).strip().split(":")
                                return int(parts[0])*60 + int(parts[1])
                            except: return 0

                        def fmt_mins(m):
                            if m <= 0: return "—"
                            return f"{m//60}:{m%60:02d}"

                        # ══════════════════════════════
                        # تقرير معلمة واحدة
                        # ══════════════════════════════
                        if rpt_view == "👤 تقرير معلمة":
                            emp_name   = str(rows[0].get("الاسم الثلاثي","")).strip()
                            emp_school = str(rows[0].get("اسم المدرسة","")).strip()
                            emp_task   = str(rows[0].get("المهمة","")).strip()

                            st.markdown(f"""
                            <div style="background:#0c3460;color:#fff;border-radius:12px;padding:14px 18px;margin-bottom:16px;direction:rtl;">
                            <b>👤 {emp_name}</b> — {emp_school} — {emp_task}<br>
                            <small>من {date_from} إلى {date_to}</small>
                            </div>
                            """, unsafe_allow_html=True)

                            total_work = 0; total_extra = 0; total_days = 0
                            df_rows = []
                            for r in sorted(rows, key=lambda x: norm_date(x.get("التاريخ",""))):
                                work_m  = parse_hours(r.get("ساعات العمل",""))
                                extra_m = parse_hours(r.get("الساعات الإضافية",""))
                                total_work  += work_m
                                total_extra += extra_m
                                total_days  += 1
                                df_rows.append({
                                    "التاريخ":          norm_date(r.get("التاريخ","")),
                                    "اليوم":            r.get("اليوم",""),
                                    "وقت الحضور":       r.get("وقت الحضور","—"),
                                    "وقت الانصراف":     r.get("وقت الانصراف","—"),
                                    "ساعات العمل":      r.get("ساعات العمل","—"),
                                    "الساعات الإضافية": r.get("الساعات الإضافية","—") or "—",
                                    "رعاية":            "نعم" if is_care_day(r) else "لا",
                                    "نوع الدوام":       r.get("نوع الدوام اليومي",""),
                                    "حالة الدوام":      r.get("حالة الدوام",""),
                                })

                            df = pd.DataFrame(df_rows)
                            st.dataframe(df, use_container_width=True, hide_index=True)

                            c1,c2,c3 = st.columns(3)
                            c1.metric("عدد الأيام", total_days)
                            c2.metric("إجمالي ساعات العمل", fmt_mins(total_work))
                            c3.metric("إجمالي الساعات الإضافية", fmt_mins(total_extra))

                        # ══════════════════════════════
                        # تقرير مدرسة أو كل المدارس
                        # ══════════════════════════════
                        else:
                            # تجميع حسب المدرسة ثم الموظفة
                            school_emp = {}
                            for r in rows:
                                sch = str(r.get("اسم المدرسة","")).strip() or "غير محدد"
                                eid = str(r.get("الرقم الشخصي","")).strip()
                                if sch not in school_emp:
                                    school_emp[sch] = {}
                                if eid not in school_emp[sch]:
                                    school_emp[sch][eid] = {
                                        "الاسم":    str(r.get("الاسم الثلاثي","")).strip(),
                                        "المهمة":   str(r.get("المهمة","")).strip(),
                                        "أيام":     0,
                                        "work_m":   0,
                                        "extra_m":  0,
                                        "تواريخ":   [],
                                    }
                                school_emp[sch][eid]["أيام"]   += 1
                                school_emp[sch][eid]["work_m"]  += parse_hours(r.get("ساعات العمل",""))
                                school_emp[sch][eid]["extra_m"] += parse_hours(r.get("الساعات الإضافية",""))
                                school_emp[sch][eid]["تواريخ"].append(norm_date(r.get("التاريخ","")))

                            grand_total_extra = 0

                            for sch in sorted(school_emp.keys()):
                                emps = school_emp[sch]
                                sch_extra = sum(e["extra_m"] for e in emps.values())
                                grand_total_extra += sch_extra

                                st.markdown(f"""
                                <div style="background:#0c3460;color:#fff;border-radius:10px;padding:10px 16px;margin:16px 0 8px 0;direction:rtl;">
                                <b>🏫 {sch}</b> — {len(emps)} معلمة — إجمالي إضافي: <b>{fmt_mins(sch_extra)}</b>
                                </div>
                                """, unsafe_allow_html=True)

                                sch_rows = []
                                for eid, emp in sorted(emps.items(), key=lambda x: x[1]["الاسم"]):
                                    sch_rows.append({
                                        "الاسم":              emp["الاسم"],
                                        "الرقم الشخصي":       eid,
                                        "المهمة":             emp["المهمة"],
                                        "عدد الأيام":         emp["أيام"],
                                        "إجمالي ساعات العمل": fmt_mins(emp["work_m"]),
                                        "الساعات الإضافية":   fmt_mins(emp["extra_m"]),
                                    })
                                st.dataframe(pd.DataFrame(sch_rows), use_container_width=True, hide_index=True)

                            if rpt_view == "🏫 كل المدارس":
                                st.metric("إجمالي الساعات الإضافية — كل المدارس", fmt_mins(grand_total_extra))

                        # ── تصدير Excel ──────────────────────────────
                        st.markdown("---")
                        try:
                            from openpyxl import load_workbook
                            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                            from openpyxl.utils import get_column_letter

                            buf = BytesIO()
                            if rpt_view == "👤 تقرير معلمة":
                                export_rows = df_rows
                            else:
                                export_rows = []
                                for sch in sorted(school_emp.keys()):
                                    for eid, emp in sorted(school_emp[sch].items(), key=lambda x: x[1]["الاسم"]):
                                        export_rows.append({
                                            "المدرسة":            sch,
                                            "الاسم":              emp["الاسم"],
                                            "الرقم الشخصي":       eid,
                                            "المهمة":             emp["المهمة"],
                                            "عدد الأيام":         emp["أيام"],
                                            "إجمالي ساعات العمل": fmt_mins(emp["work_m"]),
                                            "الساعات الإضافية":   fmt_mins(emp["extra_m"]),
                                        })

                            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                                pd.DataFrame(export_rows).to_excel(writer, index=False, sheet_name="التقرير")

                            buf.seek(0)
                            wb = load_workbook(buf)
                            ws = wb["التقرير"]
                            ws.sheet_view.rightToLeft = True
                            ws.freeze_panes = "A2"

                            hdr_fill = PatternFill("solid", fgColor="0C3460")
                            hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                            body_font= Font(name="Arial", size=10)
                            alt_fill = PatternFill("solid", fgColor="F5F5F5")
                            wht_fill = PatternFill("solid", fgColor="FFFFFF")
                            ctr = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
                            rgt = Alignment(horizontal="right",  vertical="center", wrap_text=True, readingOrder=2)
                            thin= Side(style="thin", color="CCCCCC")
                            brd = Border(left=thin, right=thin, top=thin, bottom=thin)

                            for cell in ws[1]:
                                cell.font=hdr_font; cell.fill=hdr_fill
                                cell.alignment=ctr; cell.border=brd
                                ws.column_dimensions[get_column_letter(cell.column)].width = 18
                            ws.row_dimensions[1].height = 28

                            for ri, row_cells in enumerate(ws.iter_rows(min_row=2), 2):
                                ws.row_dimensions[ri].height = 18
                                fl = alt_fill if ri%2==0 else wht_fill
                                for cell in row_cells:
                                    cell.font=body_font; cell.fill=fl; cell.border=brd
                                    cell.alignment=ctr if str(ws.cell(1,cell.column).value or "") in [
                                        "عدد الأيام","إجمالي ساعات العمل","الساعات الإضافية",
                                        "وقت الحضور","وقت الانصراف","ساعات العمل","التاريخ","اليوم","رعاية"
                                    ] else rgt
                                    if str(ws.cell(1,cell.column).value or "") == "الرقم الشخصي":
                                        cell.value=str(cell.value or ""); cell.number_format="@"

                            ws.page_setup.orientation="landscape"; ws.page_setup.paperSize=9
                            ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1
                            ws.print_title_rows="1:1"
                            ws.oddHeader.center.text = f"مركز جدحفص الثانوية للتصحيح المركزي\nنظام الحضور والانصراف — {date_from} إلى {date_to}"
                            ws.oddHeader.center.font = "Arial,Bold"
                            ws.oddHeader.right.text  = "تصميم وبرمجة: أ. عفاف حسين"
                            ws.oddFooter.right.text  = "صفحة &P من &N"
                            ws.oddFooter.left.text   = "رئيسة المركز: أ. خلود يعقوب بدو"
                            ws.oddFooter.left.font   = "Arial,Bold"

                            buf2 = BytesIO()
                            wb.save(buf2); buf2.seek(0)
                            st.download_button("📥 تحميل Excel — منسق وجاهز للطباعة",
                                data=buf2,
                                file_name=f"تقرير_{date_from}_{date_to}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True)
                        except Exception as e:
                            st.warning(f"⚠️ تعذّر إنشاء Excel: {e}")

                except Exception as e:
                    st.error(f"❌ خطأ في إنشاء التقرير: {e}")

            # ══════════════════════════════════════════════
            elif rpt_main_tab == "✏️ بحث وتعديل سجل":
                st.markdown("##### 🔍 بحث وتعديل سجل في شيت1")

                with st.container(border=True):
                    st.markdown("##### اختيار سريع من القائمة البيضاء")
                    sel_school_edit = st.selectbox("اختاري المدرسة", ["الكل"] + schools, key="edit_sel_school")
                    edit_tasks_available = sorted(set(str(e.get("المهمة", "")).strip() for e in get_whitelist().values() if str(e.get("المهمة", "")).strip() and (sel_school_edit == "الكل" or str(e.get("المدرسة", "")).strip() == sel_school_edit)))
                    sel_task_edit = st.selectbox("اختاري المهمة", ["الكل"] + edit_tasks_available, key="edit_sel_task")
                    edit_options = whitelist_options_by_filters(sel_school_edit, sel_task_edit)
                    edit_labels = [x[2] for x in edit_options]
                    selected_edit_label = st.selectbox("اختاري الاسم", ["اختاري"] + edit_labels, key="edit_sel_emp")
                    edit_search_term = st.text_input("أو ابحثي بالاسم / الرقم الشخصي", key="edit_smart_search")
                    edit_search_matches = find_whitelist_matches(edit_search_term) if edit_search_term else []
                    selected_search_label = st.selectbox("نتائج البحث", ["اختاري"] + [x[2] for x in edit_search_matches], key="edit_search_result") if edit_search_matches else "اختاري"
                    col_q1, col_q2 = st.columns(2)
                    with col_q1:
                        quick_from = st.date_input("من تاريخ", value=now_bh().date().replace(day=1), key="edit_quick_from")
                    with col_q2:
                        quick_to = st.date_input("إلى تاريخ", value=now_bh().date(), key="edit_quick_to")
                    if st.button("🔍 عرض سجلات المختارة", use_container_width=True, type="primary", key="btn_edit_quick_search"):
                        chosen = selected_search_label if selected_search_label != "اختاري" else selected_edit_label
                        if chosen == "اختاري":
                            st.error("❌ اختاري اسمًا أو ابحثي بالرقم/الاسم.")
                        else:
                            chosen_id = chosen.split("#")[-1].split("—")[0].strip()
                            ef = quick_from.strftime("%Y-%m-%d")
                            et = quick_to.strftime("%Y-%m-%d")
                            data = get_sheet_data_fresh()
                            results = []
                            for i, r in enumerate(data):
                                d = str(r.get("التاريخ", "")).strip().replace("/", "-")
                                if ef <= d <= et and str(r.get("الرقم الشخصي", "")).strip() == chosen_id:
                                    results.append((i+2, r))
                            if results:
                                st.session_state.edit_results = results
                                st.success(f"✅ وجد {len(results)} سجل.")
                            else:
                                st.warning("⚠️ لا توجد سجلات لهذه الموظفة في النطاق المحدد.")

                with st.container(border=True):
                    search_by = st.radio("البحث بـ", ["رقم شخصي","اسم","مدرسة"], horizontal=True, key="edit_search_by")
                    search_val = st.text_input("أدخلي قيمة البحث", key="edit_search_val")

                    col_ed1, col_ed2 = st.columns(2)
                    with col_ed1:
                        edit_date_from = st.date_input("من تاريخ", value=now_bh().date().replace(day=1), key="edit_from")
                    with col_ed2:
                        edit_date_to = st.date_input("إلى تاريخ", value=now_bh().date(), key="edit_to")

                    if st.button("🔍 بحث", use_container_width=True, type="primary", key="btn_edit_search"):
                        if not search_val.strip():
                            st.error("❌ أدخلي قيمة البحث.")
                        else:
                            data = get_sheet_data_fresh()
                            sv   = search_val.strip()
                            ef   = edit_date_from.strftime("%Y-%m-%d")
                            et   = edit_date_to.strftime("%Y-%m-%d")

                            def norm_d(d): return str(d).strip().replace("/","-")

                            results = []
                            for i, r in enumerate(data):
                                if not (ef <= norm_d(r.get("التاريخ","")) <= et):
                                    continue
                                if search_by == "رقم شخصي":
                                    match = ar_to_en_digits(sv) == str(r.get("الرقم الشخصي","")).strip()
                                elif search_by == "اسم":
                                    match = sv in str(r.get("الاسم الثلاثي",""))
                                else:
                                    match = sv in str(r.get("اسم المدرسة",""))
                                if match:
                                    results.append((i+2, r))

                            if not results:
                                st.warning("⚠️ لا توجد نتائج.")
                            else:
                                st.success(f"✅ وجد {len(results)} سجل.")
                                st.session_state.edit_results = results

                if st.session_state.get("edit_results"):
                    results = st.session_state.edit_results
                    st.markdown(f"**{len(results)} سجل — اختاري السجل للتعديل:**")

                    # عرض النتائج
                    for rn, r in results:
                        label = f"{r.get('التاريخ','')} | {r.get('الاسم الثلاثي','')} | {r.get('اسم المدرسة','')} | حضور: {r.get('وقت الحضور','—')} | انصراف: {r.get('وقت الانصراف','—')}"
                        with st.expander(label, expanded=False):
                            st.markdown(f"**الصف:** {rn}")

                            col1, col2 = st.columns(2)
                            with col1:
                                new_name    = st.text_input("الاسم",           value=str(r.get("الاسم الثلاثي","")).strip(),  key=f"en_{rn}")
                                new_school  = st.text_input("المدرسة",         value=str(r.get("اسم المدرسة","")).strip(),    key=f"es_{rn}")
                                new_task    = st.text_input("المهمة",          value=str(r.get("المهمة","")).strip(),         key=f"et_{rn}")
                                new_att     = st.text_input("وقت الحضور",      value=str(r.get("وقت الحضور","")).strip(),     key=f"ea_{rn}")
                                new_att_rsn = st.text_input("سبب التأخير",     value=str(r.get("سبب التأخير","")).strip(),    key=f"ear_{rn}")
                            with col2:
                                new_dep     = st.text_input("وقت الانصراف",    value=str(r.get("وقت الانصراف","")).strip(),   key=f"ed_{rn}")
                                new_dep_rsn = st.text_input("سبب الانصراف",    value=str(r.get("سبب الانصراف","")).strip(),   key=f"edr_{rn}")
                                new_exit    = st.text_input("خروج استئذان",    value=str(r.get("خروج استئذان","")).strip(),   key=f"ee_{rn}")
                                new_return  = st.text_input("عودة",            value=str(r.get("عودة","")).strip(),           key=f"er_{rn}")
                                new_day_type= st.text_input("نوع الدوام اليومي", value=str(r.get("نوع الدوام اليومي","")).strip(), key=f"edy_{rn}")
                                care_current = "نعم" if is_care_day(r) else "لا"
                                new_care_conf = st.selectbox("هل لديها رعاية؟", ["لا", "نعم"], index=1 if care_current == "نعم" else 0, key=f"ecare_{rn}")

                            col_s1, col_s2 = st.columns(2)
                            with col_s1:
                                if st.button("💾 حفظ التعديلات", use_container_width=True, type="primary", key=f"save_{rn}"):
                                    try:
                                        updates = {
                                            COL_NAME:          new_name,
                                            3:                 new_school,
                                            4:                 new_task,
                                            COL_ATTEND:        new_att,
                                            COL_LATE_REASON:   new_att_rsn,
                                            COL_DEPART:        new_dep,
                                            COL_DEPART_REASON: new_dep_rsn,
                                            COL_EXIT:          new_exit,
                                            COL_RETURN:        new_return,
                                            COL_CARE_CONF:     new_care_conf,
                                            COL_DAILY_TYPE:    new_day_type,
                                        }
                                        for col, val in updates.items():
                                            safe_update(sheet, rn, col, val)
                                        # إعادة حساب الساعات
                                        new_row = dict(r)
                                        new_row.update({
                                            "الاسم الثلاثي": new_name,
                                            "اسم المدرسة":   new_school,
                                            "المهمة":        new_task,
                                            "وقت الحضور":   new_att,
                                            "سبب التأخير":  new_att_rsn,
                                            "وقت الانصراف": new_dep,
                                            "سبب الانصراف": new_dep_rsn,
                                            "خروج استئذان": new_exit,
                                            "عودة":         new_return,
                                            "تأكيد الرعاية": new_care_conf,
                                            "نوع الدوام اليومي": new_day_type,
                                        })
                                        update_work_calculation(rn, new_row)
                                        clear_caches()
                                        log_audit(str(r.get("الرقم الشخصي","")), new_name, "تعديل سجل من التقارير", f"تاريخ:{r.get('التاريخ','')}")
                                        st.success("✅ تم الحفظ وإعادة حساب الساعات.")
                                        st.session_state.edit_results = None
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"❌ خطأ: {e}")
                            with col_s2:
                                if st.button("🗑️ حذف هذا السجل", use_container_width=True, key=f"del_{rn}"):
                                    confirm_key = f"confirm_del_{rn}"
                                    st.session_state[confirm_key] = True

                                if st.session_state.get(f"confirm_del_{rn}"):
                                    st.warning("⚠️ هل أنتِ متأكدة من الحذف؟")
                                    if st.button("نعم، احذف", key=f"yes_del_{rn}", use_container_width=True):
                                        try:
                                            sheet.delete_rows(rn)
                                            clear_caches()
                                            log_audit(str(r.get("الرقم الشخصي","")), str(r.get("الاسم الثلاثي","")), "حذف سجل من التقارير", f"تاريخ:{r.get('التاريخ','')}")
                                            st.success("✅ تم الحذف.")
                                            st.session_state.edit_results = None
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"❌ خطأ: {e}")

            # ══════════════════════════════════════════════
            elif rpt_main_tab == "➕ إضافة سجل جديد":
                st.markdown("##### ➕ إضافة سجل كامل لموظفة")
                st.caption("لو سجّلت الموظفة ورقياً وما سجّلت إلكترونياً، أضيفي سجلها هنا.")

                with st.container(border=True):
                    st.markdown("##### اختيار الموظفة")
                    add_school_filter = st.selectbox("اختاري المدرسة", ["الكل"] + schools, key="add_sel_school")
                    add_tasks_available = sorted(set(str(e.get("المهمة", "")).strip() for e in get_whitelist().values() if str(e.get("المهمة", "")).strip() and (add_school_filter == "الكل" or str(e.get("المدرسة", "")).strip() == add_school_filter)))
                    add_task_filter = st.selectbox("اختاري المهمة", ["الكل"] + add_tasks_available, key="add_sel_task")
                    add_options = whitelist_options_by_filters(add_school_filter, add_task_filter)
                    add_labels = [x[2] for x in add_options]
                    selected_add_label = st.selectbox("اختاري الاسم", ["اختاري"] + add_labels, key="add_sel_emp")
                    add_search_term = st.text_input("أو ابحثي بالاسم / الرقم الشخصي", key="add_smart_search")
                    add_search_matches = find_whitelist_matches(add_search_term) if add_search_term else []
                    selected_add_search_label = st.selectbox("نتائج البحث", ["اختاري"] + [x[2] for x in add_search_matches], key="add_search_result") if add_search_matches else "اختاري"

                    chosen_add_label = selected_add_search_label if selected_add_search_label != "اختاري" else selected_add_label
                    add_id = ""
                    add_emp = None
                    if chosen_add_label != "اختاري":
                        add_id = chosen_add_label.split("#")[-1].split("—")[0].strip()
                        add_emp = validate_employee(add_id)

                    if add_emp:
                        st.success(f"✅ تم اختيار: {add_emp.get('الاسم','')} — {add_emp.get('المدرسة','')} — {add_emp.get('المهمة','')}")
                        add_name   = add_emp.get("الاسم","")
                        add_school = add_emp.get("المدرسة","")
                        add_task   = add_emp.get("المهمة","")
                        add_sup    = "نعم" if is_support_employee_record(add_emp) else "لا"
                    else:
                        add_id_raw = st.text_input("الرقم الشخصي *", key="add_id_manual")
                        add_id = ar_to_en_digits(add_id_raw).strip()
                        add_emp_manual = validate_employee(add_id) if add_id else None
                        if add_emp_manual:
                            st.success(f"✅ {add_emp_manual.get('الاسم','')} — {add_emp_manual.get('المدرسة','')} — {add_emp_manual.get('المهمة','')}")
                            add_name   = add_emp_manual.get("الاسم","")
                            add_school = add_emp_manual.get("المدرسة","")
                            add_task   = add_emp_manual.get("المهمة","")
                            add_sup    = "نعم" if is_support_employee_record(add_emp_manual) else "لا"
                        else:
                            if add_id: st.warning("⚠️ الرقم غير موجود — أدخلي البيانات يدوياً.")
                            add_name   = st.text_input("الاسم الثلاثي *", key="add_name")
                            add_school = st.selectbox("المدرسة *", schools + ["أخرى"], key="add_school")
                            if add_school == "أخرى":
                                add_school = st.text_input("اكتبي اسم المدرسة", key="add_school_other").strip()
                            add_task   = st.selectbox("المهمة *", TASKS_ALL, key="add_task")
                            add_sup    = "لا"

                    add_date = st.date_input("التاريخ *", value=now_bh().date(), key="add_date")
                    add_date_str = add_date.strftime("%Y-%m-%d")
                    add_day  = add_date.strftime("%A")

                    col_a1, col_a2 = st.columns(2)
                    with col_a1:
                        add_att     = st.text_input("وقت الحضور (مثال: 07:00:00)", key="add_att")
                        add_att_rsn = st.text_input("سبب التأخير (اختياري)", key="add_att_rsn")
                        add_exit    = st.text_input("خروج استئذان (اختياري)", key="add_exit")
                    with col_a2:
                        add_dep     = st.text_input("وقت الانصراف (مثال: 14:00:00)", key="add_dep")
                        add_dep_rsn = st.text_input("سبب الانصراف (اختياري)", key="add_dep_rsn")
                        add_return  = st.text_input("عودة من استئذان (اختياري)", key="add_return")

                    add_care = st.selectbox("هل لديها رعاية؟", ["لا", "نعم"], key="add_care")
                    add_note = st.text_input("ملاحظة", value="تسجيل يدوي من التقارير", key="add_note")

                    if st.button("➕ إضافة السجل", use_container_width=True, type="primary", key="btn_add_rec"):
                        if not add_id or not add_name or not add_att:
                            st.error("❌ الرقم الشخصي والاسم ووقت الحضور مطلوبة.")
                        else:
                            try:
                                care_reason = "رعاية" if add_care == "نعم" else add_att_rsn
                                care_confirm = "نعم" if add_care == "نعم" else ""
                                daily_type_to_save = "رعاية" if add_care == "نعم" else ""
                                new_row_data = [
                                    add_date_str, add_day,
                                    add_school, add_task, add_sup,
                                    add_name, add_id,
                                    add_att, care_reason,
                                    add_dep, add_dep_rsn,
                                    add_exit, add_return,
                                    "", "", "", "", "", "", daily_type_to_save, "", care_confirm, "تسجيل يدوي من التقارير"
                                ]
                                safe_append(sheet, new_row_data)
                                # إعادة حساب الساعات
                                _idx2, _row2 = find_today_row_fresh(add_date_str, add_id)
                                if _idx2: update_work_calculation(_idx2, _row2)
                                clear_caches()
                                log_audit(add_id, add_name, "إضافة سجل يدوي من التقارير", f"تاريخ:{add_date_str}")
                                st.success(f"✅ تم إضافة السجل بنجاح — {add_name} — {add_date_str}")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ خطأ: {e}")

        # ── التواصل والمتابعة ─────────────────────────────────────
        elif admin_tab=="📞 التواصل والمتابعة":
            st.markdown("#### 📞 التواصل مع الموظفة / متابعة السجلات المضافة")
            st.info("ابحثي بالاسم أو الرقم الشخصي لفتح واتساب برسالة جاهزة، أو راجعي السجلات التي أُضيفت يدوياً.")

            wl_all = get_whitelist()

            def _phone_to_wa(phone):
                ph = str(phone or "").strip().replace(" ", "").replace("-", "")
                if not ph:
                    return ""
                ph = ar_to_en_digits(ph)
                if ph.startswith("00"):
                    ph = ph[2:]
                if ph.startswith("+"):
                    ph = ph[1:]
                if not ph.startswith("973"):
                    ph = "973" + ph.lstrip("0")
                return ph

            def _wa_button_for_emp(phone, msg):
                ph = _phone_to_wa(phone)
                if ph:
                    wa_url = "https://wa.me/" + ph + "?text=" + urllib.parse.quote(msg)
                    st.link_button("📩 فتح واتساب", wa_url, use_container_width=True)
                else:
                    st.warning("⚠️ لا يوجد رقم تواصل محفوظ لهذه الموظفة في القائمة البيضاء.")

            contact_tab = st.radio("اختاري القسم", ["🔍 بحث سريع", "📝 السجلات المضافة يدويًا"], horizontal=True, key="contact_main_tab")

            if contact_tab == "🔍 بحث سريع":
                with st.container(border=True):
                    st.markdown("##### 🔍 بحث بالاسم أو الرقم الشخصي")
                    search_txt = st.text_input("اكتبي الاسم أو الرقم الشخصي", key="contact_search_txt", placeholder="مثال: 123456789 أو عفاف")

                    matches = []
                    if search_txt.strip():
                        sv_raw = search_txt.strip()
                        sv_num = ar_to_en_digits(sv_raw)
                        sv_name = normalize_name(sv_raw)
                        for eid, emp in wl_all.items():
                            emp_name0 = str(emp.get("الاسم", "")).strip()
                            emp_name_norm = normalize_name(emp_name0)
                            if sv_num in str(eid).strip() or sv_name in emp_name_norm or sv_raw in emp_name0:
                                matches.append((eid, emp))

                    if search_txt.strip() and not matches:
                        st.warning("⚠️ لا توجد نتيجة مطابقة في القائمة البيضاء.")

                    if matches:
                        labels = []
                        label_map = {}
                        for eid, emp in matches[:50]:
                            label = f"{emp.get('الاسم','')} — #{eid} — {emp.get('المدرسة','')} — {emp.get('المهمة','')}"
                            labels.append(label)
                            label_map[label] = (eid, emp)
                        selected = st.selectbox("اختاري الموظفة", labels, key="contact_selected_emp")
                        eid, emp = label_map[selected]
                        emp_name = str(emp.get("الاسم", "")).strip()
                        school = str(emp.get("المدرسة", "")).strip()
                        task = str(emp.get("المهمة", "")).strip()
                        phone = str(emp.get("رقم التواصل", "")).strip()

                        st.markdown(f'''
                        <div class="audit-row">
                            <b>{emp_name}</b><br>
                            الرقم الشخصي: {eid}<br>
                            المدرسة: {school}<br>
                            المهمة: {task}<br>
                            رقم التواصل: {phone or 'غير موجود'}
                        </div>
                        ''', unsafe_allow_html=True)

                        msg_type = st.selectbox("نوع الرسالة", [
                            "تنبيه: تسجيل ورقي بدون إلكتروني",
                            "لم يظهر تسجيلك لهذا اليوم",
                            "تذكير بتسجيل الانصراف",
                            "رسالة مخصصة"
                        ], key="contact_msg_type")

                        default_msg = f"""السلام عليكم أ. {emp_name.split()[0] if emp_name else ''} 🌷
(الرقم الشخصي: {eid})

تبين لنا أنكِ سجلتِ في الورقة اليدوية ولم يتم تسجيل حضوركِ إلكترونيًا في النظام لهذا اليوم.

نأمل الالتزام بتسجيل الحضور والانصراف إلكترونيًا بشكل مباشر، لأن التسجيل الورقي فقط لا يُعتمد وقد لا يتم احتساب الحضور لهذا اليوم في حال عدم وجود تسجيل إلكتروني.

شكرًا لتعاونكِ.

🔗 {APP_URL}
"""
                        if msg_type == "لم يظهر تسجيلك لهذا اليوم":
                            default_msg = f"""السلام عليكم أ. {emp_name.split()[0] if emp_name else ''} 🌷
(الرقم الشخصي: {eid})

لم يظهر تسجيلكِ في نظام الحضور لهذا اليوم.
يرجى التأكد من التسجيل أو التواصل مع الأدمن في حال وجود مشكلة تقنية.

🔗 {APP_URL}
"""
                        elif msg_type == "تذكير بتسجيل الانصراف":
                            default_msg = f"""السلام عليكم أ. {emp_name.split()[0] if emp_name else ''} 🌷
(الرقم الشخصي: {eid})

يُرجى تسجيل الانصراف في نظام الحضور قبل مغادرة المركز.

🔗 {APP_URL}
"""
                        elif msg_type == "رسالة مخصصة":
                            default_msg = ""

                        final_msg = st.text_area("نص الرسالة", value=default_msg, height=180, key="contact_final_msg")
                        _wa_button_for_emp(phone, final_msg)

            else:
                with st.container(border=True):
                    st.markdown("##### 📝 السجلات المضافة يدويًا")
                    col_m1, col_m2 = st.columns(2)
                    with col_m1:
                        manual_from = st.date_input("من تاريخ", value=now_bh().date(), key="manual_added_from")
                    with col_m2:
                        manual_to = st.date_input("إلى تاريخ", value=now_bh().date(), key="manual_added_to")
                    mf = manual_from.strftime("%Y-%m-%d")
                    mt = manual_to.strftime("%Y-%m-%d")

                    data = get_sheet_data_fresh()
                    manual_rows = []
                    for r in data:
                        d = str(r.get("التاريخ", "")).strip().replace("/", "-")
                        if not (mf <= d <= mt):
                            continue
                        combined_txt = " ".join([
                            str(r.get("نوع التسجيل", "")),
                            str(r.get("محاولة", "")),
                            str(r.get("سبب التأخير", "")),
                            str(r.get("سبب الانصراف", "")),
                        ])
                        if any(x in combined_txt for x in ["يدوي", "طلب", "تعديل", "بدون تحقق GPS", "تحويل غياب"]):
                            manual_rows.append(r)

                    st.metric("عدد السجلات المضافة/المعدلة يدويًا", len(manual_rows))

                    if not manual_rows:
                        st.success("✅ لا توجد سجلات يدوية في هذا النطاق.")
                    else:
                        st.caption("يعرض الاسم والرقم الشخصي مع زر واتساب إذا كان رقم التواصل موجودًا في القائمة البيضاء.")
                        for idx, r in enumerate(reversed(manual_rows[-150:])):
                            eid = str(r.get("الرقم الشخصي", "")).strip()
                            emp_wl = wl_all.get(eid, {})
                            name = str(r.get("الاسم الثلاثي", "") or emp_wl.get("الاسم", "")).strip()
                            school = str(r.get("اسم المدرسة", "") or r.get("المدرسة", "") or emp_wl.get("المدرسة", "")).strip()
                            task = str(r.get("المهمة", "") or emp_wl.get("المهمة", "")).strip()
                            phone = str(emp_wl.get("رقم التواصل", "")).strip()
                            reg_type = str(r.get("نوع التسجيل", "") or r.get("محاولة", "") or "سجل يدوي/معدل").strip()
                            d = str(r.get("التاريخ", "")).strip()

                            with st.expander(f"📝 {name} — #{eid} — {d}", expanded=False):
                                st.markdown(f'''
                                <div class="audit-row">
                                    <b>{name}</b><br>
                                    الرقم الشخصي: {eid}<br>
                                    التاريخ: {d}<br>
                                    المدرسة: {school}<br>
                                    المهمة: {task}<br>
                                    حضور: {r.get('وقت الحضور','') or '—'} | انصراف: {r.get('وقت الانصراف','') or '—'}<br>
                                    نوع/ملاحظة السجل: {reg_type}<br>
                                    رقم التواصل: {phone or 'غير موجود'}
                                </div>
                                ''', unsafe_allow_html=True)

                                msg = f"""السلام عليكم أ. {name.split()[0] if name else ''} 🌷
(الرقم الشخصي: {eid})

تبين لنا أنكِ سجلتِ في الورقة اليدوية ولم يتم تسجيل حضوركِ إلكترونيًا في النظام بتاريخ {d}.

نأمل الالتزام بتسجيل الحضور والانصراف إلكترونيًا بشكل مباشر، لأن التسجيل الورقي فقط لا يُعتمد وقد لا يتم احتساب الحضور لهذا اليوم في حال عدم وجود تسجيل إلكتروني.

شكرًا لتعاونكِ.

🔗 {APP_URL}
"""
                                msg_edit = st.text_area("نص رسالة واتساب", value=msg, height=150, key=f"manual_msg_{idx}_{eid}_{d}")
                                _wa_button_for_emp(phone, msg_edit)


        elif admin_tab=="🛠️ إصلاح شامل":
            st.markdown("#### 🛠️ إصلاح شامل")
            st.warning("⚠️ هذه الأدوات تعدّل البيانات مباشرة. استخدميها بحذر.")

            # ── إعادة حساب الساعات ──
            with st.container(border=True):
                st.markdown("##### 🔄 إعادة حساب ساعات العمل")
                st.caption("يعيد احتساب ساعات العمل والساعات الإضافية وحالة الدوام لجميع السجلات أو لتاريخ محدد.")
                recalc_mode = st.radio("النطاق", ["تاريخ محدد", "كل السجلات"], horizontal=True, key="recalc_mode")
                if recalc_mode == "تاريخ محدد":
                    recalc_date = st.date_input("التاريخ", value=now_bh().date(), key="recalc_date")
                    recalc_date_str = recalc_date.strftime("%Y-%m-%d")
                else:
                    recalc_date_str = None

                if st.button("🔄 بدء إعادة الحساب", use_container_width=True, type="primary", key="btn_recalc"):
                    try:
                        data = get_sheet_data()
                        updated = 0
                        errors  = 0
                        progress = st.progress(0)
                        total_r = len(data)
                        for i, row in enumerate(data):
                            progress.progress((i+1)/max(total_r,1))
                            if recalc_date_str and str(row.get("التاريخ","")).strip() != recalc_date_str:
                                continue
                            if not row.get("وقت الحضور",""):
                                continue
                            try:
                                if update_work_calculation(i+2, row):
                                    updated += 1
                            except Exception:
                                errors += 1
                        clear_caches()
                        st.success(f"✅ تم إعادة حساب {updated} سجل. أخطاء: {errors}")
                        log_audit("—","أدمن","إعادة حساب شاملة",f"نطاق:{recalc_date_str or 'الكل'}|محدّث:{updated}|أخطاء:{errors}")
                    except Exception as e:
                        st.error(f"❌ خطأ: {e}")

            st.markdown("---")

            # ── تنظيف التكرارات لكل التواريخ ──
            with st.container(border=True):
                st.markdown("##### 🧹 تنظيف تكرارات كل التواريخ")
                st.caption("يفحص جميع التواريخ ويحذف السجلات المكررة تلقائياً — يحتفظ بالأكمل.")
                if st.button("🔍 فحص وتنظيف كل التكرارات", use_container_width=True, type="primary", key="btn_clean_all_dups"):
                    try:
                        groups = find_duplicate_attendance_groups()
                        if not groups:
                            st.success("✅ لا توجد تكرارات في أي تاريخ.")
                        else:
                            total_deleted = 0
                            for (date_val, emp_id_val), _ in groups.items():
                                deleted = auto_cleanup_duplicate_attendance_for_emp(date_val, emp_id_val, "إصلاح شامل")
                                total_deleted += deleted
                            clear_caches()
                            st.success(f"✅ تم حذف {total_deleted} سجل مكرر.")
                            log_audit("—","أدمن","تنظيف شامل للتكرارات",f"سجلات محذوفة:{total_deleted}")
                    except Exception as e:
                        st.error(f"❌ خطأ: {e}")

            st.markdown("---")

            # ── إغلاق السجلات المفتوحة ──
            with st.container(border=True):
                st.markdown("##### 🔒 إغلاق السجلات المفتوحة من أيام سابقة")
                st.caption("يغلق تلقائياً أي سجل من أيام سابقة لم يُسجَّل له انصراف.")
                if st.button("🔒 تنفيذ الإغلاق التلقائي الآن", use_container_width=True, key="btn_force_autoclose"):
                    try:
                        auto_close_previous_open_records()
                        clear_caches()
                        st.success("✅ تم تنفيذ الإغلاق التلقائي لجميع السجلات المفتوحة.")
                        log_audit("—","أدمن","إغلاق تلقائي يدوي","تنفيذ من لوحة الإصلاح الشامل")
                    except Exception as e:
                        st.error(f"❌ خطأ: {e}")
        # ── إعدادات التسجيل اليدوي ───────────────────────────────
        elif admin_tab=="⚙️ إعدادات التسجيل اليدوي":
            st.markdown("#### ⚙️ إعدادات التسجيل اليدوي للموظفات")

            # ── وقت الإغلاق التلقائي ──
            with st.container(border=True):
                st.markdown("##### 🕙 وقت الإغلاق التلقائي لسجلات اليوم")
                st.caption("عند فتح البرنامج بعد هذا الوقت، تُغلق سجلات اليوم المفتوحة تلقائياً.")
                cur_close_time = get_system_setting("auto_close_time", "22:00")
                try:
                    _ch, _cm = map(int, cur_close_time.split(":"))
                except:
                    _ch, _cm = 22, 0
                new_close_h = st.number_input("الساعة", 18, 23, _ch, key="close_hour")
                new_close_m = st.number_input("الدقيقة", 0, 59, _cm, step=15, key="close_min")
                new_close_str = f"{int(new_close_h):02d}:{int(new_close_m):02d}"
                st.info(f"الوقت الحالي المحدد: **{cur_close_time}** — الجديد: **{new_close_str}**")
                if st.button("💾 حفظ وقت الإغلاق", use_container_width=True, key="btn_save_close_time"):
                    if set_system_setting("auto_close_time", new_close_str, f"تغيير وقت الإغلاق التلقائي إلى {new_close_str}"):
                        log_audit("—","أدمن","تغيير وقت الإغلاق التلقائي",f"الجديد: {new_close_str}")
                        clear_caches()
                        st.success(f"✅ تم حفظ وقت الإغلاق: {new_close_str}")
                        st.rerun()

            st.markdown("---")
            current_enabled = manual_requests_enabled()
            if current_enabled:
                st.success("✅ طلبات التسجيل اليدوي من واجهة الموظفة مفعّلة حاليًا.")
            else:
                st.warning("⚠️ طلبات التسجيل اليدوي من واجهة الموظفة معطّلة حاليًا. الموظفات سيظهر لهن تنبيه للتواصل مع الأدمن فقط.")

            new_enabled = st.radio(
                "حالة الخاصية",
                ["مفعّلة", "معطّلة"],
                index=0 if current_enabled else 1,
                horizontal=True,
                key="manual_requests_enabled_radio"
            )
            setting_note = st.text_input("ملاحظة/سبب التغيير", value="تغيير إعداد طلبات التسجيل اليدوي", key="manual_requests_setting_note")

            if st.button("💾 حفظ إعداد التسجيل اليدوي", use_container_width=True, type="primary"):
                value = "true" if new_enabled == "مفعّلة" else "false"
                if set_system_setting("manual_requests_enabled", value, setting_note):
                    log_audit("—", "أدمن", "تعديل إعداد التسجيل اليدوي", f"manual_requests_enabled={value}|{setting_note}")
                    clear_caches()
                    st.success("✅ تم حفظ الإعداد بنجاح.")
                    st.rerun()

            st.info("حتى عند تعطيل هذه الخاصية، يبقى قسم ➕ تسجيل يدوي في لوحة الأدمن شغالًا، وتستطيعين تسجيل الموظفة يدويًا بعد التواصل معها.")

        # ── طلبات التسجيل اليدوي ─────────────────────────────────
        elif admin_tab=="🆘 طلبات التسجيل اليدوي":
            st.markdown("#### 🆘 طلبات التسجيل اليدوي / مشاكل الموقع والمتصفح")

            # تهيئة قائمة الطلبات المعتمدة/المرفوضة محلياً لتجنب اختفائها بعد rerun
            if "approved_req_rows" not in st.session_state:
                st.session_state.approved_req_rows = set()

            reqs = get_manual_requests()
            pending = [(i+2, r) for i, r in enumerate(reqs)
                       if str(r.get("الحالة", "")).strip() in ["", "بانتظار الاعتماد"]
                       and (i+2) not in st.session_state.approved_req_rows]
            done = [(i+2, r) for i, r in enumerate(reqs)
                    if str(r.get("الحالة", "")).strip() not in ["", "بانتظار الاعتماد"]
                    or (i+2) in st.session_state.approved_req_rows]

            st.metric("طلبات بانتظار الاعتماد", len(pending))

            # زر تحديث يدوي
            if st.button("🔄 تحديث القائمة", key="btn_refresh_reqs"):
                st.session_state.approved_req_rows = set()
                get_manual_requests.clear()
                st.rerun()

            if not pending:
                st.success("✅ لا توجد طلبات معلقة حالياً.")
            else:
                for row_num, r in reversed(pending[-80:]):
                    eid = str(r.get("الرقم الشخصي", "")).strip()
                    title = f"{r.get('الاسم','')} — #{eid} — {r.get('نوع الطلب','')} — {r.get('نوع المشكلة','')}"
                    with st.expander(title, expanded=False):
                        st.markdown(f'<div class="warn-row"><b>تاريخ الطلب:</b> {r.get("تاريخ الطلب","")} — <b>وقت الإرسال المعتمد:</b> {r.get("وقت الطلب","")}<br><b>نوع الطلب:</b> {r.get("نوع الطلب","")}<br><b>المدرسة:</b> {r.get("المدرسة","")}<br><b>المهمة:</b> {r.get("المهمة","")}<br><b>الملاحظات:</b> {r.get("ملاحظات","") or "—"}</div>', unsafe_allow_html=True)
                        c1, c2, c3 = st.columns(3)
                        with c1:
                            if st.button("✅ اعتماد حضور", key=f"approve_att_{row_num}", use_container_width=True, type="primary"):
                                if approve_manual_request(row_num, r, "حضور", False):
                                    st.session_state.approved_req_rows.add(row_num)
                                    st.success("✅ تم اعتماد الحضور اليدوي")
                                    st.rerun()
                        with c2:
                            if st.button("🔵 اعتماد انصراف", key=f"approve_dep_{row_num}", use_container_width=True):
                                if approve_manual_request(row_num, r, "انصراف", False):
                                    st.session_state.approved_req_rows.add(row_num)
                                    st.success("✅ تم اعتماد الانصراف اليدوي")
                                    st.rerun()
                        with c3:
                            if st.button("❌ رفض الطلب", key=f"reject_req_{row_num}", use_container_width=True):
                                try:
                                    manual_requests_sheet.update_cell(row_num, 12, "مرفوض")
                                    manual_requests_sheet.update_cell(row_num, 14, now_bh().strftime("%Y-%m-%d %H:%M:%S"))
                                    manual_requests_sheet.update_cell(row_num, 15, "أدمن")
                                    log_audit(eid, r.get("الاسم", ""), "رفض طلب تسجيل يدوي", f"نوع المشكلة:{r.get('نوع المشكلة','')}|وقت الطلب:{r.get('وقت الطلب','')}")
                                    st.session_state.approved_req_rows.add(row_num)
                                    clear_caches()
                                    st.success("تم رفض الطلب")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"تعذر رفض الطلب: {e}")
            with st.expander("الطلبات المعتمدة/المرفوضة الأخيرة", expanded=False):
                for row_num, r in reversed(done[-40:]):
                    st.markdown(f'<div class="audit-row">{r.get("الحالة","")} — {r.get("الاسم","")} — #{r.get("الرقم الشخصي","")} — {r.get("تاريخ الطلب","")} {r.get("وقت الطلب","")}</div>', unsafe_allow_html=True)

        # ── تسجيل الغياب ────────────────────────────────────────
        elif admin_tab=="🔴 تسجيل الغياب":
            abs_date=st.date_input("تاريخ الغياب",value=now_bh().date(),key="abs_date")
            abs_date_str=str(abs_date)
            abs_day_ar=day_ar_from_date(abs_date)

            wl_all=get_whitelist()
            required_from_people = required_people_for_date(abs_date_str)
            scheduled_tasks = None
            schedule_source = ""

            if required_from_people is not None:
                st.success("✅ تسجيل الغياب يعتمد على ورقة مطلوبات_اليوم لهذا التاريخ.")
                with st.expander(f"📋 مطلوبات اليوم للغياب — {abs_date_str}", expanded=False):
                    last_task = None
                    last_school = None
                    for eid, emp in required_from_people.items():
                        task = str(emp.get("المهمة", "")).strip() or "غير محدد"
                        school = str(emp.get("المدرسة", "")).strip() or "غير محدد"
                        if task != last_task:
                            st.markdown(f"### 📌 {task}")
                            last_task = task
                            last_school = None
                        if school != last_school:
                            st.markdown(f"**🏫 {school}**")
                            last_school = school
                        st.markdown(f"- {emp.get('الاسم','')} — #{eid}")
            else:
                scheduled_tasks, schedule_source = scheduled_tasks_for_date(abs_date_str)
                if scheduled_tasks is None:
                    st.warning("⚠️ لا توجد قائمة مطلوبات اليوم ولم يتم تحديد دوام أقسام لهذا اليوم، لذلك سيتم حصر الغياب على جميع القائمة البيضاء.")
                else:
                    with st.expander(f"📅 الأقسام المطلوب دوامها يوم {abs_day_ar} — مصدر الجدول: {schedule_source}", expanded=False):
                        for t in scheduled_tasks:
                            st.markdown(f"- {t}")

            if not wl_all:
                st.warning("⚠️ القائمة البيضاء فارغة")
            else:
                # الأولوية لقائمة مطلوبات اليوم، وإذا غير موجودة يرجع لجدول الدوام
                if required_from_people is not None:
                    required_wl = required_from_people
                else:
                    required_wl={eid:emp for eid,emp in wl_all.items() if emp_required_on_day(emp, scheduled_tasks)}

                data=get_sheet_data()
                attended_ids=set(str(r.get("الرقم الشخصي","")).strip() for r in data if r.get("التاريخ")==abs_date_str and r.get("وقت الحضور"))
                try:
                    abs_records=absence_sheet.get_all_records()
                    absent_ids=set(str(r.get("الرقم الشخصي","")).strip() for r in abs_records if r.get("التاريخ")==abs_date_str)
                except:
                    abs_records=[]; absent_ids=set()

                not_registered={eid:emp for eid,emp in required_wl.items() if eid not in attended_ids and eid not in absent_ids}
                already_absent={eid:emp for eid,emp in required_wl.items() if eid in absent_ids}

                c1,c2,c3=st.columns(3)
                c1.metric("المطلوب دوامهم",len(required_wl))
                c2.metric("حاضرات",len(attended_ids.intersection(set(required_wl.keys()))))
                c3.metric("لم يسجّلن",len(not_registered))

                if required_from_people is not None:
                    st.info("✅ الحصر الحالي يعتمد فقط على المعلمات الموجودات في ورقة مطلوبات_اليوم لهذا التاريخ.")
                elif scheduled_tasks is not None:
                    st.info("✅ الحصر الحالي يعتمد فقط على الأقسام/المهام المحددة في جدول دوام الأقسام لهذا اليوم.")

                if already_absent:
                    st.markdown("#### تم تسجيل غيابهن / تعديل الحالة")
                    for eid,emp in already_absent.items():
                        rec=next((r for r in abs_records if str(r.get("الرقم الشخصي",""))==eid and r.get("التاريخ")==abs_date_str),{})
                        st.markdown(f'<div class="absent-row">🔴 {emp.get("الاسم","")} — {emp.get("المهمة","")} — سبب: {rec.get("سبب الغياب","")}</div>',unsafe_allow_html=True)
                        with st.expander(f"✏️ تعديل غياب / تحويل إلى حضور يدوي — {emp.get('الاسم','')}", expanded=False):
                            new_abs_reason = st.selectbox("تعديل سبب الغياب", abs_reasons, index=abs_reasons.index(rec.get("سبب الغياب","مرض")) if rec.get("سبب الغياب","") in abs_reasons else 0, key=f"edit_abs_reason_{eid}")
                            new_abs_note = st.text_input("ملاحظات الغياب", value=rec.get("ملاحظات", ""), key=f"edit_abs_note_{eid}")
                            c_abs1, c_abs2 = st.columns(2)
                            with c_abs1:
                                if st.button("💾 حفظ تعديل الغياب", key=f"save_abs_edit_{eid}", use_container_width=True):
                                    try:
                                        for i, rr in enumerate(abs_records):
                                            if str(rr.get("الرقم الشخصي", "")).strip() == eid and rr.get("التاريخ") == abs_date_str:
                                                absence_sheet.update_cell(i+2, 7, new_abs_reason)
                                                absence_sheet.update_cell(i+2, 8, new_abs_note)
                                                log_audit(eid, emp.get("الاسم", ""), "تعديل غياب", f"التاريخ:{abs_date_str}|السبب:{new_abs_reason}|ملاحظات:{new_abs_note}")
                                                st.success("✅ تم تعديل الغياب")
                                                st.rerun()
                                    except Exception as e:
                                        st.error(f"❌ تعذر تعديل الغياب: {e}")
                            with c_abs2:
                                st.caption("إذا كانت الموظفة موجودة لكن المتصفح لا يعمل، سجليها يدويًا وسيتم إلغاء الغياب.")

                            manual_att = st.text_input("وقت الحضور اليدوي", value="07:00:00", key=f"abs_manual_att_{eid}")
                            manual_dep = st.text_input("وقت الانصراف اليدوي (اختياري)", key=f"abs_manual_dep_{eid}")
                            manual_note = st.text_input("سبب التسجيل اليدوي", value="مشكلة في المتصفح / البرنامج لم يفتح", key=f"abs_manual_note_{eid}")
                            if st.button("✅ تحويل إلى تسجيل حضور يدوي وإلغاء الغياب", key=f"convert_abs_manual_{eid}", use_container_width=True, type="primary"):
                                if not manual_note.strip():
                                    st.error("❌ سبب التسجيل اليدوي مطلوب")
                                else:
                                    try:
                                        existing_matches = find_daily_rows_fresh(abs_date_str, eid)
                                        existing_idx, existing_row = pick_main_daily_row(existing_matches)
                                        full_name = normalize_name(emp.get("الاسم", ""))
                                        support_raw = str(emp.get("دعم", "")).strip()
                                        task = str(emp.get("المهمة", "")).strip()
                                        support_value = "نعم" if support_raw in ["نعم","yes","Yes","TRUE","true","1"] or "دعم" in task else "لا"
                                        if existing_idx:
                                            safe_update(sheet, existing_idx, COL_ATTEND, manual_att)
                                            safe_update(sheet, existing_idx, COL_LATE_REASON, f"[يدوي] {manual_note.strip()}")
                                            safe_update(sheet, existing_idx, COL_DEPART, manual_dep)
                                            existing_row["وقت الحضور"] = manual_att
                                            existing_row["وقت الانصراف"] = manual_dep
                                            update_work_calculation(existing_idx, existing_row)
                                        else:
                                            safe_append(sheet, [abs_date_str, abs_day_ar, emp.get("المدرسة", ""), task, support_value, full_name, eid, manual_att, f"[يدوي] {manual_note.strip()}", manual_dep, "", "", "", "", "", "", "", "", "", "", "", "", "تحويل غياب إلى تسجيل يدوي"])
                                            idx_after, row_after = find_today_row_fresh(abs_date_str, eid)
                                            if idx_after:
                                                update_work_calculation(idx_after, row_after)
                                        # حذف سجل الغياب لنفس التاريخ والرقم
                                        fresh_abs = absence_sheet.get_all_records()
                                        for i in range(len(fresh_abs)-1, -1, -1):
                                            rr = fresh_abs[i]
                                            if str(rr.get("الرقم الشخصي", "")).strip() == eid and rr.get("التاريخ") == abs_date_str:
                                                absence_sheet.delete_rows(i+2)
                                        log_audit(eid, full_name, "تحويل غياب إلى حضور يدوي", f"التاريخ:{abs_date_str}|حضور:{manual_att}|انصراف:{manual_dep}|السبب:{manual_note.strip()}")
                                        clear_caches()
                                        st.success("✅ تم تسجيل الحضور اليدوي وإلغاء الغياب")
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"❌ تعذر التحويل: {e}")

                if not_registered:
                    st.markdown("#### لم يسجّلن بعد")
                    for eid,emp in not_registered.items():
                        with st.expander(f"🔴 {emp.get('الاسم','')} — {emp.get('المدرسة','')} — {emp.get('المهمة','')}"):
                            sel=st.selectbox("سبب الغياب",abs_reasons,key=f"ar_{eid}")
                            other_txt=""
                            if sel=="أخرى": other_txt=st.text_input("اكتبي السبب",key=f"ao_{eid}")
                            note_txt=st.text_input("ملاحظات (اختياري)",key=f"an_{eid}")
                            final_r=other_txt.strip() if sel=="أخرى" else sel
                            if st.button("تسجيل غياب",key=f"ab_{eid}",use_container_width=True):
                                absence_sheet.append_row([abs_date_str,abs_day_ar,eid,emp.get("الاسم",""),emp.get("المدرسة",""),emp.get("المهمة",""),final_r,note_txt,"أدمن"])
                                log_audit(eid,emp.get("الاسم",""),"تسجيل غياب",f"التاريخ:{abs_date_str}|السبب:{final_r}")
                                st.success(f"✅ تم تسجيل غياب {emp.get('الاسم','')}"); st.rerun()
                else:
                    st.success("✅ تم تسجيل وضع جميع الموظفات المطلوب دوامهن لهذا اليوم")

        # ── مطلوبات اليوم ────────────────────────────────────────
        elif admin_tab=="📋 مطلوبات اليوم":
            st.markdown("#### 📋 تحديد المعلمات المطلوبات لليوم")
            st.info("إذا تم إدخال قائمة لهذا التاريخ، ستعتمد إحصائيات اليوم وتسجيل الغياب على هذه القائمة فقط، وبالترتيب: المهمة ← المدرسة ← المعلمات.")

            req_date = st.date_input("تاريخ القائمة", value=now_bh().date(), key="req_today_date")
            req_date_str = req_date.strftime("%Y-%m-%d")
            wl_all = get_whitelist()

            with st.container(border=True):
                st.markdown("##### ➕ إضافة من القائمة البيضاء")
                col_rt1, col_rt2 = st.columns(2)
                with col_rt1:
                    rt_task_filter = st.selectbox("المهمة", ["الكل"] + sorted(set(str(e.get("المهمة", "")).strip() for e in wl_all.values() if str(e.get("المهمة", "")).strip())), key="rt_task_filter")
                with col_rt2:
                    whitelist_schools = sorted(set(get_emp_school(e) for e in wl_all.values() if get_emp_school(e)))
                    rt_school_filter = st.selectbox("المدرسة", ["الكل"] + whitelist_schools, key="rt_school_filter")

                rt_options = []
                rt_task_filter_norm = normalize_name(rt_task_filter)
                rt_school_filter_norm = normalize_school_name(rt_school_filter)
                for eid, emp in wl_all.items():
                    task = str(emp.get("المهمة", "") or "").strip()
                    school = get_emp_school(emp)
                    name = get_emp_name(emp)
                    if rt_task_filter != "الكل" and normalize_name(task) != rt_task_filter_norm:
                        continue
                    if rt_school_filter != "الكل" and normalize_school_name(school) != rt_school_filter_norm:
                        continue
                    if eid and name:
                        rt_options.append((eid, emp, f"{task} — {school} — {name} — #{eid}"))
                rt_options = sorted(rt_options, key=lambda x: x[2])
                rt_label_map = {lbl: (eid, emp) for eid, emp, lbl in rt_options}

                st.caption("اختاري من الفلاتر، أو ابحثي بالاسم/الرقم من كل القائمة البيضاء حتى لو كانت مسجلة تحت مدرسة أخرى.")
                selected_labels = st.multiselect("اختاري المعلمات المطلوبات حسب الفلتر", list(rt_label_map.keys()), key="rt_selected_people")

                rt_search_term = st.text_input("بحث عام بالاسم أو الرقم الشخصي — يتجاوز فلتر المدرسة والمهمة", key="rt_global_search")
                rt_search_matches = find_whitelist_matches(rt_search_term, limit=80) if rt_search_term else []
                rt_search_map = {lbl: (eid, emp) for eid, emp, lbl in rt_search_matches}
                selected_search_labels = []
                if rt_search_term:
                    if rt_search_matches:
                        selected_search_labels = st.multiselect("نتائج البحث العام", list(rt_search_map.keys()), key="rt_selected_search_people")
                    else:
                        st.warning("⚠️ لم يتم العثور على اسم/رقم مطابق في القائمة البيضاء.")

                rt_note = st.text_input("ملاحظات اختيارية", key="rt_note")

                if st.button("💾 إضافة المختارات لقائمة اليوم", use_container_width=True, type="primary", key="btn_add_required_today"):
                    if not selected_labels and not selected_search_labels:
                        st.error("❌ اختاري معلمة واحدة على الأقل من الفلتر أو من البحث العام.")
                    else:
                        try:
                            existing = get_required_today_records()
                            existing_keys = set(
                                (str(r.get("التاريخ", "")).strip().replace("/", "-"), str(r.get("الرقم الشخصي", "")).strip())
                                for r in existing
                            )
                            added = 0
                            selected_all_labels = list(dict.fromkeys((selected_labels or []) + (selected_search_labels or [])))
                            combined_map = {}
                            combined_map.update(rt_label_map)
                            combined_map.update(rt_search_map)

                            for lbl in selected_all_labels:
                                eid, emp = combined_map[lbl]
                                if (req_date_str, str(eid).strip()) in existing_keys:
                                    continue
                                required_today_sheet.append_row([
                                    req_date_str,
                                    str(emp.get("المهمة", "")).strip(),
                                    get_emp_school(emp),
                                    str(eid).strip(),
                                    get_emp_name(emp),
                                    "نعم",
                                    rt_note,
                                ], value_input_option="USER_ENTERED")
                                added += 1
                            get_required_today_records.clear()
                            clear_caches()
                            st.success(f"✅ تم إضافة {added} معلمة لقائمة {req_date_str}.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ تعذر الحفظ: {e}")

            with st.container(border=True):
                st.markdown("##### 📥 رفع قائمة من Excel")
                st.caption("الأعمدة المقبولة: التاريخ، المهمة، المدرسة، الرقم الشخصي، الاسم، نشط، ملاحظات. أهم عمود هو الرقم الشخصي. إذا التاريخ فارغ يستخدم تاريخ القائمة المختار.")
                up_req = st.file_uploader("ارفعي ملف Excel لقائمة المطلوبات", type=["xlsx"], key="upload_required_today")
                if up_req is not None:
                    try:
                        df_req = pd.read_excel(up_req)
                        df_req.columns = [str(c).strip() for c in df_req.columns]
                        st.dataframe(df_req.head(30), use_container_width=True)
                        if st.button("✅ اعتماد ورفع القائمة", use_container_width=True, type="primary", key="btn_import_required_today"):
                            id_col = next((c for c in df_req.columns if c in ["الرقم الشخصي", "رقم شخصي", "CPR", "cpr", "ID", "id"]), None)
                            if not id_col:
                                st.error("❌ لم أجد عمود الرقم الشخصي في الملف.")
                            else:
                                added = 0
                                existing = get_required_today_records()
                                existing_keys = set(
                                    (str(r.get("التاريخ", "")).strip().replace("/", "-"), str(r.get("الرقم الشخصي", "")).strip())
                                    for r in existing
                                )
                                for _, row in df_req.iterrows():
                                    eid = ar_to_en_digits(str(row.get(id_col, "")).strip()).replace(".0", "")
                                    if not eid or eid.lower() == "nan":
                                        continue
                                    emp = wl_all.get(eid, {})
                                    d_val = str(row.get("التاريخ", req_date_str) or req_date_str).strip()
                                    if d_val.lower() == "nan" or not d_val:
                                        d_val = req_date_str
                                    try:
                                        d_val = pd.to_datetime(d_val).strftime("%Y-%m-%d")
                                    except Exception:
                                        d_val = d_val.replace("/", "-")
                                    if (d_val, eid) in existing_keys:
                                        continue
                                    task_val = str(row.get("المهمة", "") or emp.get("المهمة", "")).strip()
                                    school_val = str(row.get("المدرسة", row.get("اسم المدرسة", "")) or get_emp_school(emp)).strip()
                                    name_val = str(row.get("الاسم", row.get("الاسم الثلاثي", "")) or get_emp_name(emp)).strip()
                                    active_val = str(row.get("نشط", "نعم") or "نعم").strip()
                                    note_val = str(row.get("ملاحظات", "") or "").strip()
                                    if active_val.lower() == "nan": active_val = "نعم"
                                    if note_val.lower() == "nan": note_val = ""
                                    required_today_sheet.append_row([d_val, task_val, school_val, eid, name_val, active_val, note_val], value_input_option="USER_ENTERED")
                                    added += 1
                                get_required_today_records.clear()
                                clear_caches()
                                st.success(f"✅ تم رفع {added} سجل في مطلوبات اليوم.")
                                st.rerun()
                    except Exception as e:
                        st.error(f"❌ تعذر قراءة الملف: {e}")

            existing_req = required_people_for_date(req_date_str)
            st.markdown("##### 📌 القائمة الحالية")
            if not existing_req:
                st.warning("لا توجد مطلوبات مسجلة لهذا التاريخ. ستعمل الإحصائيات حسب جدول الأقسام.")
            else:
                st.metric("عدد المطلوبات", len(existing_req))
                rows_show = []
                for eid, emp in existing_req.items():
                    rows_show.append({
                        "المهمة": emp.get("المهمة", ""),
                        "المدرسة": emp.get("المدرسة", ""),
                        "الاسم": emp.get("الاسم", ""),
                        "الرقم الشخصي": eid,
                    })
                st.dataframe(pd.DataFrame(rows_show), use_container_width=True, hide_index=True)

                with st.expander("🗑️ حذف/تعطيل من قائمة هذا التاريخ", expanded=False):
                    labels_del = [f"{r['المهمة']} — {r['المدرسة']} — {r['الاسم']} — #{r['الرقم الشخصي']}" for r in rows_show]
                    selected_del = st.multiselect("اختاري من تريدين تعطيله من القائمة", labels_del, key="rt_delete_select")
                    if st.button("🚫 تعطيل المختارات", use_container_width=True, key="btn_disable_required_today"):
                        try:
                            records = required_today_sheet.get_all_records()
                            ids_to_disable = set(lbl.split("#")[-1].strip() for lbl in selected_del)
                            changed = 0
                            for i, r in enumerate(records, start=2):
                                if str(r.get("التاريخ", "")).strip().replace("/", "-") == req_date_str and str(r.get("الرقم الشخصي", "")).strip() in ids_to_disable:
                                    required_today_sheet.update_cell(i, 6, "لا")
                                    changed += 1
                            get_required_today_records.clear()
                            clear_caches()
                            st.success(f"✅ تم تعطيل {changed} سجل.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ تعذر التعطيل: {e}")

                    if st.button("🧹 تعطيل كل قائمة هذا التاريخ", use_container_width=True, key="btn_disable_all_required_today"):
                        try:
                            records = required_today_sheet.get_all_records()
                            changed = 0
                            for i, r in enumerate(records, start=2):
                                if str(r.get("التاريخ", "")).strip().replace("/", "-") == req_date_str and is_yes(r.get("نشط", "نعم")):
                                    required_today_sheet.update_cell(i, 6, "لا")
                                    changed += 1
                            get_required_today_records.clear()
                            clear_caches()
                            st.success(f"✅ تم تعطيل كل قائمة هذا التاريخ: {changed} سجل.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ تعذر التعطيل: {e}")

        # ── دوام الأقسام ────────────────────────────────────────
        elif admin_tab=="📅 دوام الأقسام":
            st.markdown("#### 📅 تحديد الأقسام/المهام التي تداوم في كل يوم")
            st.info("حددي الأيام المطلوبة لكل مهمة. عند حصر الغياب سيحسب البرنامج فقط الموظفات التابعات للمهام المفعّلة في يوم الغياب.")

            st.markdown("### 🗓️ دوام يوم محدد — مرن")
            st.caption("استخدمي هذا إذا تغير الجدول في يوم معيّن. هذا الجدول له أولوية على الجدول الأسبوعي لذلك التاريخ.")
            daily_date = st.date_input("تاريخ الدوام الخاص", value=now_bh().date(), key="daily_sch_date")
            daily_date_str = str(daily_date)
            daily_tasks_selected = st.multiselect("اختاري الأقسام/المهام التي ستداوم في هذا التاريخ", TASKS_ALL, key="daily_sch_tasks")
            daily_note = st.text_input("ملاحظات اختيارية", key="daily_sch_note")
            if st.button("💾 حفظ دوام هذا التاريخ", use_container_width=True, type="primary", key="save_daily_schedule"):
                if not daily_tasks_selected:
                    st.error("❌ اختاري مهمة واحدة على الأقل.")
                else:
                    try:
                        records = daily_schedule_sheet.get_all_records()
                        # نجمع كل التحديثات في batch واحد
                        batch_updates = []
                        for i, r in enumerate(records):
                            if str(r.get("التاريخ","")).strip() == daily_date_str:
                                batch_updates.append({
                                    "range": f"C{i+2}",
                                    "values": [["لا"]]
                                })
                        if batch_updates:
                            daily_schedule_sheet.batch_update(batch_updates)
                        import time as _t; _t.sleep(1)
                        for t in daily_tasks_selected:
                            daily_schedule_sheet.append_row([daily_date_str, t, "نعم", daily_note], value_input_option="USER_ENTERED")
                            _t.sleep(0.3)
                        get_daily_schedule_records.clear()
                        st.success("✅ تم حفظ دوام التاريخ المحدد.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ خطأ أثناء حفظ دوام اليوم: {e}")

            current_daily = [r for r in get_daily_schedule_records() if str(r.get("التاريخ","")).strip()==daily_date_str and (str(r.get("نشط","")).strip()=="" or is_yes(r.get("نشط","")))]
            if current_daily:
                with st.expander(f"📌 دوام محفوظ لتاريخ {daily_date_str}", expanded=True):
                    for i, r in enumerate(current_daily, start=1):
                        st.markdown(f'<div class="audit-row"><b>{r.get("المهمة","")}</b><br><small>{r.get("ملاحظات","")}</small></div>', unsafe_allow_html=True)
                    if st.button("🗑️ تعطيل دوام هذا التاريخ", use_container_width=True, key="disable_daily_schedule"):
                        try:
                            records = daily_schedule_sheet.get_all_records()
                            batch_updates = []
                            for i, r in enumerate(records):
                                if str(r.get("التاريخ","")).strip() == daily_date_str:
                                    batch_updates.append({"range": f"C{i+2}", "values": [["لا"]]})
                            if batch_updates:
                                daily_schedule_sheet.batch_update(batch_updates)
                            get_daily_schedule_records.clear()
                            st.success("✅ تم تعطيل دوام هذا التاريخ.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ تعذر التعطيل: {e}")

            st.markdown("---")
            st.markdown("### 📅 الجدول الأسبوعي الثابت")

            with st.expander("➕ إضافة / تحديث مهمة", expanded=True):
                sch_task=st.selectbox("المهمة", TASKS_ALL, key="sch_task")
                days_cols=st.columns(4)
                with days_cols[0]: d_sat=st.checkbox("السبت", key="d_sat")
                with days_cols[1]: d_sun=st.checkbox("الأحد", key="d_sun")
                with days_cols[2]: d_mon=st.checkbox("الاثنين", key="d_mon")
                with days_cols[3]: d_tue=st.checkbox("الثلاثاء", key="d_tue")
                days_cols2=st.columns(3)
                with days_cols2[0]: d_wed=st.checkbox("الأربعاء", key="d_wed")
                with days_cols2[1]: d_thu=st.checkbox("الخميس", key="d_thu")
                with days_cols2[2]: d_fri=st.checkbox("الجمعة", key="d_fri")

                active_task=st.checkbox("نشط", value=True, key="sch_active")
                if st.button("💾 حفظ دوام المهمة", use_container_width=True, type="primary"):
                    row_values=[
                        sch_task,
                        "نعم" if d_sat else "لا",
                        "نعم" if d_sun else "لا",
                        "نعم" if d_mon else "لا",
                        "نعم" if d_tue else "لا",
                        "نعم" if d_wed else "لا",
                        "نعم" if d_thu else "لا",
                        "نعم" if d_fri else "لا",
                        "نعم" if active_task else "لا",
                    ]
                    try:
                        records=schedule_sheet.get_all_records()
                        found_row=None
                        for i,r in enumerate(records):
                            if str(r.get("المهمة","")).strip()==sch_task:
                                found_row=i+2; break
                        if found_row:
                            schedule_sheet.update(f"A{found_row}:I{found_row}",[row_values])
                        else:
                            schedule_sheet.append_row(row_values, value_input_option="USER_ENTERED")
                        get_schedule_records.clear()
                        st.success("✅ تم حفظ جدول دوام المهمة")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ خطأ أثناء حفظ جدول الدوام: {e}")

            st.markdown("#### الجدول الحالي")
            try:
                records=get_schedule_records()
                if not records:
                    st.warning("لم يتم إدخال أي دوام أقسام حتى الآن.")
                else:
                    for r in records:
                        task=str(r.get("المهمة","")).strip()
                        active="نشط" if is_yes(r.get("نشط","")) or str(r.get("نشط","")).strip()=="" else "غير نشط"
                        days=[d for d in ["السبت","الأحد","الاثنين","الثلاثاء","الأربعاء","الخميس","الجمعة"] if is_yes(r.get(d,""))]
                        days_txt="، ".join(days) if days else "لا توجد أيام محددة"
                        st.markdown(f'<div class="audit-row"><b>{task}</b><br><small>{days_txt} — {active}</small></div>',unsafe_allow_html=True)
            except Exception as e:
                st.error(f"خطأ: {e}")

        # ── تنظيف التكرارات ───────────────────────────────────────
        elif admin_tab=="🧹 تنظيف التكرارات":
            st.markdown("#### 🧹 تنظيف تكرار سجلات sheet1")
            st.info("يفحص النظام التكرار في نفس اليوم فقط حسب: التاريخ + الرقم الشخصي. اختاري التاريخ أولاً، ثم افتحي السهم للمعلمة واحذفي الصفوف الزائدة.")

            if st.button("🔍 فحص التكرارات الآن", use_container_width=True, type="primary"):
                st.session_state.duplicate_groups = find_duplicate_attendance_groups()

            duplicate_groups = st.session_state.get("duplicate_groups", None)
            if duplicate_groups is None:
                duplicate_groups = find_duplicate_attendance_groups()
                st.session_state.duplicate_groups = duplicate_groups

            if not duplicate_groups:
                st.success("✅ لا توجد سجلات مكررة حالياً في sheet1.")
            else:
                # ترتيب التكرارات حسب التاريخ حتى يكون واضح أن التكرار داخل نفس اليوم
                dates_available = sorted(set(k[0] for k in duplicate_groups.keys()), reverse=True)
                selected_date_filter = st.selectbox(
                    "اختاري التاريخ لعرض تكرارات نفس اليوم",
                    ["كل التواريخ"] + dates_available,
                    key="dup_date_filter"
                )

                filtered_groups = {
                    k: v for k, v in duplicate_groups.items()
                    if selected_date_filter == "كل التواريخ" or k[0] == selected_date_filter
                }

                st.warning(f"⚠️ تم العثور على {len(filtered_groups)} حالة تكرار حسب التاريخ + الرقم الشخصي.")

                # نجمع العرض تحت كل تاريخ
                groups_by_date = {}
                for (dup_date, dup_id), rows_list in filtered_groups.items():
                    groups_by_date.setdefault(dup_date, []).append((dup_id, rows_list))

                for dup_date in sorted(groups_by_date.keys(), reverse=True):
                    with st.expander(f"📅 تاريخ {dup_date} — عدد حالات التكرار: {len(groups_by_date[dup_date])}", expanded=(selected_date_filter != "كل التواريخ")):
                        for group_no, (dup_id, rows_list) in enumerate(groups_by_date[dup_date], start=1):
                            first_row = rows_list[0][1]
                            emp_name = first_row.get("الاسم الثلاثي", "") or first_row.get("الاسم", "")
                            task_name = first_row.get("المهمة", "")

                            with st.expander(f"🔁 {emp_name} — #{dup_id} — عدد السجلات في نفس اليوم: {len(rows_list)}", expanded=False):
                                options = []
                                option_map = {}
                                st.caption("اختاري فقط الصفوف الزائدة للحذف. انتبهي: الحذف من Google Sheet نهائي.")

                                for row_num, r in rows_list:
                                    label = f"صف {row_num} | التاريخ: {dup_date} | حضور: {r.get('وقت الحضور','—') or '—'} | انصراف: {r.get('وقت الانصراف','—') or '—'} | مهمة: {r.get('المهمة','—')}"
                                    options.append(label)
                                    option_map[label] = row_num
                                    st.markdown(f"""
                                    <div class="audit-row">
                                        <b>صف {row_num}</b><br>
                                        التاريخ: {dup_date}<br>
                                        الاسم: {r.get('الاسم الثلاثي','')}<br>
                                        الرقم الشخصي: {r.get('الرقم الشخصي','')}<br>
                                        المدرسة: {r.get('اسم المدرسة', r.get('المدرسة',''))}<br>
                                        المهمة: {r.get('المهمة','')}<br>
                                        وقت الحضور: {r.get('وقت الحضور','') or '—'}<br>
                                        وقت الانصراف: {r.get('وقت الانصراف','') or '—'}<br>
                                        سبب التأخير: {r.get('سبب التأخير','') or '—'}<br>
                                        سبب الانصراف: {r.get('سبب الانصراف','') or '—'}
                                    </div>
                                    """, unsafe_allow_html=True)

                                selected = st.multiselect(
                                    "حددي الصفوف التي تريدين حذفها",
                                    options,
                                    key=f"dup_select_{dup_date}_{dup_id}_{group_no}"
                                )
                                confirm_delete = st.checkbox(
                                    "أؤكد حذف الصفوف المحددة من sheet1",
                                    key=f"dup_confirm_{dup_date}_{dup_id}_{group_no}"
                                )
                                if st.button("🗑️ حذف الصفوف المحددة", key=f"dup_delete_{dup_date}_{dup_id}_{group_no}", use_container_width=True):
                                    if not selected:
                                        st.error("❌ لم تختاري أي صف للحذف.")
                                    elif not confirm_delete:
                                        st.error("❌ يجب تفعيل خانة التأكيد قبل الحذف.")
                                    elif len(selected) >= len(rows_list):
                                        st.error("❌ لا يمكن حذف كل سجلات المعلمة لهذا اليوم. اتركي سجل واحد صحيح على الأقل.")
                                    else:
                                        row_numbers = sorted([option_map[x] for x in selected], reverse=True)
                                        try:
                                            for rn in row_numbers:
                                                sheet.delete_rows(rn)
                                            log_audit(dup_id, emp_name, "حذف تكرار", f"التاريخ:{dup_date}|الصفوف المحذوفة:{row_numbers}|المهمة:{task_name}")
                                            clear_caches()
                                            st.session_state.duplicate_groups = find_duplicate_attendance_groups()
                                            st.success(f"✅ تم حذف {len(row_numbers)} صف/صفوف بنجاح من تكرارات تاريخ {dup_date}.")
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"❌ تعذر حذف الصفوف: {e}")

        # ── تعديل سجل ────────────────────────────────────────────
        elif admin_tab=="✏️ تعديل سجل":
            search_id=st.text_input("الرقم الشخصي",key="edit_id")
            search_date=st.date_input("التاريخ",value=now_bh().date(),key="edit_date")
            if st.button("بحث",key="btn_search"):
                data=get_sheet_data(); idx,row=find_today_row(data,str(search_date),search_id)
                if row: st.session_state.edit_row_idx=idx; st.session_state.edit_row=row
                else: st.error("لا يوجد سجل"); st.session_state.edit_row=None
            if st.session_state.get("edit_row"):
                row=st.session_state.edit_row; idx=st.session_state.edit_row_idx
                st.info(f"السجل: {row.get('الاسم الثلاثي','')} — {row.get('التاريخ','')}")
                new_att=st.text_input("وقت الحضور",value=row.get("وقت الحضور",""),key="new_att")
                new_dep=st.text_input("وقت الانصراف",value=row.get("وقت الانصراف",""),key="new_dep")
                edit_reason=st.text_input("سبب التعديل (مطلوب)",key="edit_reason")
                if st.button("حفظ التعديل",use_container_width=True):
                    if not edit_reason.strip(): st.error("سبب التعديل مطلوب")
                    else:
                        safe_update(sheet,idx,COL_ATTEND,new_att); safe_update(sheet,idx,COL_DEPART,new_dep)
                        row["وقت الحضور"] = new_att
                        row["وقت الانصراف"] = new_dep
                        update_work_calculation(idx, row)
                        log_audit(search_id,row.get("الاسم الثلاثي",""),"تعديل أدمن",f"حضور:{row.get('وقت الحضور','')}→{new_att}|انصراف:{row.get('وقت الانصراف','')}→{new_dep}|السبب:{edit_reason}")
                        clear_caches(); st.success("✅ تم الحفظ"); st.session_state.edit_row=None

        # ── تسجيل يدوي ──────────────────────────────────────────
        elif admin_tab=="➕ تسجيل يدوي":
            st.markdown("#### ➕ تسجيل حضور/انصراف يدوي")
            st.info("يتم التسجيل اليدوي بنفس ترتيب أعمدة sheet1 بدون لخبطة: التاريخ، اليوم، المدرسة، المهمة، دعم، الاسم، الرقم، وقت الحضور...")

            m_id=ar_to_en_digits(st.text_input("الرقم الشخصي",key="mid")).strip()
            m_date=st.date_input("التاريخ",value=now_bh().date(),key="mdate")
            m_att=st.text_input("وقت الحضور",value="07:00:00",key="matt")
            m_dep=st.text_input("وقت الانصراف (اختياري)",key="mdep")
            m_note=st.text_input("سبب الإضافة اليدوية (مطلوب)",key="mnote")

            emp_preview = validate_employee(m_id) if m_id else None
            manual_support_mode = False
            if m_id and not emp_preview:
                st.warning("⚠️ الرقم غير موجود في القائمة البيضاء. إذا كانت الموظفة دعم خارجي، فعّلي الخيار التالي وأدخلي بياناتها يدويًا.")
                manual_support_mode = st.checkbox("تسجيل كدعم خارجي غير موجود في القائمة البيضاء", key="manual_support_mode")
            if m_id and emp_preview:
                st.success(f"✅ تم العثور على البيانات: {emp_preview.get('الاسم','')} — {emp_preview.get('المدرسة','')} — {emp_preview.get('المهمة','')}")

            if manual_support_mode:
                m_support_name = st.text_input("اسم الدعم", key="manual_support_name")
                school_choice = st.selectbox("مدرسة الدعم", schools + ["أخرى"], key="manual_support_school_choice")
                if school_choice == "أخرى":
                    m_support_school = st.text_input("اكتبي اسم المدرسة", key="manual_support_school_other").strip()
                else:
                    m_support_school = school_choice
                m_support_task = st.selectbox("مهمة الدعم", TASKS_SUPPORT, key="manual_support_task")

            if st.button("تسجيل يدوي",use_container_width=True,type="primary"):
                if not m_id:
                    st.error("❌ الرقم الشخصي مطلوب")
                elif not m_note.strip():
                    st.error("❌ سبب الإضافة اليدوية مطلوب")
                else:
                    emp=validate_employee(m_id)
                    if not emp and not manual_support_mode:
                        st.error("❌ الرقم غير موجود في القائمة البيضاء. إذا كانت الموظفة دعم خارجي فعّلي خيار تسجيل كدعم خارجي.")
                    else:
                        if not emp and manual_support_mode:
                            if not m_support_name.strip() or not m_support_school.strip():
                                st.error("❌ اسم الدعم والمدرسة مطلوبان")
                                st.stop()
                            emp = {"الاسم": normalize_name(m_support_name), "المدرسة": m_support_school, "المهمة": m_support_task, "دعم": "نعم"}
                        date_str=str(m_date)
                        day_name=m_date.strftime("%A")
                        task=str(emp.get("المهمة","")).strip()

                        # تحديد قيمة عمود دعم من القائمة البيضاء أو من اسم المهمة أو خيار الدعم اليدوي
                        support_raw=str(emp.get("دعم","")).strip()
                        support_value="نعم" if manual_support_mode or support_raw in ["نعم","yes","Yes","TRUE","true","1"] or "دعم" in task else "لا"
                        full_name=normalize_name(emp.get("الاسم",""))

                        row_data=[
                            date_str,                       # A التاريخ
                            day_name,                       # B اليوم
                            emp.get("المدرسة",""),          # C اسم المدرسة
                            task,                           # D المهمة
                            support_value,                  # E دعم
                            full_name,                      # F الاسم الثلاثي
                            m_id,                           # G الرقم الشخصي
                            m_att,                          # H وقت الحضور
                            f"[يدوي] {m_note.strip()}",     # I سبب التأخير
                            m_dep,                          # J وقت الانصراف
                            "",                             # K سبب الانصراف
                            "",                             # L خروج استئذان
                            "",                             # M عودة
                            "",                             # N محاولة
                            "", "", "", "", "", "", "", "", # O:V احتسابات لاحقة
                            "تسجيل يدوي" if support_value != "نعم" else "تسجيل دعم يدوي"  # W نوع التسجيل
                        ]

                        existing_matches = find_daily_rows_fresh(date_str, m_id)
                        existing_idx, existing_row = pick_main_daily_row(existing_matches)
                        if existing_idx and existing_row and str(existing_row.get("وقت الحضور", "")).strip():
                            st.error(f"❌ يوجد سجل حضور مسبق لهذا الرقم في تاريخ {date_str} الساعة {existing_row.get('وقت الحضور','')}. لا يمكن إضافة سجل يدوي مكرر.")
                        elif len(existing_matches) > 1:
                            st.error("❌ يوجد تكرار سابق لهذا الرقم في نفس التاريخ. نظّفي التكرارات أولاً من قسم تنظيف التكرارات.")
                        elif existing_idx:
                            safe_update(sheet,existing_idx,COL_ATTEND,m_att)
                            safe_update(sheet,existing_idx,COL_LATE_REASON,f"[يدوي] {m_note.strip()}")
                            safe_update(sheet,existing_idx,COL_DEPART,m_dep)
                            existing_row["وقت الحضور"] = m_att
                            existing_row["وقت الانصراف"] = m_dep
                            update_work_calculation(existing_idx,existing_row)
                            log_audit(m_id,full_name,"تسجيل يدوي",f"تحديث سجل موجود|التاريخ:{date_str}|حضور:{m_att}|انصراف:{m_dep}|السبب:{m_note.strip()}")
                            clear_caches()
                            st.success("✅ تم تحديث السجل الموجود بدون إنشاء تكرار")
                        else:
                            ok=safe_append(sheet,row_data)
                            if ok:
                                try:
                                    data_after=get_sheet_data_fresh()
                                    idx_after,row_after=find_today_row(data_after,date_str,m_id)
                                    if idx_after:
                                        update_work_calculation(idx_after,row_after)
                                except Exception:
                                    pass
                                log_audit(
                                    m_id,
                                    full_name,
                                    "تسجيل يدوي",
                                    f"التاريخ:{date_str}|حضور:{m_att}|انصراف:{m_dep}|السبب:{m_note.strip()}"
                                )
                                clear_caches()
                                st.success("✅ تم التسجيل اليدوي بنجاح وبالأعمدة الصحيحة")
                            else:
                                st.error("❌ تعذر حفظ التسجيل اليدوي، حاولي مرة أخرى")

        # ── القائمة البيضاء ──────────────────────────────────────
        elif admin_tab=="📋 القائمة البيضاء":

            # ── تنظيف تكرارات القائمة البيضاء ──
            with st.container(border=True):
                st.markdown("##### 🧹 تنظيف تكرارات القائمة البيضاء")
                st.caption("يبحث عن الأرقام الشخصية المكررة ويحذف النسخ الزائدة — يبقي آخر صف لكل رقم.")

                try:
                    all_wl = whitelist_sheet.get_all_records()
                    dup_ids = {}
                    for i, r in enumerate(all_wl):
                        eid = str(r.get("الرقم الشخصي","")).strip()
                        if eid:
                            dup_ids.setdefault(eid, []).append(i+2)
                    dups = {k:v for k,v in dup_ids.items() if len(v)>1}

                    if not dups:
                        st.success("✅ لا توجد تكرارات في القائمة البيضاء.")
                    else:
                        total_dup = sum(len(v)-1 for v in dups.values())
                        st.warning(f"⚠️ وجد {len(dups)} رقم مكرر — إجمالي الصفوف الزائدة: {total_dup}")
                        for eid, rows in list(dups.items())[:5]:
                            name = str(all_wl[rows[0]-2].get("الاسم","")).strip()
                            st.markdown(f"- **{name}** (#{eid}) — {len(rows)} مرة")
                        if len(dups) > 5:
                            st.markdown(f"- ... و{len(dups)-5} آخرين")

                        if st.button("🗑️ حذف التكرارات — إبقاء الصف الأكمل بياناتً لكل رقم",
                                     use_container_width=True, type="primary", key="btn_clean_wl_dups"):
                            rows_to_delete = []
                            for eid, rows in dups.items():
                                # احسب اكتمال كل صف (عدد الخلايا غير الفارغة)
                                scored = []
                                for rn in rows:
                                    r = all_wl[rn-2]
                                    score = sum(1 for v in r.values() if str(v).strip())
                                    scored.append((score, rn))
                                # ابقي الصف الأكمل، احذف الباقي
                                best_rn = max(scored, key=lambda x: x[0])[1]
                                rows_to_delete.extend([rn for rn in rows if rn != best_rn])
                            # نحذف من الأكبر للأصغر عشان ما تتأثر الأرقام
                            rows_to_delete = sorted(set(rows_to_delete), reverse=True)
                            deleted = 0
                            errors  = 0
                            prog = st.progress(0)
                            for j, rn in enumerate(rows_to_delete):
                                prog.progress((j+1)/len(rows_to_delete))
                                try:
                                    whitelist_sheet.delete_rows(rn)
                                    deleted += 1
                                except Exception:
                                    errors += 1
                            get_whitelist.clear()
                            log_audit("—","أدمن","تنظيف تكرارات القائمة البيضاء",
                                      f"محذوف:{deleted}|أخطاء:{errors}")
                            st.success(f"✅ تم حذف {deleted} صف مكرر. أخطاء: {errors}")
                            st.rerun()
                except Exception as e:
                    st.error(f"❌ خطأ: {e}")

            st.markdown("---")

            # ── مزامنة بيانات القائمة البيضاء مع شيت1 ──
            with st.container(border=True):
                st.markdown("##### 🔄 مزامنة بيانات موظفة مع شيت1")
                st.caption("لما تتعدل بيانات موظفة في القائمة البيضاء (اسم / مدرسة / مهمة)، اضغطي هنا لتحديث كل سجلاتها في شيت1.")

                sync_id_raw = st.text_input("الرقم الشخصي للمزامنة", key="sync_wl_id")
                sync_id = ar_to_en_digits(sync_id_raw).strip()
                sync_all = st.checkbox("مزامنة كل القائمة البيضاء مع شيت1", key="sync_wl_all")

                if sync_id and not sync_all:
                    found_sync = validate_employee(sync_id)
                    if found_sync:
                        st.success(f"✅ {found_sync.get('الاسم','')} — {found_sync.get('المدرسة','')} — {found_sync.get('المهمة','')}")
                    else:
                        st.warning("⚠️ الرقم غير موجود في القائمة البيضاء")

                if st.button("🔄 تنفيذ المزامنة", use_container_width=True, type="primary", key="btn_sync_wl"):
                    try:
                        wl_data    = get_whitelist()
                        sheet_data = get_sheet_data_fresh()
                        updated    = 0
                        errors     = 0
                        prog       = st.progress(0)

                        # حدد الموظفات للمزامنة
                        if sync_all:
                            sync_ids = list(wl_data.keys())
                        elif sync_id and sync_id in wl_data:
                            sync_ids = [sync_id]
                        else:
                            st.error("❌ أدخلي رقماً صحيحاً أو اختاري مزامنة الكل.")
                            sync_ids = []

                        total = len(sync_ids)
                        for j, eid in enumerate(sync_ids):
                            prog.progress((j+1)/max(total,1))
                            emp_wl = wl_data[eid]
                            new_name   = str(emp_wl.get("الاسم","")).strip()
                            new_school = str(emp_wl.get("المدرسة","")).strip()
                            new_task   = str(emp_wl.get("المهمة","")).strip()

                            for i, row in enumerate(sheet_data):
                                if str(row.get("الرقم الشخصي","")).strip() != eid:
                                    continue
                                row_num = i + 2
                                try:
                                    # أعمدة: الاسم=6، المدرسة=3، المهمة=4
                                    if new_name   and str(row.get("الاسم الثلاثي","")).strip()  != new_name:
                                        safe_update(sheet, row_num, COL_NAME, new_name)
                                    if new_school and str(row.get("اسم المدرسة","")).strip() != new_school:
                                        safe_update(sheet, row_num, 3, new_school)
                                    if new_task   and str(row.get("المهمة","")).strip()       != new_task:
                                        safe_update(sheet, row_num, 4, new_task)
                                    updated += 1
                                except Exception:
                                    errors += 1

                        clear_caches()
                        log_audit("—","أدمن","مزامنة القائمة البيضاء مع شيت1",
                                  f"محدّث:{updated}|أخطاء:{errors}|{'الكل' if sync_all else sync_id}")
                        st.success(f"✅ تم تحديث {updated} سجل في شيت1. أخطاء: {errors}")
                    except Exception as e:
                        st.error(f"❌ خطأ: {e}")

            st.markdown("---")
            st.markdown("#### إضافة موظفة")
            wl_id=ar_to_en_digits(st.text_input("الرقم الشخصي",key="wlid")).strip()
            wl_name=st.text_input("الاسم",key="wlname")
            wl_school=st.selectbox("المدرسة",schools + ["أخرى"],key="wlschool")
            if wl_school == "أخرى":
                wl_school_final = st.text_input("اكتبي اسم المدرسة", key="wlschool_other").strip()
            else:
                wl_school_final = wl_school
            wl_task=st.selectbox("المهمة",TASKS_ALL,key="wltask")
            wl_job=st.selectbox("المسمى الوظيفي",JOB_TITLES,key="wljob")
            wl_phone=st.text_input("رقم التواصل", key="wlphone")
            wl_email=st.text_input("البريد الإلكتروني", key="wlemail")
            if st.button("إضافة",use_container_width=True):
                if not wl_id.strip() or not wl_name.strip() or not wl_school_final.strip():
                    st.error("الرقم والاسم والمدرسة مطلوبة")
                else:
                    ok=safe_append(whitelist_sheet,[wl_id,normalize_name(wl_name),wl_school_final,wl_task,"نعم" if "دعم" in wl_task else "لا",wl_phone,wl_email,wl_job,"نعم"])
                    get_whitelist.clear()
                    if ok:
                        log_audit(wl_id, normalize_name(wl_name), "إضافة للقائمة البيضاء", f"المدرسة:{wl_school_final}|المهمة:{wl_task}")
                        st.success("✅ تمت الإضافة")
                    else: st.error("❌ تعذرت الإضافة")

            st.markdown("---")
            st.markdown("#### ✏️ بحث وتعديل بيانات موظفة")
            search_wl = st.text_input("ابحثي بالرقم الشخصي أو الاسم", key="wl_search_edit")
            wl_records_full = []
            try:
                wl_records_full = whitelist_sheet.get_all_records()
            except Exception:
                wl_records_full = []

            matches = []
            if search_wl.strip():
                q = normalize_name(search_wl)
                for i, r in enumerate(wl_records_full, start=2):
                    eid = str(r.get("الرقم الشخصي", "")).strip()
                    nm = normalize_name(r.get("الاسم", ""))
                    if q in normalize_name(eid) or q in nm:
                        matches.append((i, r))
            else:
                matches = list(enumerate(wl_records_full[-30:], start=max(2, len(wl_records_full)-28)))
                st.caption("يعرض آخر 30 سجل. للبحث الدقيق اكتبي الرقم الشخصي أو جزءًا من الاسم.")

            if not matches:
                st.info("لا توجد نتائج مطابقة.")
            else:
                for row_num, emp in matches[:20]:
                    eid = str(emp.get("الرقم الشخصي", "")).strip()
                    with st.expander(f"✏️ {emp.get('الاسم','')} — #{eid} — {emp.get('المدرسة','')}", expanded=False):
                        edit_name = st.text_input("الاسم", value=emp.get("الاسم", ""), key=f"edit_wl_name_{row_num}")
                        school_current = emp.get("المدرسة", "") if emp.get("المدرسة", "") in schools else "أخرى"
                        edit_school_choice = st.selectbox("المدرسة", schools + ["أخرى"], index=(schools + ["أخرى"]).index(school_current), key=f"edit_wl_school_{row_num}")
                        if edit_school_choice == "أخرى":
                            edit_school = st.text_input("اكتبي اسم المدرسة", value=emp.get("المدرسة", ""), key=f"edit_wl_school_other_{row_num}").strip()
                        else:
                            edit_school = edit_school_choice
                        task_list = TASKS_ALL
                        task_current = emp.get("المهمة", "") if emp.get("المهمة", "") in task_list else task_list[0]
                        edit_task = st.selectbox("المهمة", task_list, index=task_list.index(task_current), key=f"edit_wl_task_{row_num}")
                        edit_support = st.selectbox("دعم", ["لا", "نعم"], index=1 if is_yes(emp.get("دعم", "")) or "دعم" in str(emp.get("المهمة", "")) else 0, key=f"edit_wl_support_{row_num}")
                        edit_phone = st.text_input("رقم التواصل", value=emp.get("رقم التواصل", ""), key=f"edit_wl_phone_{row_num}")
                        edit_email = st.text_input("البريد الإلكتروني", value=emp.get("البريد الإلكتروني", ""), key=f"edit_wl_email_{row_num}")
                        job_current = emp.get("المسمى الوظيفي", "") if emp.get("المسمى الوظيفي", "") in JOB_TITLES else JOB_TITLES[0]
                        edit_job = st.selectbox("المسمى الوظيفي", JOB_TITLES, index=JOB_TITLES.index(job_current), key=f"edit_wl_job_{row_num}")
                        edit_active = st.selectbox("نشط", ["نعم", "لا"], index=0 if is_yes(emp.get("نشط", "")) else 1, key=f"edit_wl_active_{row_num}")
                        update_today = st.checkbox("تحديث سجل اليوم تلقائيًا أيضًا", value=True, key=f"edit_wl_update_today_{row_num}")
                        edit_reason = st.text_input("سبب التعديل", value="تصحيح بيانات الموظفة", key=f"edit_wl_reason_{row_num}")

                        if st.button("💾 حفظ تعديل بيانات الموظفة", key=f"save_wl_edit_{row_num}", use_container_width=True, type="primary"):
                            if not edit_name.strip() or not edit_school.strip() or not edit_reason.strip():
                                st.error("❌ الاسم والمدرسة وسبب التعديل مطلوبة")
                            else:
                                try:
                                    values = [
                                        eid,
                                        normalize_name(edit_name),
                                        edit_school,
                                        edit_task,
                                        edit_support,
                                        edit_phone,
                                        edit_email,
                                        edit_job,
                                        edit_active
                                    ]
                                    whitelist_sheet.update(f"A{row_num}:I{row_num}", [values], value_input_option="USER_ENTERED")
                                    get_whitelist.clear()
                                    details = f"الاسم:{emp.get('الاسم','')}→{normalize_name(edit_name)}|المدرسة:{emp.get('المدرسة','')}→{edit_school}|المهمة:{emp.get('المهمة','')}→{edit_task}|دعم:{edit_support}|نشط:{edit_active}|السبب:{edit_reason}"

                                    if update_today:
                                        today_matches = find_daily_rows_fresh(today_str, eid)
                                        for idx_today, row_today in today_matches:
                                            safe_update(sheet, idx_today, COL_SCHOOL, edit_school)
                                            safe_update(sheet, idx_today, COL_TASK, edit_task)
                                            safe_update(sheet, idx_today, COL_SUPPORT, edit_support)
                                            safe_update(sheet, idx_today, COL_NAME, normalize_name(edit_name))
                                        if today_matches:
                                            details += f"|تم تحديث سجل اليوم للصفوف:{[x[0] for x in today_matches]}"

                                    log_audit(eid, normalize_name(edit_name), "تعديل القائمة البيضاء", details)
                                    clear_caches()
                                    st.success("✅ تم تعديل بيانات الموظفة وتحديث سجل اليوم تلقائيًا")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ تعذر حفظ التعديل: {e}")

            st.markdown("---")
            st.markdown("#### الموظفات المسجّلات")
            wl_all=get_whitelist()
            for eid,emp in list(wl_all.items())[-50:]:
                st.markdown(f'<div class="audit-row"><b>{emp.get("الاسم","")}</b> — #{eid} — {emp.get("المدرسة","")} — {emp.get("المهمة","")}</div>',unsafe_allow_html=True)

        # ── محاولات التسجيل ─────────────────────────────────────
        elif admin_tab=="🚫 محاولات التسجيل":
            try:
                records=attempts_sheet.get_all_records()
                today_att=[r for r in records if r.get("التاريخ")==today_str]
                st.metric("محاولات اليوم",len(today_att))
                for r in reversed(today_att[-50:]):
                    st.markdown(f'<div class="warn-row">⚠️ {r.get("اسم_المقفول_عليه","")} ({r.get("الرقم_المقفول_عليه","")}) ← حاول: {r.get("اسم_المحاول","")} — {r.get("وقت_المحاولة","")}</div>',unsafe_allow_html=True)
            except Exception as e: st.error(f"خطأ: {e}")

        # ── تجاوز الموقع ────────────────────────────────────────
        elif admin_tab=="📡 تجاوز الموقع":
            st.markdown("#### 📡 إدارة تجاوز الموقع")
            active,end_dt=get_location_override()

            if active and end_dt:
                remaining_seconds = max(0, int((end_dt - now_bh()).total_seconds()))
                remaining_minutes = remaining_seconds // 60
                st.success(f"✅ تجاوز الموقع مفعّل حالياً حتى الساعة {end_dt.strftime('%H:%M')}")
                st.info(f"المتبقي تقريباً: {remaining_minutes} دقيقة. الموظفات يمكنهن التسجيل بدون تحديد الموقع خلال هذه المدة.")

                if st.button("🔴 تعطيل تجاوز الموقع الآن", use_container_width=True, type="primary"):
                    ok = disable_location_override()
                    if ok:
                        st.success("✅ تم تعطيل تجاوز الموقع.")
                    st.rerun()

            else:
                st.warning("⚠️ تجاوز الموقع غير مفعّل حالياً.")
                duration=st.selectbox("مدة تفعيل تجاوز الموقع بالدقائق",[30,60,90,120,180], key="override_duration")
                reason=st.text_input("سبب تفعيل تجاوز الموقع", key="override_reason")

                if st.button("🟢 تفعيل تجاوز الموقع", use_container_width=True, type="primary"):
                    if not reason.strip():
                        st.error("❌ السبب مطلوب.")
                    else:
                        ok,end_dt=set_location_override(duration,reason.strip())
                        if ok and end_dt:
                            # نتحقق مباشرة من الشيت بعد الحفظ حتى نضمن أن الزر سيتغير للتعطيل
                            get_settings_records.clear()
                            active_check,end_check=get_location_override()
                            if active_check:
                                st.success(f"✅ تم تفعيل تجاوز الموقع حتى الساعة {end_check.strftime('%H:%M')}")
                                st.rerun()
                            else:
                                st.error("❌ تم الحفظ لكن لم يظهر التفعيل. تأكدي من شيت إعدادات_النظام أن الأعمدة هي: المفتاح، القيمة، تاريخ_الانتهاء، ملاحظات.")
                        else:
                            st.error("❌ تعذر تفعيل تجاوز الموقع.")

        # ── إدارة قفل الجهاز ────────────────────────────────────────
        elif admin_tab=="⚙️ قفل الجهاز":
            st.markdown("#### ⚙️ إدارة قفل الجهاز")
            st.info("الوضع الطبيعي: نفس الجهاز/المتصفح يسجل لشخص واحد فقط طوال اليوم. يمكن تعطيله مؤقتًا للجميع أو استثناء رقم معيّن.")

            global_off, global_end = get_device_lock_global_override()
            if global_off and global_end:
                remaining = max(0, int((global_end - now_bh()).total_seconds() // 60))
                st.warning(f"⚠️ قفل الجهاز معطّل للجميع مؤقتًا حتى {global_end.strftime('%H:%M')} — المتبقي {remaining} دقيقة")
                if st.button("🔒 إعادة تفعيل قفل الجهاز للجميع الآن", use_container_width=True, type="primary"):
                    if disable_device_lock_global_override():
                        log_audit("—", "أدمن", "تفعيل قفل الجهاز", "إلغاء التعطيل العام لقفل الجهاز")
                        st.success("✅ تم إعادة تفعيل قفل الجهاز.")
                        st.rerun()
            else:
                st.success("✅ قفل الجهاز مفعّل حاليًا للجميع.")
                mins = st.selectbox("مدة تعطيل قفل الجهاز للجميع", [15, 30, 60, 90, 120, 180], key="dev_global_minutes")
                note = st.text_input("سبب تعطيل قفل الجهاز للجميع", key="dev_global_note")
                if st.button("🔓 تعطيل قفل الجهاز للجميع مؤقتًا", use_container_width=True):
                    if not note.strip():
                        st.error("❌ السبب مطلوب")
                    else:
                        ok, end_dt = set_device_lock_global_override(mins, note.strip())
                        if ok:
                            log_audit("—", "أدمن", "تعطيل قفل الجهاز للجميع", f"حتى:{end_dt.strftime('%Y-%m-%d %H:%M:%S')}|السبب:{note.strip()}")
                            st.success(f"✅ تم تعطيل قفل الجهاز للجميع حتى {end_dt.strftime('%H:%M')}")
                            st.rerun()

            st.markdown("---")
            st.markdown("#### 👤 استثناء رقم شخصي من قفل الجهاز")
            ex_id = ar_to_en_digits(st.text_input("الرقم الشخصي للاستثناء", key="dev_ex_id")).strip()
            ex_emp = validate_employee(ex_id) if ex_id else None
            ex_name = ex_emp.get("الاسم", "") if ex_emp else ""
            if ex_id and ex_emp:
                st.success(f"الموظفة: {ex_name} — {ex_emp.get('المدرسة','')} — {ex_emp.get('المهمة','')}")
            elif ex_id:
                st.warning("الرقم غير موجود في القائمة البيضاء، يمكن إضافة الاستثناء بالرقم فقط إذا لزم.")
                ex_name = st.text_input("اسم الموظفة للاستثناء", key="dev_ex_name_manual")
            ex_minutes = st.selectbox("مدة الاستثناء", [15, 30, 60, 90, 120, 180, 240], key="dev_ex_minutes")
            ex_note = st.text_input("سبب الاستثناء", key="dev_ex_note")
            cex1, cex2 = st.columns(2)
            with cex1:
                if st.button("➕ إضافة/تفعيل استثناء", use_container_width=True, type="primary"):
                    if not ex_id or not ex_note.strip():
                        st.error("❌ الرقم الشخصي وسبب الاستثناء مطلوبان")
                    else:
                        ok, end_dt = add_device_exception_for_employee(ex_id, ex_name, ex_minutes, ex_note.strip())
                        if ok:
                            log_audit(ex_id, ex_name or "—", "استثناء قفل الجهاز", f"حتى:{end_dt.strftime('%Y-%m-%d %H:%M:%S')}|السبب:{ex_note.strip()}")
                            st.success(f"✅ تم إضافة الاستثناء حتى {end_dt.strftime('%H:%M')}")
                            st.rerun()
            with cex2:
                if st.button("🛑 تعطيل استثناءات هذا الرقم", use_container_width=True):
                    if not ex_id:
                        st.error("❌ أدخلي الرقم الشخصي")
                    else:
                        changed = disable_device_exception_for_employee(ex_id)
                        log_audit(ex_id, ex_name or "—", "تعطيل استثناء قفل الجهاز", f"عدد الاستثناءات المعطلة:{changed}")
                        st.success(f"✅ تم تعطيل {changed} استثناء/استثناءات لهذا الرقم")
                        st.rerun()

            st.markdown("---")
            st.markdown("#### الاستثناءات النشطة")
            active_ex = []
            for r in get_device_exceptions():
                if is_yes(r.get("نشط", "")):
                    end_dt = parse_bahrain_datetime(r.get("تاريخ_الانتهاء", ""))
                    if end_dt and now_bh() < end_dt:
                        active_ex.append((r, end_dt))
            if not active_ex:
                st.success("لا توجد استثناءات نشطة حاليًا.")
            else:
                for r, end_dt in active_ex[-50:]:
                    st.markdown(f'<div class="audit-row">🔓 {r.get("الاسم","")} — #{r.get("الرقم الشخصي","")} — حتى {end_dt.strftime("%H:%M")}<br><small>{r.get("ملاحظات","")}</small></div>', unsafe_allow_html=True)


        # ── الأجهزة الموثوقة ───────────────────────────────────────
        elif admin_tab=="📱 الأجهزة الموثوقة":
            st.markdown("#### 📱 الأجهزة الموثوقة")
            st.info("افتحي هذه الصفحة من جهاز الحضور الاحتياطي داخل المركز، ثم اعتمديه. الجهاز الموثوق يسمح بتسجيل أكثر من موظفة من نفس الجهاز بدون قفل الجهاز.")
            current_fp = get_device_fingerprint()
            trusted_now, trusted_row = is_current_device_trusted()
            if trusted_now:
                st.success(f"✅ هذا الجهاز موثوق حاليًا: {trusted_row.get('اسم الجهاز','جهاز موثوق')}")
            else:
                st.warning("⚠️ هذا الجهاز غير معتمد كجهاز موثوق.")
            dev_name = st.text_input("اسم الجهاز", value="جهاز الحضور الاحتياطي", key="trusted_dev_name")
            dev_note = st.text_input("ملاحظات", value="جهاز مخصص للتسجيل داخل المركز", key="trusted_dev_note")
            if st.button("⭐ اعتماد هذا الجهاز كجهاز موثوق", use_container_width=True, type="primary"):
                if approve_current_device_as_trusted(dev_name, dev_note):
                    log_audit("—", "أدمن", "اعتماد جهاز موثوق", f"اسم الجهاز:{dev_name}|بصمة:{current_fp[:8]}...")
                    st.success("✅ تم اعتماد هذا الجهاز كجهاز موثوق.")
                    st.rerun()
            st.markdown("---")
            st.markdown("#### قائمة الأجهزة الموثوقة")
            rows = get_trusted_devices()
            active_rows = [r for r in rows if is_yes(r.get("نشط", ""))]
            if not active_rows:
                st.info("لا توجد أجهزة موثوقة نشطة.")
            else:
                for idx, r in enumerate(active_rows, start=1):
                    fp = str(r.get("بصمة الجهاز", "")).strip()
                    with st.expander(f"📱 {r.get('اسم الجهاز','جهاز موثوق')} — آخر استخدام: {r.get('آخر استخدام','')}"):
                        st.write(f"ملاحظات: {r.get('ملاحظات','')}")
                        st.code(fp)
                        if st.button("🛑 تعطيل هذا الجهاز", key=f"disable_trusted_{idx}", use_container_width=True):
                            if disable_trusted_device(fp):
                                log_audit("—", "أدمن", "تعطيل جهاز موثوق", f"اسم الجهاز:{r.get('اسم الجهاز','')}|بصمة:{fp[:8]}...")
                                st.success("✅ تم تعطيل الجهاز.")
                                st.rerun()

        # ── فتح قفل جهاز ────────────────────────────────────────
        elif admin_tab=="🔓 فتح قفل جهاز":
            st.info("استخدمي هذه الخاصية عند وجود سبب رسمي.")
            unlock_id=st.text_input("الرقم الشخصي المقفول عليه")
            if st.button("حذف قفل اليوم لهذا الرقم",use_container_width=True):
                try:
                    records=device_sheet.get_all_records(); deleted=False
                    for i,r in enumerate(records):
                        if str(r.get("التاريخ","")).strip()==today_str and str(r.get("الرقم الشخصي","")).strip()==unlock_id.strip():
                            device_sheet.delete_rows(i+2); deleted=True; break
                    get_device_locks.clear()
                    if deleted: log_audit(unlock_id,"—","فتح قفل أدمن",f"فتح قفل الرقم {unlock_id} ليوم {today_str}"); st.success("✅ تم حذف القفل.")
                    else: st.warning("لم يتم العثور على قفل لهذا الرقم اليوم.")
                except Exception as e: st.error(f"خطأ: {e}")

        # ── سجل التدقيق ─────────────────────────────────────────
        elif admin_tab=="🔍 سجل التدقيق":
            try:
                audit_data=audit_sheet.get_all_records()
                for r in reversed(audit_data[-30:]):
                    st.markdown(f'<div class="audit-row"><b>{r.get("نوع العملية","")}</b> — {r.get("الوقت","")}<br><small>{r.get("المستخدم","")} (#{r.get("الرقم الشخصي","")}) — {r.get("التفاصيل","")}</small></div>',unsafe_allow_html=True)
            except Exception as e: st.error(f"خطأ: {e}")

        # ── تقرير الأجهزة ────────────────────────────────────────
        elif admin_tab=="⚠️ تقرير الأجهزة":
            try:
                audit_data=audit_sheet.get_all_records(); device_map={}
                for r in audit_data:
                    if r.get("التاريخ")==today_str and r.get("نوع العملية")=="تسجيل حضور":
                        fp_key=str(r.get("بصمة الجهاز",""))
                        if fp_key not in device_map: device_map[fp_key]=[]
                        name=f"{r.get('المستخدم','')} (#{r.get('الرقم الشخصي','')})"
                        if name not in device_map[fp_key]: device_map[fp_key].append(name)
                found=False
                for fp_key,names in device_map.items():
                    if len(names)>1:
                        found=True
                        st.markdown(f'<div class="warn-row">⚠️ جهاز واحد سجّل لـ {len(names)} موظفات: {"، ".join(names)}</div>',unsafe_allow_html=True)
                if not found: st.success("✅ لا يوجد تسجيل مشبوه اليوم")
            except Exception as e: st.error(f"خطأ: {e}")

        # ── تصاريح الوقت اليدوي ──────────────────────────────────
        elif admin_tab=="⏰ تصاريح الوقت اليدوي":
            st.markdown("#### ⏰ تصاريح الوقت اليدوي")
            st.info("تسمح للموظفة بتعديل وقت الحضور أو الانصراف يدوياً لتواريخ محددة، بدون GPS وبدون انتظار اعتماد.")

            # ── إضافة تصريح جديد ──
            with st.container(border=True):
                st.markdown("##### ➕ إضافة تصريح جديد")

                permit_scope = st.radio("التصريح لـ", ["موظفة محددة","كل الموظفات"], horizontal=True, key="permit_scope")
                if permit_scope == "موظفة محددة":
                    permit_id_raw = st.text_input("الرقم الشخصي", key="permit_emp_id")
                    permit_id = ar_to_en_digits(permit_id_raw).strip()
                    if permit_id:
                        found_emp = validate_employee(permit_id)
                        if found_emp:
                            st.success(f"✅ {found_emp.get('الاسم','')} — {found_emp.get('المدرسة','')}")
                        else:
                            st.warning("⚠️ الرقم غير موجود في القائمة البيضاء")
                else:
                    permit_id = "الكل"
                    st.warning("⚠️ سيُفعَّل التصريح لجميع الموظفات")

                permit_type = st.selectbox("نوع التصريح", ["كليهما","حضور","انصراف"], key="permit_type")

                st.markdown("**نطاق التاريخ**")
                col_d1, col_d2 = st.columns(2)
                with col_d1:
                    permit_date_from = st.date_input("من تاريخ", value=now_bh().date(), key="permit_from")
                with col_d2:
                    permit_date_to   = st.date_input("إلى تاريخ", value=now_bh().date(), key="permit_to")

                st.markdown("**نافذة الوقت** (اتركيها فارغة ليكون مفتوحاً طول اليوم)")
                col_t1, col_t2 = st.columns(2)
                with col_t1:
                    all_day = st.checkbox("يوم كامل (بدون قيد وقت)", value=True, key="permit_allday")
                with col_t2:
                    permit_note = st.text_input("ملاحظة", key="permit_note")

                if not all_day:
                    col_t3, col_t4 = st.columns(2)
                    with col_t3:
                        permit_open  = st.time_input("وقت الفتح",  value=time(6,0),  key="permit_open")
                    with col_t4:
                        permit_close = st.time_input("وقت الإغلاق", value=time(15,0), key="permit_close")
                    t_open  = permit_open.strftime("%H:%M")
                    t_close = permit_close.strftime("%H:%M")
                else:
                    t_open = t_close = ""

                if permit_date_from > permit_date_to:
                    st.error("❌ تاريخ البداية يجب أن يكون قبل تاريخ النهاية.")
                elif st.button("✅ تفعيل التصريح", use_container_width=True, type="primary", key="btn_add_permit"):
                    if not permit_id:
                        st.error("❌ أدخل الرقم الشخصي أولاً.")
                    else:
                        add_time_permit(
                            permit_id, permit_type,
                            permit_date_from.strftime("%Y-%m-%d"),
                            permit_date_to.strftime("%Y-%m-%d"),
                            t_open, t_close, permit_note
                        )
                        log_audit("—","أدمن","تفعيل تصريح وقت يدوي",
                                  f"رقم:{permit_id}|نوع:{permit_type}|من:{permit_date_from}|إلى:{permit_date_to}|وقت:{t_open or 'يوم كامل'}")
                        scope_lbl = "كل الموظفات" if permit_id=="الكل" else permit_id
                        time_lbl  = "يوم كامل" if all_day else f"{t_open}–{t_close}"
                        st.success(f"✅ تم التفعيل — {scope_lbl} — {permit_type} — {permit_date_from} إلى {permit_date_to} — {time_lbl}")
                        st.rerun()

            # ── التصاريح النشطة ──
            st.markdown("---")
            st.markdown("##### 📋 التصاريح الحالية")
            permits = get_time_permits()
            today_s = now_bh().strftime("%Y-%m-%d")
            for i, p in enumerate(permits):
                is_active = str(p.get("نشط","")).strip() in ["نعم","yes","1","TRUE","true"]
                if not is_active: continue
                p_id    = str(p.get("الرقم الشخصي","")).strip() or "الكل"
                p_type  = str(p.get("نوع التصريح","")).strip()
                d_from  = str(p.get("تاريخ البداية","")).strip()
                d_to    = str(p.get("تاريخ النهاية","")).strip()
                t_open  = str(p.get("وقت الفتح","")).strip()
                t_close = str(p.get("وقت الإغلاق","")).strip()
                p_note  = str(p.get("ملاحظات","")).strip()
                expired = d_to and d_to < today_s
                color   = "#F8D7DA" if expired else "#D4EDDA"
                status  = "منتهي ❌" if expired else "نشط ✅"
                time_w  = f"{t_open}–{t_close}" if t_open else "يوم كامل"
                st.markdown(f"""
                <div style="background:{color};border-radius:10px;padding:10px 14px;margin-bottom:8px;direction:rtl;">
                <b>{'كل الموظفات' if p_id in ['الكل','','*'] else f'#{p_id}'}</b> —
                {p_type} — {d_from} إلى {d_to} — {time_w} — {status}
                {f'<br><small>ملاحظة: {p_note}</small>' if p_note else ''}
                </div>
                """, unsafe_allow_html=True)
                if st.button("🚫 إلغاء", key=f"revoke_{i+2}"):
                    revoke_time_permit_row(i+2)
                    log_audit("—","أدمن","إلغاء تصريح وقت يدوي",f"رقم:{p_id}|نوع:{p_type}")
                    st.success("✅ تم إلغاء التصريح.")
                    st.rerun()

        # ── إعدادات الدوام المخصص ──────────────────────────────
        elif admin_tab=="📅 إعدادات الدوام المخصص":
            st.markdown("#### 📅 إعدادات الدوام المخصص")
            st.info("هنا تحددين قواعد الدوام المرنة. الأولوية في الحساب: شخص/أرقام محددة ← مهمة/قسم ← مدرسة ← الكل. إذا للمدرسة دوام عادي لكن موظفة عندها رعاية، قاعدة الموظفة تتغلب على قاعدة المدرسة.")

            with st.container(border=True):
                st.markdown("##### ➕ إضافة قاعدة دوام")

                scope_type = st.selectbox("تطبيق القاعدة على", [
                    "مدرسة", "مدارس (أكثر من واحدة)",
                    "مهمة", "مهام (أكثر من واحدة)",
                    "رقم شخصي", "أرقام (أكثر من واحد)",
                    "اسم", "أسماء (أكثر من واحد)",
                    "الكل"
                ], key="cs_scope_type")

                # قيمة النطاق
                if scope_type == "مدرسة":
                    cs_val = st.selectbox("المدرسة", schools, key="cs_val_school")
                    scope_key = "مدرسة"
                elif scope_type == "مدارس (أكثر من واحدة)":
                    cs_vals = st.multiselect("المدارس", schools, key="cs_val_schools")
                    cs_val  = "،".join(cs_vals)
                    scope_key = "مدارس"
                elif scope_type == "مهمة":
                    cs_val = st.selectbox("المهمة", TASKS_ALL, key="cs_val_task")
                    scope_key = "مهمة"
                elif scope_type == "مهام (أكثر من واحدة)":
                    cs_vals = st.multiselect("المهام", TASKS_ALL, key="cs_val_tasks")
                    cs_val  = "،".join(cs_vals)
                    scope_key = "مهام"
                elif scope_type == "رقم شخصي":
                    cs_raw = st.text_input("الرقم الشخصي", key="cs_val_id")
                    cs_val = ar_to_en_digits(cs_raw).strip()
                    scope_key = "رقم شخصي"
                    if cs_val:
                        found = validate_employee(cs_val)
                        if found:
                            st.success(f"✅ {found.get('الاسم','')} — {found.get('المدرسة','')} — {found.get('المهمة','')}")
                        else:
                            st.warning("⚠️ الرقم غير موجود في القائمة البيضاء.")
                elif scope_type == "أرقام (أكثر من واحد)":
                    cs_raw = st.text_area("الأرقام الشخصية (كل رقم في سطر)", key="cs_val_ids")
                    cs_val = "،".join([ar_to_en_digits(x).strip() for x in cs_raw.splitlines() if x.strip()])
                    scope_key = "أرقام"
                elif scope_type == "اسم":
                    cs_val = st.text_input("الاسم", key="cs_val_name")
                    scope_key = "اسم"
                elif scope_type == "أسماء (أكثر من واحد)":
                    cs_raw = st.text_area("الأسماء (كل اسم في سطر)", key="cs_val_names")
                    cs_val = "،".join([x.strip() for x in cs_raw.splitlines() if x.strip()])
                    scope_key = "أسماء"
                else:
                    cs_val = "الكل"
                    scope_key = "الكل"
                    st.warning("⚠️ سيُطبَّق على جميع الموظفات إلا من لديهن قاعدة أعلى أولوية.")

                cs_work_type = st.selectbox("نوع الدوام", [
                    "دوام عادي",
                    "رعاية",
                    "دوام مرن",
                ], key="cs_work_type_new")

                col_time1, col_time2 = st.columns(2)
                with col_time1:
                    cs_start_time = st.time_input("وقت بداية الحساب", value=time(7,0), key="cs_start_time")
                with col_time2:
                    default_hours = 5.0 if cs_work_type == "رعاية" else 7.0
                    cs_hours = st.number_input("عدد ساعات الدوام المطلوبة", min_value=1.0, max_value=12.0, value=default_hours, step=0.5, key="cs_required_hours")

                st.caption("مثال: بداية 07:00 وعدد الساعات 7 يعني نهاية الدوام المتوقع 14:00. أي خروج بعد ذلك يُحسب إضافي.")

                col_cs1, col_cs2 = st.columns(2)
                with col_cs1:
                    cs_date_from = st.date_input("من تاريخ", value=now_bh().date(), key="cs_from_new")
                with col_cs2:
                    cs_date_to   = st.date_input("إلى تاريخ", value=now_bh().date(), key="cs_to_new")
                cs_note = st.text_input("ملاحظة (اختياري)", key="cs_note_new")

                if st.button("✅ حفظ قاعدة الدوام", use_container_width=True, type="primary", key="btn_save_cs_new"):
                    if not cs_val:
                        st.error("❌ يرجى تحديد النطاق أولاً.")
                    elif cs_date_from > cs_date_to:
                        st.error("❌ تاريخ البداية يجب أن يكون قبل النهاية.")
                    else:
                        custom_schedules_sheet.append_row([
                            now_bh().strftime("%Y-%m-%d"),
                            scope_key, cs_val, cs_work_type,
                            cs_date_from.strftime("%Y-%m-%d"),
                            cs_date_to.strftime("%Y-%m-%d"),
                            "نعم", cs_note,
                            cs_start_time.strftime("%H:%M"),
                            str(cs_hours)
                        ], value_input_option="USER_ENTERED")
                        get_custom_schedules.clear()
                        log_audit("—","أدمن","إضافة إعداد دوام مخصص",
                                  f"نطاق:{scope_key}={cs_val}|نوع:{cs_work_type}|بداية:{cs_start_time.strftime('%H:%M')}|ساعات:{cs_hours}|من:{cs_date_from}|إلى:{cs_date_to}")
                        st.success("✅ تم حفظ قاعدة الدوام. لإعادة احتساب السجلات القديمة استخدمي زر إعادة الحساب بالأسفل.")
                        st.rerun()

            # ── الإعدادات الحالية ──
            st.markdown("---")
            st.markdown("##### 📋 القواعد الحالية")
            cs_all = get_custom_schedules()
            cs_active = [s for s in cs_all if str(s.get("نشط","")).strip() in ["نعم","yes","Yes","1","TRUE","true","✅"]]
            if not cs_active:
                st.success("✅ لا توجد قواعد دوام مخصصة نشطة.")
            else:
                today_s = now_bh().strftime("%Y-%m-%d")
                priority_label = {
                    1: "أولوية 1 — شخصي",
                    2: "أولوية 2 — مهمة/قسم",
                    3: "أولوية 3 — مدرسة",
                    4: "أولوية 4 — عام",
                    9: "أولوية غير محددة",
                }
                for i, s in enumerate(cs_active):
                    d_from = str(s.get("تاريخ البداية","")).strip()
                    d_to   = str(s.get("تاريخ النهاية","")).strip()
                    expired = d_to and d_to < today_s
                    color  = "#F8D7DA" if expired else "#D4EDDA"
                    status = "منتهي ❌" if expired else "نشط ✅"
                    scope_t = str(s.get("نوع النطاق","")).strip()
                    pri = _custom_scope_priority(scope_t)
                    start_txt = str(s.get("وقت البداية","") or "07:00").strip()
                    hours_txt = str(s.get("عدد الساعات","") or ("5" if str(s.get("نوع الدوام","")).strip()=="رعاية" else "7")).strip()
                    st.markdown(f"""
                    <div style="background:{color};border-radius:10px;padding:10px 14px;margin-bottom:8px;direction:rtl;">
                    <b>{priority_label.get(pri, '')}</b><br>
                    <b>{s.get('نوع النطاق','')}</b>: {s.get('قيمة النطاق','')} —
                    <b>{s.get('نوع الدوام','')}</b> — بداية: <b>{start_txt}</b> — ساعات مطلوبة: <b>{hours_txt}</b><br>
                    {d_from} إلى {d_to} — {status}
                    {f'<br><small>{s.get("ملاحظات","")}</small>' if s.get("ملاحظات") else ''}
                    </div>
                    """, unsafe_allow_html=True)
                    row_num = cs_all.index(s) + 2
                    col_disable, col_enable = st.columns(2)
                    with col_disable:
                        if st.button("🚫 تعطيل القاعدة", key=f"cs_revoke_{row_num}", use_container_width=True):
                            custom_schedules_sheet.update_cell(row_num, 7, "لا")
                            get_custom_schedules.clear()
                            log_audit("—","أدمن","تعطيل إعداد دوام مخصص",f"{s.get('نوع النطاق','')}:{s.get('قيمة النطاق','')}")
                            st.success("✅ تم تعطيل القاعدة.")
                            st.rerun()

            # ── إعادة حساب للمتأثرين ──
            st.markdown("---")
            with st.container(border=True):
                st.markdown("##### 🔄 إعادة حساب الساعات")
                st.caption("استخدميها بعد إضافة/تعطيل قاعدة لتحديث السجلات السابقة. السجلات الجديدة تُحسب تلقائياً.")
                recalc_scope = st.radio("نطاق إعادة الحساب", ["تاريخ محدد", "نطاق تاريخ"], horizontal=True, key="cs_recalc_scope")
                if recalc_scope == "تاريخ محدد":
                    cs_recalc_from = st.date_input("التاريخ", value=now_bh().date(), key="cs_recalc_date_new")
                    cs_recalc_to = cs_recalc_from
                else:
                    col_rf, col_rt = st.columns(2)
                    with col_rf:
                        cs_recalc_from = st.date_input("من تاريخ", value=now_bh().date(), key="cs_recalc_from_new")
                    with col_rt:
                        cs_recalc_to = st.date_input("إلى تاريخ", value=now_bh().date(), key="cs_recalc_to_new")

                if st.button("🔄 إعادة الحساب الآن", use_container_width=True, key="btn_cs_recalc_new"):
                    try:
                        if cs_recalc_from > cs_recalc_to:
                            st.error("❌ تاريخ البداية يجب أن يكون قبل النهاية.")
                        else:
                            data = get_sheet_data_fresh()
                            updated = 0
                            skipped = 0
                            start_s = cs_recalc_from.strftime("%Y-%m-%d")
                            end_s   = cs_recalc_to.strftime("%Y-%m-%d")
                            progress = st.progress(0)
                            total = len(data)
                            for i, row in enumerate(data):
                                progress.progress((i+1)/max(total,1))
                                d = str(row.get("التاريخ","")).strip().replace("/","-")
                                if not (start_s <= d <= end_s):
                                    skipped += 1
                                    continue
                                if not row.get("وقت الحضور",""):
                                    skipped += 1
                                    continue
                                if update_work_calculation(i+2, row):
                                    updated += 1
                            clear_caches()
                            log_audit("—","أدمن","إعادة حساب دوام مخصص",f"من:{start_s}|إلى:{end_s}|تم:{updated}")
                            st.success(f"✅ تم إعادة حساب {updated} سجل.")
                    except Exception as e:
                        st.error(f"❌ خطأ: {e}")


        if st.button("🚪 تسجيل خروج الأدمن",use_container_width=True):
            st.session_state.admin_logged_in=False; st.session_state.admin_last_active=None; st.rerun()


# ─── Footer ─────────────────────────────────────────────────────
st.markdown("""
<div class="footer-bar">
    <span>تصميم وبرمجة: <span class="hl">أ. عفاف حسين</span></span>
    <span>رئيسة المركز: <span class="hl">أ. خلود يعقوب بدو</span></span>
</div>
""", unsafe_allow_html=True)




